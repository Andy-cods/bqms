"""
Tool 2 - BQMS Analytics engine (column mapping by real worksheet layout).
"""

import os
import re
import unicodedata
import shutil
import tempfile
import subprocess
import json
from datetime import datetime

import pandas as pd
import openpyxl


CANONICAL_ALIASES = {
    "date": ["ngay", "date", "input", "end date"],
    "owner": ["nguoi phu trach", "phu trach", "owner"],
    "rfq_no": ["rfq no", "rfq", "so rfq", "don hang", "quotation no", "folder s name"],
    "bqms_code": ["bqms code", "bqms", "ma bqms", "code", "ma san pham"],
    "spec": ["spec", "ten hang", "mo ta", "model", "quy cach", "thong so ky thuat"],
    "maker": ["maker", "hang", "brand", "nha cung cap", "supplier", "vendor"],
    "qty_forecast": ["so luong du kien", "so luong", "qty", "quantity"],
    "in_rmb": ["gia nhap rmb", "rmb"],
    "in_vnd": ["gia nhap vnd", "vnd"],
    "quote_ama": ["gia bao cho ama", "bao gia ama"],
    "quote_v1": ["gia bao cho bqms v1", "v1"],
    "quote_v2": ["v2"],
    "quote_v3": ["v3"],
    "quote_v4": ["v4"],
    "note": ["ghi chu", "note"],
    "supplier": ["ncc", "supplier", "vendor"],
    "status": ["ket qua", "status", "win", "lose", "thang", "thua", "hien trang", "bo cuoc"],
}


def _norm_text(v) -> str:
    if v is None:
        return ""
    return str(v).replace("\xa0", " ").replace("\n", " ").strip()


def _fold(s: str) -> str:
    s = _norm_text(s).lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    # Normalize Vietnamese "đ" and mojibake variants before tokenizing.
    for ch in ("đ", "Đ", "ð", "Ð", "Ä‘", "Ä"):
        s = s.replace(ch, "d")
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _to_number(v):
    s = _norm_text(v)
    if not s:
        return None
    s = s.replace(",", "")
    m = re.search(r"[-+]?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        return float(m.group(0))
    except Exception:
        return None


def _to_datetime_safe(v):
    if v is None:
        return pd.NaT
    try:
        if pd.isna(v):
            return pd.NaT
    except Exception:
        pass

    if isinstance(v, (pd.Timestamp, datetime)):
        return pd.to_datetime(v, errors="coerce")

    if isinstance(v, (int, float)):
        try:
            fv = float(v)
            # Excel serial dates are day-based from 1899-12-30.
            if 59 <= fv <= 90000:
                return pd.to_datetime("1899-12-30") + pd.to_timedelta(fv, unit="D")
            return pd.to_datetime(v, errors="coerce", unit="ns")
        except Exception:
            return pd.NaT

    s = _norm_text(v)
    if not s:
        return pd.NaT
    s_clean = re.sub(r"^\(gmt[^\)]*\)\s*", "", s, flags=re.IGNORECASE)
    dt = pd.to_datetime(s_clean, errors="coerce", dayfirst=True)
    if pd.isna(dt):
        dt = pd.to_datetime(s_clean, errors="coerce", dayfirst=False)
    return dt


def _find_header_row(raw_df: pd.DataFrame, max_scan_rows: int = 25) -> int:
    """
    Detect header row by scoring known column labels.
    Returns 0 when no confident row is found.
    """
    if raw_df is None or raw_df.empty:
        return 0
    best_idx = 0
    best_score = -1
    sample_rows = min(max_scan_rows, len(raw_df))
    core_labels = ["ngay", "rfq", "bqms", "spec", "maker", "so luong", "gia", "v1", "v2", "v3", "v4", "ncc"]
    for i in range(sample_rows):
        row_vals = [_fold(v) for v in raw_df.iloc[i].tolist()]
        row_text = " | ".join(row_vals)
        score = 0
        for lb in core_labels:
            if lb in row_text:
                score += 1
        if score > best_score:
            best_score = score
            best_idx = i
    return best_idx if best_score >= 3 else 0


def _sheet_score(name: str) -> int:
    n = _fold(name)
    score = 0
    if "bqms" in n:
        score += 3
    if "hoi hang" in n:
        score += 3
    if "thong ke" in n:
        score += 2
    if "tong hop" in n:
        score += 1
    return score


def _detect_cols(columns: list[str]) -> dict:
    folded = {c: _fold(c) for c in columns}
    out = {}
    for key, aliases in CANONICAL_ALIASES.items():
        target = None
        for c, fc in folded.items():
            for alias in aliases:
                a = _fold(alias)
                if fc == a or fc.startswith(a) or a in fc:
                    target = c
                    break
            if target:
                break
        out[key] = target
    return out


def _make_unique_cols(cols: list[str]) -> list[str]:
    """
    Make column names stable and unique to avoid pandas blow-ups on
    duplicate labels during later column selection/indexing.
    """
    seen = {}
    out = []
    for i, c in enumerate(cols):
        base = _norm_text(c)
        if not base:
            base = f"__empty_{i+1}"
        n = seen.get(base, 0) + 1
        seen[base] = n
        out.append(base if n == 1 else f"{base}__{n}")
    return out


def _normalize_result(status_text: str, note_text: str) -> str:
    s = f"{_fold(status_text)} {_fold(note_text)}"
    if any(k in s for k in ["win", "thang", "trung", "awarded"]):
        return "win"
    if any(k in s for k in ["lose", "thua", "truot", "rot", "fail"]):
        return "lose"
    if any(k in s for k in ["pending", "cho", "dang xu ly", "dang"]) :
        return "pending"
    return "unknown"


def _extract_round(note_text: str, q1, q2, q3, q4):
    if q4 is not None:
        return "v4"
    if q3 is not None:
        return "v3"
    if q2 is not None:
        return "v2"
    if q1 is not None:
        return "v1"
    s = _fold(note_text)
    m = re.search(r"\bv\s*([1-5])\b", s)
    if m:
        return f"v{m.group(1)}"
    return ""


def _is_header_like_row(row: pd.Series) -> bool:
    text = " ".join(_fold(v) for v in row.tolist())
    keys = ["ngay", "rfq", "bqms", "spec", "maker", "gia", "v1", "v2", "v3", "v4", "ncc"]
    hit = sum(1 for k in keys if k in text)
    return hit >= 4


def _to_hex_rgb(argb: str) -> str:
    s = (argb or "").strip()
    if len(s) == 8:
        s = s[2:]
    if len(s) != 6:
        return ""
    return f"#{s.upper()}"


def _cell_fill_color(cell) -> str:
    try:
        fill = cell.fill
        if fill is None:
            return ""
        fg = getattr(fill, "fgColor", None) or getattr(fill, "start_color", None)
        if fg is None:
            return ""
        ctype = getattr(fg, "type", "")
        if ctype == "rgb":
            rgb = _to_hex_rgb(getattr(fg, "rgb", ""))
            # Ignore common "no fill" placeholders.
            if rgb in ("#000000", "#FFFFFF"):
                return ""
            return rgb
        if ctype == "indexed":
            idx = getattr(fg, "indexed", None)
            return f"indexed:{idx}" if idx is not None else ""
        if ctype == "theme":
            th = getattr(fg, "theme", None)
            tint = getattr(fg, "tint", 0)
            return f"theme:{th}:{tint}"
    except Exception:
        return ""
    return ""


def _row_color(ws, excel_row: int, probe_cols: list[int]) -> str:
    for c in probe_cols:
        try:
            color = _cell_fill_color(ws.cell(row=int(excel_row), column=int(c)))
            if color:
                return color
        except Exception:
            continue
    return ""


def _make_read_snapshot(path: str) -> str:
    ext = os.path.splitext(path)[1] or ".xlsx"
    fd, snap = tempfile.mkstemp(prefix="bqms_read_", suffix=ext)
    os.close(fd)
    try:
        shutil.copy2(path, snap)
        return snap
    except Exception:
        pass

    # Fallback: Excel COM SaveCopyAs for opened/locked workbook.
    src_ps = path.replace("'", "''")
    dst_ps = snap.replace("'", "''")
    ps = f"""
$ErrorActionPreference = 'Stop'
$src = '{src_ps}'
$dst = '{dst_ps}'
$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$wb = $excel.Workbooks.Open($src, 0, $true)
try {{
  $wb.SaveCopyAs($dst)
}} finally {{
  $wb.Close($false)
  $excel.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb) | Out-Null
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}}
"""
    try:
        rr = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", ps],
            timeout=90,
            capture_output=True,
            text=True,
        )
        if rr.returncode == 0 and os.path.isfile(snap):
            return snap
    except Exception:
        pass

    try:
        os.remove(snap)
    except Exception:
        pass
    return ""


def _close_quiet(obj) -> None:
    try:
        close_fn = getattr(obj, "close", None)
        if callable(close_fn):
            close_fn()
    except Exception:
        pass


def _load_one_workbook(path: str, include_colors: bool = False) -> tuple[pd.DataFrame, dict, str]:
    if not os.path.isfile(path):
        return pd.DataFrame(), {}, f"File not found: {path}"
    read_path = path
    temp_snap = ""
    try:
        xls = pd.ExcelFile(read_path, engine="openpyxl")
    except PermissionError:
        temp_snap = _make_read_snapshot(path)
        if not temp_snap:
            return pd.DataFrame(), {}, f"Permission denied: {path}"
        read_path = temp_snap
        xls = pd.ExcelFile(read_path, engine="openpyxl")
    wb_color = None
    if include_colors:
        try:
            wb_color = openpyxl.load_workbook(read_path, data_only=True)
        except Exception:
            wb_color = None
    frames = []

    mapped_debug = []
    sorted_sheets = sorted(xls.sheet_names, key=_sheet_score, reverse=True)
    target_sheets = [s for s in sorted_sheets if _sheet_score(s) >= 3] or sorted_sheets[:1]

    raw_total_rows = 0
    kept_rows = 0
    for sheet in target_sheets:
        try:
            raw = pd.read_excel(read_path, sheet_name=sheet, engine="openpyxl", header=None)
        except Exception:
            continue
        if raw is None or raw.empty:
            continue

        raw_total_rows += int(len(raw))
        hdr_idx = _find_header_row(raw)
        header_vals = [_norm_text(v) for v in raw.iloc[hdr_idx].tolist()]
        df = raw.iloc[hdr_idx + 1 :].copy()
        df.columns = _make_unique_cols(header_vals)
        df = df.reset_index(drop=True)
        # Use positional boolean mask instead of label-based `.loc[:, labels]`
        # to avoid expensive/non-linear reindexing when header labels are duplicated.
        keep_mask = [not str(c).startswith("__empty_") for c in df.columns]
        if any(keep_mask):
            df = df.iloc[:, keep_mask]
        if df.empty:
            continue

        col = _detect_cols(list(df.columns))
        header_pos = {}
        for i, v in enumerate(header_vals):
            k = _norm_text(v)
            if k and k not in header_pos:
                header_pos[k] = i + 1
        mapped_debug.append(
            {
                "sheet": sheet,
                "header_row": int(hdr_idx) + 1,
                "mapped": {k: v for k, v in col.items() if v},
            }
        )

        out = pd.DataFrame()
        out["date"] = df[col["date"]].map(_to_datetime_safe) if col["date"] else pd.NaT
        out["date"] = pd.to_datetime(out["date"], errors="coerce")
        out["owner"] = df[col["owner"]].map(_norm_text) if col["owner"] else ""
        out["rfq_no"] = df[col["rfq_no"]].map(_norm_text) if col["rfq_no"] else ""
        out["bqms_code"] = df[col["bqms_code"]].map(_norm_text) if col["bqms_code"] else ""
        out["spec"] = df[col["spec"]].map(_norm_text) if col["spec"] else ""
        out["maker"] = df[col["maker"]].map(_norm_text) if col["maker"] else ""
        out["qty_forecast"] = df[col["qty_forecast"]].map(_to_number) if col["qty_forecast"] else None
        out["in_rmb"] = df[col["in_rmb"]].map(_to_number) if col["in_rmb"] else None
        out["in_vnd"] = df[col["in_vnd"]].map(_to_number) if col["in_vnd"] else None
        out["quote_ama"] = df[col["quote_ama"]].map(_to_number) if col["quote_ama"] else None
        out["quote_v1"] = df[col["quote_v1"]].map(_to_number) if col["quote_v1"] else None
        out["quote_v2"] = df[col["quote_v2"]].map(_to_number) if col["quote_v2"] else None
        out["quote_v3"] = df[col["quote_v3"]].map(_to_number) if col["quote_v3"] else None
        out["quote_v4"] = df[col["quote_v4"]].map(_to_number) if col["quote_v4"] else None
        out["note"] = df[col["note"]].map(_norm_text) if col["note"] else ""
        out["supplier"] = df[col["supplier"]].map(_norm_text) if col["supplier"] else ""
        out["status_raw"] = df[col["status"]].map(_norm_text) if col["status"] else ""
        out["status"] = out["status_raw"]
        out["sheet"] = sheet
        out["source_file"] = os.path.basename(path)
        out["source_path"] = os.path.normpath(path)
        out["excel_row"] = (df.index + int(hdr_idx) + 2).astype(int)
        if include_colors and wb_color is not None and sheet in wb_color.sheetnames:
            ws_color = wb_color[sheet]
            probe_names = [col.get(k) for k in ["rfq_no", "bqms_code", "spec", "status"] if col.get(k)]
            probe_cols = [header_pos.get(_norm_text(n), 0) for n in probe_names]
            probe_cols = [c for c in probe_cols if c > 0]
            if not probe_cols:
                probe_cols = list(range(1, min(ws_color.max_column, 12) + 1))
            out["row_color"] = out["excel_row"].map(lambda r: _row_color(ws_color, int(r), probe_cols))
        else:
            out["row_color"] = ""

        # Keep data rows aggressively (for max coverage), drop only obvious empty/header-like rows.
        non_empty_signal = (
            (out["bqms_code"] != "")
            | (out["spec"] != "")
            | (out["rfq_no"] != "")
            | (out["maker"] != "")
            | (out["note"] != "")
            | out["qty_forecast"].notna()
            | out["quote_ama"].notna()
            | out["quote_v1"].notna()
            | out["quote_v2"].notna()
            | out["quote_v3"].notna()
            | out["quote_v4"].notna()
        )
        out = out[non_empty_signal]
        if not out.empty:
            out = out[~out.apply(_is_header_like_row, axis=1)]
        if out.empty:
            continue

        out["result"] = out.apply(lambda r: _normalize_result(r["status_raw"], r["note"]), axis=1)
        out["round"] = out.apply(
            lambda r: _extract_round(r["note"], r["quote_v1"], r["quote_v2"], r["quote_v3"], r["quote_v4"]),
            axis=1,
        )
        out["year_month"] = out["date"].dt.to_period("M").astype(str)
        out.loc[out["year_month"] == "NaT", "year_month"] = ""
        out["has_quote"] = out[["quote_v1", "quote_v2", "quote_v3", "quote_v4"]].notna().any(axis=1)
        kept_rows += int(len(out))
        frames.append(out)

    if not frames:
        _close_quiet(wb_color)
        _close_quiet(xls)
        if temp_snap and os.path.isfile(temp_snap):
            try:
                os.remove(temp_snap)
            except Exception:
                pass
        return pd.DataFrame(), {}, "No usable rows found in workbook."

    df_all = pd.concat(frames, ignore_index=True)
    wins = int((df_all["result"] == "win").sum())
    loses = int((df_all["result"] == "lose").sum())
    pending = int((df_all["result"] == "pending").sum())
    unknown = int((df_all["result"] == "unknown").sum())

    quote_series = pd.concat(
        [df_all["quote_v4"], df_all["quote_v3"], df_all["quote_v2"], df_all["quote_v1"]],
        axis=0,
    ).dropna()
    avg_quote = round(float(quote_series.mean()), 2) if not quote_series.empty else 0.0

    meta = {
        "rows": int(len(df_all)),
        "raw_rows": int(raw_total_rows),
        "kept_rows": int(kept_rows),
        "wins": wins,
        "loses": loses,
        "pending": pending,
        "unknown": unknown,
        "has_quote": int(df_all["has_quote"].sum()),
        "win_rate": round((wins / max(wins + loses, 1)) * 100, 1),
        "avg_price": avg_quote,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "mapped_debug": mapped_debug,
    }
    _close_quiet(wb_color)
    _close_quiet(xls)
    if temp_snap and os.path.isfile(temp_snap):
        try:
            os.remove(temp_snap)
        except Exception:
            pass
    return df_all, meta, ""


def load_bqms_analytics(path: str, extra_path: str = "", include_colors: bool = False) -> dict:
    paths = []
    for raw_p in [path, extra_path]:
        s = str(raw_p or "").strip()
        if not s:
            continue
        p = os.path.normpath(s)
        if p and p not in paths:
            paths.append(p)
    if not paths:
        return {"ok": False, "error": "Empty input path.", "df": None, "meta": {}}

    all_frames = []
    merged_meta = {
        "rows": 0,
        "raw_rows": 0,
        "kept_rows": 0,
        "wins": 0,
        "loses": 0,
        "pending": 0,
        "unknown": 0,
        "has_quote": 0,
        "win_rate": 0.0,
        "avg_price": 0.0,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "mapped_debug": [],
        "sources": [],
    }
    errs = []

    for p in paths:
        df_one, meta_one, err = _load_one_workbook(p, include_colors=include_colors)
        if err:
            errs.append(f"{os.path.basename(p)}: {err}")
            continue
        all_frames.append(df_one)
        merged_meta["raw_rows"] += int(meta_one.get("raw_rows", 0))
        merged_meta["kept_rows"] += int(meta_one.get("kept_rows", 0))
        merged_meta["wins"] += int(meta_one.get("wins", 0))
        merged_meta["loses"] += int(meta_one.get("loses", 0))
        merged_meta["pending"] += int(meta_one.get("pending", 0))
        merged_meta["unknown"] += int(meta_one.get("unknown", 0))
        merged_meta["has_quote"] += int(meta_one.get("has_quote", 0))
        merged_meta["mapped_debug"].extend(meta_one.get("mapped_debug", []))
        merged_meta["sources"].append(os.path.normpath(p))

    if not all_frames:
        msg = " | ".join(errs) if errs else "No usable rows found in workbook(s)."
        return {"ok": False, "error": msg, "df": None, "meta": {}}

    df_all = pd.concat(all_frames, ignore_index=True)
    merged_meta["rows"] = int(len(df_all))
    denom = max(merged_meta["wins"] + merged_meta["loses"], 1)
    merged_meta["win_rate"] = round((merged_meta["wins"] / denom) * 100, 1)
    quote_series = pd.concat(
        [df_all.get("quote_v4", pd.Series(dtype=float)),
         df_all.get("quote_v3", pd.Series(dtype=float)),
         df_all.get("quote_v2", pd.Series(dtype=float)),
         df_all.get("quote_v1", pd.Series(dtype=float))],
        axis=0,
    ).dropna()
    merged_meta["avg_price"] = round(float(quote_series.mean()), 2) if not quote_series.empty else 0.0
    return {"ok": True, "error": "", "df": df_all, "meta": merged_meta}


def _find_bc_header_row(ws, max_scan: int = 20) -> int:
    keys = ["don hang", "quotation no", "bqms", "ma san pham", "hinh anh", "maker", "so luong", "han bg", "end date"]
    best_r, best_sc = 1, -1
    for r in range(1, min(ws.max_row, max_scan) + 1):
        txt = " | ".join(_fold(ws.cell(row=r, column=c).value) for c in range(1, min(ws.max_column, 20) + 1))
        sc = sum(1 for k in keys if k in txt)
        if sc > best_sc:
            best_sc = sc
            best_r = r
    return best_r if best_sc >= 2 else 1


def _detect_bc_cols(header_vals: list[str]) -> dict:
    cols = {_fold(v): idx + 1 for idx, v in enumerate(header_vals)}
    out = {
        "date": 0,
        "rfq_no": 0,
        "bqms_code": 0,
        "spec": 0,
        "explain": 0,
        "loai_hang": 0,
        "maker": 0,
        "mark": 0,
        "image": 0,
        "unit": 0,
        "qty_forecast": 0,
        "deadline": 0,
        "status_raw": 0,
        "note": 0,
    }
    for k, c in cols.items():
        if ("ngay" in k and "thang" in k) or k == "ngay" or k == "input":
            out["date"] = c
        elif "don hang" in k or "rfq" in k or "quotation no" in k or "folder s name" in k:
            out["rfq_no"] = c
        elif "bqms" in k or "ma san pham" in k:
            out["bqms_code"] = c
        elif "ten hang" in k or "spec" in k or "thong so ky thuat" in k:
            out["spec"] = c
        elif "explain" in k or "mo ta" in k:
            out["explain"] = c
        elif "loai hang" in k:
            out["loai_hang"] = c
        elif "maker" in k or "nha cung cap" in k or "supplier" in k:
            out["maker"] = c
        elif "mark" in k or "phan so" in k:
            out["mark"] = c
        elif "hinh anh" in k:
            out["image"] = c
        elif ("don vi" in k and "tien te" not in k) or k == "unit":
            out["unit"] = c
        elif "so luong" in k or "quantity" in k or "qty" in k:
            out["qty_forecast"] = c
        elif "han bg" in k or "end date" in k or "deadline" in k:
            out["deadline"] = c
        elif "hien trang" in k or "bo cuoc" in k or "status" in k:
            out["status_raw"] = c
        elif "ghi chu" in k or "mien phi" in k or "note" in k:
            out["note"] = c
    return out


def _extract_ws_images_by_row(ws) -> dict[int, bytes]:
    out = {}
    imgs = getattr(ws, "_images", []) or []
    for img in imgs:
        try:
            anc = getattr(img, "anchor", None)
            if anc is None:
                continue
            if hasattr(anc, "_from"):
                row = int(anc._from.row) + 1
            else:
                row = int(getattr(anc, "row", 1))
            b = img._data() if hasattr(img, "_data") else None
            if b and row not in out:
                out[row] = b
        except Exception:
            continue
    return out


def _estimate_bc_data_end_row(
    ws,
    header_row: int,
    col_map: dict,
    default_row_limit: int = 12000,
    blank_streak_stop: int = 320,
) -> int:
    """
    Estimate practical end row for BC table to avoid scanning styled empty tails.
    Many workbooks have huge ws.max_row due formatting, not real data.
    """
    signal_cols = []
    for k in ["rfq_no", "bqms_code", "spec", "maker", "qty_forecast", "status_raw", "note"]:
        c = int(col_map.get(k, 0) or 0)
        if c > 0 and c not in signal_cols:
            signal_cols.append(c)
    if not signal_cols:
        signal_cols = [2, 3, 4, 7, 11, 13, 14]

    # Hard cap to keep load predictable on heavily formatted worksheets.
    hard_end = min(int(ws.max_row or 0), int(header_row) + int(default_row_limit))
    last_signal = int(header_row)
    blank_run = 0

    for r in range(int(header_row) + 1, hard_end + 1):
        has_signal = False
        for c in signal_cols:
            try:
                if _norm_text(ws.cell(row=r, column=c).value):
                    has_signal = True
                    break
            except Exception:
                continue
        if has_signal:
            last_signal = r
            blank_run = 0
        else:
            blank_run += 1
            # Stop after long empty tail once we already passed some data rows.
            if blank_run >= int(blank_streak_stop) and r > int(header_row) + 120:
                break

    return max(int(header_row), int(last_signal))


def load_bc_bqms_2026(path: str, include_colors: bool = True, include_images: bool = False) -> dict:
    if not os.path.isfile(path):
        return {"ok": False, "error": f"File not found: {path}", "df": None, "meta": {}}
    # COM is expensive. Use it only when image extraction is explicitly requested.
    if include_images:
        try:
            df_com, err_com = _load_bc_bqms_2026_via_com(path, include_colors=include_colors, include_images=include_images)
            if err_com == "" and df_com is not None and not df_com.empty:
                df = df_com
                meta = {
                    "rows": int(len(df)),
                    "raw_rows": int(len(df)),
                    "kept_rows": int(len(df)),
                    "wins": int((df["result"] == "win").sum()),
                    "loses": int((df["result"] == "lose").sum()),
                    "pending": int((df["result"] == "pending").sum()),
                    "unknown": int((df["result"] == "unknown").sum()),
                    "has_quote": 0,
                    "win_rate": 0.0,
                    "avg_price": 0.0,
                    "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                    "mapped_debug": [{"sheet": str(df["sheet"].iloc[0] if len(df) else ""), "header_row": 1, "mapped": "excel_com"}],
                    "sources": [os.path.normpath(path)],
                }
                return {"ok": True, "error": "", "df": df, "meta": meta}
        except Exception:
            pass

    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except PermissionError:
        snap = _make_read_snapshot(path)
        if not snap:
            return {"ok": False, "error": f"Permission denied: {path}", "df": None, "meta": {}}
        wb = openpyxl.load_workbook(snap, data_only=True)
    except Exception as e:
        return {"ok": False, "error": str(e), "df": None, "meta": {}}

    best = {"sheet": None, "header_row": 1, "col": {}, "records": []}
    default_col = {
        "date": 1,
        "rfq_no": 2,
        "bqms_code": 3,
        "spec": 4,
        "explain": 5,
        "loai_hang": 6,
        "maker": 7,
        "mark": 8,
        "image": 9,
        "unit": 10,
        "qty_forecast": 11,
        "deadline": 12,
        "status_raw": 13,
        "note": 14,
    }
    default_col_alt = {
        "date": 23,          # Input
        "rfq_no": 3,         # Quotation No.
        "bqms_code": 15,     # Mã sản phẩm
        "spec": 6,           # Thông số kỹ thuật
        "explain": 4,        # Mô tả
        "loai_hang": 4,      # Mô tả short type
        "maker": 20,         # Nhà cung cấp
        "mark": 21,          # Phần số
        "image": 16,         # Hình ảnh
        "unit": 7,           # Đơn vị
        "qty_forecast": 5,   # Số lượng dự kiến
        "deadline": 24,      # End date
        "status_raw": 17,    # Bỏ cuộc
        "note": 18,          # Miễn phí
    }
    # Prefer INPUT DATA sheet for BC files where pictures are stored as "picture in cell".
    sheet = None
    for ws in wb.worksheets:
        if _fold(ws.title) == "input data":
            sheet = ws
            break
    if sheet is None:
        for ws in wb.worksheets:
            if "input data" in _fold(ws.title):
                sheet = ws
                break
    if sheet is None:
        sheet = wb.worksheets[0]
    header_row = _find_bc_header_row(sheet)
    header_vals = [_norm_text(sheet.cell(row=header_row, column=c).value) for c in range(1, sheet.max_column + 1)]
    col = _detect_bc_cols(header_vals)
    mapped_score = sum(1 for k in ["rfq_no", "bqms_code", "spec", "maker", "qty_forecast", "image"] if col.get(k))
    if mapped_score < 3:
        folded_headers = [_fold(x) for x in header_vals]
        if ("quotation no" in folded_headers or "folder s name" in folded_headers) and (
            "ma san pham" in folded_headers or "hinh anh" in folded_headers
        ):
            col = default_col_alt.copy()
        else:
            col = default_col.copy()
    end_row = _estimate_bc_data_end_row(sheet, header_row, col)
    img_by_row = _extract_ws_images_by_row(sheet) if include_images else {}
    probe_cols = [x for x in [col.get("rfq_no"), col.get("bqms_code"), col.get("spec"), col.get("status_raw")] if x]
    if not probe_cols:
        probe_cols = list(range(1, min(sheet.max_column, 12) + 1))

    records = []
    for r in range(header_row + 1, end_row + 1):
        rfq = _norm_text(sheet.cell(row=r, column=col["rfq_no"]).value) if col["rfq_no"] else ""
        bqms = _norm_text(sheet.cell(row=r, column=col["bqms_code"]).value) if col["bqms_code"] else ""
        spec = _norm_text(sheet.cell(row=r, column=col["spec"]).value) if col["spec"] else ""
        maker = _norm_text(sheet.cell(row=r, column=col["maker"]).value) if col["maker"] else ""
        if not any([rfq, bqms, spec, maker]):
            continue
        dval = sheet.cell(row=r, column=col["date"]).value if col["date"] else None
        d = _to_datetime_safe(dval)
        status_raw = _norm_text(sheet.cell(row=r, column=col["status_raw"]).value) if col["status_raw"] else ""
        note = _norm_text(sheet.cell(row=r, column=col["note"]).value) if col["note"] else ""
        q = _to_number(sheet.cell(row=r, column=col["qty_forecast"]).value) if col["qty_forecast"] else None
        explain_val = _norm_text(sheet.cell(row=r, column=col["explain"]).value) if col["explain"] else ""
        loai_val = _norm_text(sheet.cell(row=r, column=col["loai_hang"]).value) if col["loai_hang"] else ""
        if not loai_val and explain_val:
            loai_val = explain_val
        color = _row_color(sheet, r, probe_cols) if include_colors else ""
        rec = {
            "date": d,
            "owner": "",
            "rfq_no": rfq,
            "bqms_code": bqms,
            "spec": spec,
            "explain": explain_val,
            "loai_hang": loai_val,
            "maker": maker,
            "mark": _norm_text(sheet.cell(row=r, column=col["mark"]).value) if col["mark"] else "",
            "unit": _norm_text(sheet.cell(row=r, column=col["unit"]).value) if col["unit"] else "",
            "qty_forecast": q,
            "deadline_text": _norm_text(sheet.cell(row=r, column=col["deadline"]).value) if col["deadline"] else "",
            "deadline": _norm_text(sheet.cell(row=r, column=col["deadline"]).value) if col["deadline"] else "",
            "note": note,
            "status_raw": status_raw,
            "status": status_raw,
            "result": _normalize_result(status_raw, note),
            "round": "",
            "quote_v1": None,
            "quote_v2": None,
            "quote_v3": None,
            "quote_v4": None,
            "quote_ama": None,
            "in_rmb": None,
            "in_vnd": None,
            "sheet": sheet.title,
            "row_color": color,
            "source_file": os.path.basename(path),
            "source_path": os.path.normpath(path),
            "excel_row": int(r),
            "image_bytes": img_by_row.get(int(r), b""),
            "has_image": int(int(r) in img_by_row),
        }
        records.append(rec)
    best = {"sheet": sheet, "header_row": header_row, "col": col, "records": records}

    if not best["records"]:
        return {"ok": False, "error": "No usable rows found in BC BQMS sheet.", "df": None, "meta": {}}

    df = pd.DataFrame(best["records"])
    meta = {
        "rows": int(len(df)),
        "raw_rows": int(end_row if best["sheet"] is not None else len(df)),
        "kept_rows": int(len(df)),
        "wins": int((df["result"] == "win").sum()),
        "loses": int((df["result"] == "lose").sum()),
        "pending": int((df["result"] == "pending").sum()),
        "unknown": int((df["result"] == "unknown").sum()),
        "has_quote": 0,
        "win_rate": 0.0,
        "avg_price": 0.0,
        "updated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "mapped_debug": [{"sheet": best["sheet"].title, "header_row": best["header_row"], "mapped": best["col"]}],
        "sources": [os.path.normpath(path)],
    }
    return {"ok": True, "error": "", "df": df, "meta": meta}


def _load_bc_bqms_2026_via_com(path: str, include_colors: bool = True, include_images: bool = True) -> tuple[pd.DataFrame, str]:
    if not os.path.isfile(path):
        return pd.DataFrame(), f"File not found: {path}"

    tmp_dir = tempfile.mkdtemp(prefix="bc_bqms_com_")
    out_json = os.path.join(tmp_dir, "rows.json")

    def _esc_ps(s: str) -> str:
        return str(s).replace("'", "''")

    script = f"""
$ErrorActionPreference = 'Stop'
$src = '{_esc_ps(path)}'
$outJson = '{_esc_ps(out_json)}'
$imgDir = '{_esc_ps(tmp_dir)}'
$withColor = @'
{json.dumps(bool(include_colors)).lower()}
'@ | ConvertFrom-Json
$withImage = @'
{json.dumps(bool(include_images)).lower()}
'@ | ConvertFrom-Json

function ColorToHex($v) {{
  if($null -eq $v) {{ return "" }}
  try {{
    $n = [int]$v
    if($n -le 0) {{ return "" }}
    $r = $n -band 255
    $g = ($n -shr 8) -band 255
    $b = ($n -shr 16) -band 255
    return "#{{0:X2}}{{1:X2}}{{2:X2}}" -f $r, $g, $b
  }} catch {{
    return ""
  }}
}}

$excel = $null
$wb = $null
$rows = New-Object System.Collections.ArrayList

try {{
  $excel = New-Object -ComObject Excel.Application
  $excel.Visible = $false
  $excel.DisplayAlerts = $false
  $excel.AskToUpdateLinks = $false
  $wb = $excel.Workbooks.Open($src, 0, $true)
  $ws = $wb.Worksheets.Item(1)

  $headerRow = 1
  $used = $ws.UsedRange
  $firstRow = [int]$used.Row
  $lastRow = [int]($used.Row + $used.Rows.Count - 1)
  if($firstRow -gt 1) {{ $headerRow = $firstRow }}

  # column layout fixed per BC form screenshot
  $cDate = 1; $cRFQ = 2; $cBQMS = 3; $cSpec = 4; $cExplain = 5; $cLoai = 6
  $cMaker = 7; $cMark = 8; $cUnit = 10; $cQty = 11; $cDeadline = 12; $cStatus = 13; $cNote = 14

  # map row -> exported image path
  $imgMap = @{{}}
  if($withImage -eq $true) {{
    foreach($sh in $ws.Shapes) {{
      try {{
        $r = [int]$sh.TopLeftCell.Row
        if($r -le $headerRow) {{ continue }}
        if(-not $imgMap.ContainsKey($r)) {{
          $png = Join-Path $imgDir ("img_r" + $r + "_" + [guid]::NewGuid().ToString("N") + ".png")
          $ok = $false
          try {{
            $sh.Export($png, 2)
            if(Test-Path $png) {{ $ok = $true }}
          }} catch {{}}
          if(-not $ok) {{
            try {{
              $sh.Copy() | Out-Null
              $co = $ws.ChartObjects().Add(0, 0, $sh.Width, $sh.Height)
              $co.Chart.Paste() | Out-Null
              $co.Chart.Export($png, "PNG") | Out-Null
              $co.Delete()
              if(Test-Path $png) {{ $ok = $true }}
            }} catch {{}}
          }}
          if($ok) {{ $imgMap[$r] = $png }}
        }}
      }} catch {{}}
    }}
  }}

  for($r = $headerRow + 1; $r -le $lastRow; $r++) {{
    $rfq = [string]$ws.Cells.Item($r, $cRFQ).Text
    $bqms = [string]$ws.Cells.Item($r, $cBQMS).Text
    $spec = [string]$ws.Cells.Item($r, $cSpec).Text
    $maker = [string]$ws.Cells.Item($r, $cMaker).Text
    if([string]::IsNullOrWhiteSpace($rfq) -and [string]::IsNullOrWhiteSpace($bqms) -and [string]::IsNullOrWhiteSpace($spec) -and [string]::IsNullOrWhiteSpace($maker)) {{
      continue
    }}

    $clr = ""
    if($withColor -eq $true) {{
      try {{
        $clr = ColorToHex($ws.Cells.Item($r, $cRFQ).Interior.Color)
      }} catch {{}}
    }}
    $imgPath = ""
    if($imgMap.ContainsKey($r)) {{ $imgPath = [string]$imgMap[$r] }}
    $obj = [ordered]@{{
      excel_row = $r
      date_text = [string]$ws.Cells.Item($r, $cDate).Text
      rfq_no = [string]$ws.Cells.Item($r, $cRFQ).Text
      bqms_code = [string]$ws.Cells.Item($r, $cBQMS).Text
      spec = [string]$ws.Cells.Item($r, $cSpec).Text
      explain = [string]$ws.Cells.Item($r, $cExplain).Text
      loai_hang = [string]$ws.Cells.Item($r, $cLoai).Text
      maker = [string]$ws.Cells.Item($r, $cMaker).Text
      mark = [string]$ws.Cells.Item($r, $cMark).Text
      unit = [string]$ws.Cells.Item($r, $cUnit).Text
      qty_forecast_text = [string]$ws.Cells.Item($r, $cQty).Text
      deadline_text = [string]$ws.Cells.Item($r, $cDeadline).Text
      status_raw = [string]$ws.Cells.Item($r, $cStatus).Text
      note = [string]$ws.Cells.Item($r, $cNote).Text
      row_color = $clr
      image_path = $imgPath
      has_image = [int]([string]::IsNullOrWhiteSpace($imgPath) -eq $false)
      sheet = [string]$ws.Name
    }}
    [void]$rows.Add($obj)
  }}

  $rows | ConvertTo-Json -Depth 6 | Out-File -FilePath $outJson -Encoding utf8
}} finally {{
  if($wb) {{ $wb.Close($false) }}
  if($excel) {{ $excel.Quit() }}
  if($wb) {{ [System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb) | Out-Null }}
  if($excel) {{ [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null }}
}}
"""
    try:
        rr = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            timeout=150,
            capture_output=True,
            text=True,
        )
    except Exception as e:
        return pd.DataFrame(), str(e)
    if rr.returncode != 0 or (not os.path.isfile(out_json)):
        err = (rr.stderr or rr.stdout or "Excel COM extraction failed.")[:600]
        return pd.DataFrame(), err

    try:
        with open(out_json, "r", encoding="utf-8-sig") as f:
            rows = json.load(f)
    except Exception as e:
        return pd.DataFrame(), f"Cannot parse COM output: {e}"
    if isinstance(rows, dict):
        rows = [rows]
    if not isinstance(rows, list) or not rows:
        return pd.DataFrame(), "No rows extracted from Excel COM."

    recs = []
    for x in rows:
        d = _to_datetime_safe(x.get("date_text", ""))
        q = _to_number(x.get("qty_forecast_text", ""))
        status_raw = _norm_text(x.get("status_raw", ""))
        note = _norm_text(x.get("note", ""))
        img_path = _norm_text(x.get("image_path", ""))
        img_bytes = b""
        if img_path and os.path.isfile(img_path):
            try:
                with open(img_path, "rb") as f:
                    img_bytes = f.read()
            except Exception:
                img_bytes = b""
        recs.append(
            {
                "date": d,
                "owner": "",
                "rfq_no": _norm_text(x.get("rfq_no", "")),
                "bqms_code": _norm_text(x.get("bqms_code", "")),
                "spec": _norm_text(x.get("spec", "")),
                "explain": _norm_text(x.get("explain", "")),
                "loai_hang": _norm_text(x.get("loai_hang", "")),
                "maker": _norm_text(x.get("maker", "")),
                "mark": _norm_text(x.get("mark", "")),
                "unit": _norm_text(x.get("unit", "")),
                "qty_forecast": q,
                "deadline_text": _norm_text(x.get("deadline_text", "")),
                "deadline": _norm_text(x.get("deadline_text", "")),
                "note": note,
                "status_raw": status_raw,
                "status": status_raw,
                "result": _normalize_result(status_raw, note),
                "round": "",
                "quote_v1": None,
                "quote_v2": None,
                "quote_v3": None,
                "quote_v4": None,
                "quote_ama": None,
                "in_rmb": None,
                "in_vnd": None,
                "sheet": _norm_text(x.get("sheet", "")),
                "row_color": _norm_text(x.get("row_color", "")),
                "source_file": os.path.basename(path),
                "source_path": os.path.normpath(path),
                "excel_row": int(x.get("excel_row", 0) or 0),
                "image_bytes": img_bytes,
                "has_image": int(1 if img_bytes else 0),
            }
        )
    df = pd.DataFrame(recs)
    # cleanup temp extraction folder
    try:
        shutil.rmtree(tmp_dir, ignore_errors=True)
    except Exception:
        pass
    return df, ""



def _try_openpyxl_image_fallback(path: str, row: int) -> bytes:
    """Fallback: extract embedded image from xlsx using openpyxl ws._images anchor scan.
    Works for floating/embedded pictures; does NOT work for =IMAGE() formula cells.
    Returns image bytes if found, else b"".
    """
    try:
        import openpyxl as _oxl
        wb = _oxl.load_workbook(path, data_only=True)
        for ws in wb.worksheets:
            img_map = _extract_ws_images_by_row(ws)
            if row in img_map and img_map[row]:
                return bytes(img_map[row])
        wb.close()
    except Exception:
        pass
    return b""


def _diagnose_cell_image_source(path: str, row: int, col_letter: str = "I") -> str:
    """
    Diagnose common no-image causes for BC row image capture.
    Returns a user-facing reason string or "" when nothing specific found.
    """
    try:
        addr = f"{str(col_letter).upper()}{int(row)}"
        wb_f = openpyxl.load_workbook(path, data_only=False, read_only=True)
        wb_v = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            ws_f = wb_f.worksheets[0]
            ws_v = wb_v.worksheets[0]
            cur_v = _norm_text(ws_v[addr].value)
            cur_f = ws_f[addr].value

            refs = []
            if isinstance(cur_f, str):
                for m in re.finditer(r"('([^']+)'|([A-Za-z0-9_ ]+))!\$?([A-Z]{1,3})\$?(\d+)", cur_f):
                    sheet_name = m.group(2) or m.group(3) or ""
                    cell_addr = f"{m.group(4)}{m.group(5)}"
                    if sheet_name and cell_addr:
                        refs.append((sheet_name, cell_addr))

            if cur_v.upper() == "#VALUE!":
                if refs:
                    for sh, a in refs:
                        if sh in wb_v.sheetnames:
                            rv = _norm_text(wb_v[sh][a].value)
                            if rv.upper() == "#VALUE!":
                                return f"Source {sh}!{a} = #VALUE! (khong phai anh that tai thoi diem doc)."
                            if rv == "":
                                return f"Source {sh}!{a} dang rong (chua co anh tai thoi diem doc)."
                return f"Cell {ws_f.title}!{addr} = #VALUE! (khong co anh de capture)."

            if cur_v == "" and refs:
                for sh, a in refs:
                    if sh in wb_v.sheetnames:
                        rv = _norm_text(wb_v[sh][a].value)
                        if rv == "":
                            return f"Source {sh}!{a} dang rong (chua co anh tai thoi diem doc)."
        finally:
            try:
                wb_f.close()
            except Exception:
                pass
            try:
                wb_v.close()
            except Exception:
                pass
    except Exception:
        return ""
    return ""

def capture_bc_row_image_via_com(path: str, excel_row: int, col_letter: str = "I") -> tuple[bytes, str]:
    if not os.path.isfile(path):
        return b"", f"File not found: {path}"
    try:
        r = int(excel_row)
    except Exception:
        return b"", f"Invalid row: {excel_row}"
    if r <= 0:
        return b"", f"Invalid row: {excel_row}"

    tmp_png = os.path.join(tempfile.gettempdir(), f"bc_cellimg_{r}_{os.getpid()}.png")

    def _esc_ps(s: str) -> str:
        return str(s).replace("'", "''")

    script = f"""
$ErrorActionPreference = 'Stop'
$src = '{_esc_ps(path)}'
$row = {r}
$cell = '{_esc_ps(col_letter)}' + $row
$out = '{_esc_ps(tmp_png)}'

function Export-RangePng($wsObj, $rangeExpr, $outPath) {{
  try {{
    if(Test-Path $outPath) {{ Remove-Item -LiteralPath $outPath -Force -ErrorAction SilentlyContinue }}
    $rng = $wsObj.Range($rangeExpr)
    try {{
      $txt = [string]$rng.Text
      if($txt -match '#VALUE!') {{ return $false }}
    }} catch {{}}
    $rng.CopyPicture(1, 2) | Out-Null
    $w = [Math]::Max([double]$rng.Width, 120)
    $h = [Math]::Max([double]$rng.Height, 120)
    $co = $wsObj.ChartObjects().Add(0, 0, $w, $h)
    $co.Chart.Paste() | Out-Null
    $co.Chart.Export($outPath, "PNG") | Out-Null
    $co.Delete()
    if(Test-Path $outPath) {{
      $fi = Get-Item -LiteralPath $outPath
      if($fi.Length -gt 500) {{ return $true }}
    }}
  }} catch {{}}
  return $false
}}

function Export-ShapePng($shapeObj, $outPath) {{
  try {{
    if(Test-Path $outPath) {{ Remove-Item -LiteralPath $outPath -Force -ErrorAction SilentlyContinue }}
    $ok = $false
    try {{
      $shapeObj.Export($outPath, 2)
      if(Test-Path $outPath) {{ $ok = $true }}
    }} catch {{}}
    if(-not $ok) {{
      try {{
        $shapeObj.Copy() | Out-Null
        $ws = $shapeObj.Parent
        $co = $ws.ChartObjects().Add(0, 0, [Math]::Max([double]$shapeObj.Width, 120), [Math]::Max([double]$shapeObj.Height, 120))
        $co.Chart.Paste() | Out-Null
        $co.Chart.Export($outPath, "PNG") | Out-Null
        $co.Delete()
        if(Test-Path $outPath) {{ $ok = $true }}
      }} catch {{}}
    }}
    if($ok -and (Test-Path $outPath)) {{
      $fi = Get-Item -LiteralPath $outPath
      if($fi.Length -gt 500) {{ return $true }}
    }}
  }} catch {{}}
  return $false
}}

function Export-ShapeAtRow($wsObj, $targetRow, $preferCol, $outPath) {{
  try {{
    $cands = New-Object System.Collections.ArrayList
    foreach($sh in $wsObj.Shapes) {{
      try {{
        $tl = $sh.TopLeftCell
        $rr = [int]$tl.Row
        if($rr -ne [int]$targetRow) {{ continue }}
        $cc = [int]$tl.Column
        $dist = [Math]::Abs($cc - [int]$preferCol)
        [void]$cands.Add([pscustomobject]@{{ shape = $sh; dist = $dist }})
      }} catch {{}}
    }}
    if($cands.Count -eq 0) {{ return $false }}
    $sorted = $cands | Sort-Object dist
    foreach($it in $sorted) {{
      if(Export-ShapePng $it.shape $outPath) {{ return $true }}
    }}
  }} catch {{}}
  return $false
}}

function Export-CellScreenPng($excelObj, $wsObj, $addr, $outPath) {{
  try {{
    if(Test-Path $outPath) {{ Remove-Item -LiteralPath $outPath -Force -ErrorAction SilentlyContinue }}
    $excelObj.Visible = $true
    $excelObj.ScreenUpdating = $true
    try {{ $excelObj.WindowState = -4143 }} catch {{}}
    Start-Sleep -Milliseconds 250
    $wsObj.Activate() | Out-Null
    $rng = $wsObj.Range($addr)
    $rng.Select() | Out-Null
    try {{ $excelObj.ActiveWindow.Zoom = 100 }} catch {{}}
    try {{ $excelObj.ActiveWindow.ScrollRow = [Math]::Max([int]$rng.Row - 5, 1) }} catch {{}}
    try {{ $excelObj.ActiveWindow.ScrollColumn = [Math]::Max([int]$rng.Column - 2, 1) }} catch {{}}
    Start-Sleep -Milliseconds 300

    $win = $excelObj.ActiveWindow
    if(-not $win) {{ return $false }}
    $leftPx = [int]$win.PointsToScreenPixelsX([double]$rng.Left)
    $topPx = [int]$win.PointsToScreenPixelsY([double]$rng.Top)
    $rightPx = [int]$win.PointsToScreenPixelsX([double]($rng.Left + $rng.Width))
    $botPx = [int]$win.PointsToScreenPixelsY([double]($rng.Top + $rng.Height))
    $w = [Math]::Max($rightPx - $leftPx, 40)
    $h = [Math]::Max($botPx - $topPx, 40)

    Add-Type -AssemblyName System.Drawing
    $bmp = New-Object System.Drawing.Bitmap($w, $h)
    $g = [System.Drawing.Graphics]::FromImage($bmp)
    $g.CopyFromScreen($leftPx, $topPx, 0, 0, [System.Drawing.Size]::new($w, $h))
    $bmp.Save($outPath, [System.Drawing.Imaging.ImageFormat]::Png)
    $g.Dispose()
    $bmp.Dispose()
    if(Test-Path $outPath) {{
      $fi = Get-Item -LiteralPath $outPath
      if($fi.Length -gt 500) {{ return $true }}
    }}
  }} catch {{}}
  return $false
}}

# Handle =IMAGE("url") formula -- Excel 365 In-Cell pictures (CopyPicture fails on these)
function Try-ImageFormula($wsObj, $addr, $outPath) {{
  try {{
    $f = [string]$wsObj.Range($addr).Formula
    $imgM = [regex]::Match($f, '=IMAGE\s*\(\s*"([^"]+)"')
    if($imgM.Success) {{
      $url = $imgM.Groups[1].Value.Trim()
      if(Test-Path $outPath) {{ Remove-Item -LiteralPath $outPath -Force -ErrorAction SilentlyContinue }}
      if($url -match '^https?://') {{
        try {{
          $wc = New-Object Net.WebClient
          $wc.DownloadFile($url, $outPath)
          $wc.Dispose()
        }} catch {{ }}
      }} elseif(Test-Path $url) {{
        Copy-Item -LiteralPath $url -Destination $outPath -Force -ErrorAction SilentlyContinue
      }}
      if(Test-Path $outPath) {{
        $fi = Get-Item -LiteralPath $outPath
        if($fi.Length -gt 500) {{ return $true }}
      }}
    }}
  }} catch {{}}
  return $false
}}

function Get-FormulaRefs($formulaText) {{
  $outRefs = New-Object System.Collections.ArrayList
  if([string]::IsNullOrWhiteSpace($formulaText)) {{ return ,$outRefs }}
  try {{
    $rx = [regex]"('([^']+)'|([A-Za-z0-9_ ]+))!\$?([A-Z]{{1,3}})\$?(\d+)"
    $ms = $rx.Matches($formulaText)
    foreach($m in $ms) {{
      $sheetName = ""
      if(-not [string]::IsNullOrWhiteSpace($m.Groups[2].Value)) {{
        $sheetName = $m.Groups[2].Value
      }} else {{
        $sheetName = $m.Groups[3].Value
      }}
      $addr = ($m.Groups[4].Value + $m.Groups[5].Value)
      if(-not [string]::IsNullOrWhiteSpace($sheetName) -and -not [string]::IsNullOrWhiteSpace($addr)) {{
        [void]$outRefs.Add([pscustomobject]@{{ sheet = $sheetName; addr = $addr }})
      }}
    }}
  }} catch {{}}
  return ,$outRefs
}}

$excel = New-Object -ComObject Excel.Application
$excel.Visible = $false
$excel.DisplayAlerts = $false
$wb = $excel.Workbooks.Open($src, 0, $true)
try {{
  $ws = $null
  try {{
    $ws = $wb.Worksheets.Item("INPUT DATA")
  }} catch {{}}
  if($null -eq $ws) {{
    try {{
      foreach($cand in $wb.Worksheets) {{
        if(([string]$cand.Name).ToLower().Contains("input data")) {{
          $ws = $cand
          break
        }}
      }}
    }} catch {{}}
  }}
  if($null -eq $ws) {{
    $ws = $wb.Worksheets.Item(1)
  }}
  $ok = Export-RangePng $ws $cell $out

  # Fallback A: direct cell has =IMAGE() formula
  if(-not $ok) {{
    $ok = Try-ImageFormula $ws $cell $out
  }}

  # Fallback B: follow all cross-sheet references in formula.
  if(-not $ok) {{
    try {{
      $f = [string]$ws.Range($cell).Formula
      $refs = Get-FormulaRefs $f
      foreach($rf in $refs) {{
        try {{
          $ws2 = $wb.Worksheets.Item([string]$rf.sheet)
          $addr = [string]$rf.addr
          $ok = Export-RangePng $ws2 $addr $out
          if(-not $ok) {{
            $ok = Try-ImageFormula $ws2 $addr $out
          }}
          if(-not $ok) {{
            $mRow = [regex]::Match($addr, "(\d+)$")
            if($mRow.Success) {{
              $refRow = [int]$mRow.Groups[1].Value
              $ok = Export-ShapeAtRow $ws2 $refRow 16 $out
            }}
          }}
          if($ok) {{ break }}
        }} catch {{}}
      }}
    }} catch {{}}
  }}

  # Fallback C: image is a floating shape on row in sheet 1.
  if(-not $ok) {{
    $ok = Export-ShapeAtRow $ws $row 9 $out
  }}

  # Fallback D: explicit INPUT DATA row match (common for =INPUT DATA!Pxx patterns).
  if(-not $ok) {{
    try {{
      $wsInput = $wb.Worksheets.Item("INPUT DATA")
      $ok = Export-ShapeAtRow $wsInput $row 16 $out
    }} catch {{}}
  }}

  # Fallback E: screen snapshot of displayed cell area.
  if(-not $ok) {{
    try {{
      $ok = Export-CellScreenPng $excel $ws $cell $out
    }} catch {{}}
  }}

  # Fallback F: for cross-sheet refs, snapshot source cell area.
  if(-not $ok) {{
    try {{
      $f = [string]$ws.Range($cell).Formula
      $refs = Get-FormulaRefs $f
      foreach($rf in $refs) {{
        try {{
          $ws2 = $wb.Worksheets.Item([string]$rf.sheet)
          $addr = [string]$rf.addr
          $ok = Export-CellScreenPng $excel $ws2 $addr $out
          if($ok) {{ break }}
        }} catch {{}}
      }}
    }} catch {{}}
  }}
}} finally {{
  $wb.Close($false)
  $excel.Quit()
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb) | Out-Null
  [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null
}}
"""
    try:
        rr = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            timeout=90,
            capture_output=True,
            text=True,
        )
    except Exception as e:
        return b"", str(e)
    if rr.returncode != 0:
        err_msg = (rr.stderr or rr.stdout or "Capture image via COM failed.")[:400]
        if "80070520" in err_msg:
            return b"", "Excel COM khong khoi tao duoc trong session hien tai (80070520). Hay chay app trong desktop session da dang nhap."
        # COM failed entirely -- try openpyxl ws._images fallback
        b = _try_openpyxl_image_fallback(path, r)
        if b:
            return b, ""
        diag = _diagnose_cell_image_source(path, r, col_letter=col_letter)
        if diag:
            return b"", diag
        return b"", err_msg
    if not os.path.isfile(tmp_png):
        # No output file -- try openpyxl fallback
        b = _try_openpyxl_image_fallback(path, r)
        if b:
            return b, ""
        diag = _diagnose_cell_image_source(path, r, col_letter=col_letter)
        if diag:
            return b"", diag
        return b"", "No image output file."
    try:
        with open(tmp_png, "rb") as f:
            b = f.read()
        if len(b) < 500:
            # Too small -- try openpyxl fallback before giving up
            fb = _try_openpyxl_image_fallback(path, r)
            if fb:
                return fb, ""
            diag = _diagnose_cell_image_source(path, r, col_letter=col_letter)
            if diag:
                return b"", diag
            return b"", "Captured image is empty/small."
        return b, ""
    except Exception as e:
        return b"", str(e)
    finally:
        try:
            os.remove(tmp_png)
        except Exception:
            pass
