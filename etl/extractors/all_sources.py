"""Extract ALL data sources from OneDrive for complete migration.

Sources:
1. Thong ke hoi hang BQMS.xlsx → quotations + quotation_details
2. Thong ke giao hang 2025/2026.xlsx → po_deliveries
3. Thong ke dat hang.xlsx → order_tracking
4. BC BQMS monthly files → monthly_input_data
5. DANH BA (from delivery files) → contacts
"""

import hashlib
import os
from datetime import datetime
from typing import Optional

import openpyxl


def _s(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ").strip()
    if s in ("#VALUE!", "#REF!", "#N/A", "#DIV/0!", "#NAME?"):
        return ""
    return s


def _n(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("\xa0", "").strip()
    if not s or s.startswith("#"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _d(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s or len(s) < 6 or s.startswith("#"):
        return None
    return s[:10]


def _cell(row, idx):
    return row[idx] if idx < len(row) else None


def _hash(*parts) -> str:
    key = "|".join(str(p or "") for p in parts)
    return hashlib.sha256(key.encode("utf-8")).hexdigest()[:32]


# ─── 1. Thong ke hoi hang BQMS ─────────────────────────

def extract_quotations(filepath: str) -> tuple[list[dict], list[dict]]:
    """Extract from TONG HOP BQMS + TRUNG BG sheets."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    fname = os.path.basename(filepath)
    quotations = []
    details = []

    # Sheet: TONG HOP BQMS
    for sname in wb.sheetnames:
        if "TONG HOP" in sname.upper() or "TỔNG HỢP" in sname.upper():
            ws = wb[sname]
            seen = set()
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if not row or len(row) < 6:
                    continue
                rfq = _s(_cell(row, 2))
                bqms = _s(_cell(row, 3))
                spec = _s(_cell(row, 4))
                if not rfq and not bqms and not spec:
                    continue
                h = _hash(rfq, bqms, spec, _d(_cell(row, 0)))
                if h in seen:
                    continue
                seen.add(h)

                quotations.append({
                    "date": _d(_cell(row, 0)),
                    "person_in_charge": _s(_cell(row, 1)),
                    "rfq_no": rfq[:100] if rfq else None,
                    "bqms_code": bqms[:100] if bqms else None,
                    "spec": spec[:500] if spec else None,
                    "maker": _s(_cell(row, 5))[:200] or None,
                    "quantity_forecast": _n(_cell(row, 6)),
                    "price_rmb": _n(_cell(row, 7)),
                    "price_vnd": _n(_cell(row, 8)),
                    "price_ama": _n(_cell(row, 9)),
                    "price_bqms_v1": _n(_cell(row, 10)),
                    "price_bqms_v2": _n(_cell(row, 11)),
                    "price_bqms_v3": _n(_cell(row, 12)),
                    "price_bqms_v4": _n(_cell(row, 13)) if len(row) > 13 else None,
                    "notes": _s(_cell(row, 14))[:500] or None,
                    "supplier": _s(_cell(row, 15))[:200] or None,
                    "result": _s(_cell(row, 16))[:50] or None,
                    "report": _s(_cell(row, 17))[:200] if len(row) > 17 else None,
                    "source_file": fname,
                    "source_sheet": sname,
                    "source_row": i + 2,
                    "row_hash": h,
                })
            break

    # Sheet: TRUNG BG
    for sname in wb.sheetnames:
        if "TRUNG" in sname.upper():
            ws = wb[sname]
            seen = set()
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if not row or len(row) < 5:
                    continue
                rfq = _s(_cell(row, 1))
                bqms = _s(_cell(row, 2))
                if not rfq and not bqms:
                    continue
                h = _hash("TRUNG", rfq, bqms, _s(_cell(row, 4)))
                if h in seen:
                    continue
                seen.add(h)

                details.append({
                    "person_in_charge": _s(_cell(row, 0))[:100] or None,
                    "rfq_no": rfq[:100] if rfq else None,
                    "bqms_code": bqms[:100] if bqms else None,
                    "description": _s(_cell(row, 3))[:300] or None,
                    "spec": _s(_cell(row, 4))[:500] or None,
                    "quantity_forecast": _n(_cell(row, 5)),
                    "unit": _s(_cell(row, 6))[:20] or None,
                    "price_po": _n(_cell(row, 7)),
                    "po_deadline": _s(_cell(row, 8))[:100] or None,
                    "notes": _s(_cell(row, 9))[:500] or None,
                    "supplier": _s(_cell(row, 10))[:200] or None,
                    "hs_code": _s(_cell(row, 11))[:20] or None,
                    "desc_import": _s(_cell(row, 12))[:500] or None,
                    "source_file": fname,
                    "source_row": i + 2,
                    "row_hash": h,
                })
            break

    wb.close()
    return quotations, details


# ─── 2. Thong ke giao hang ──────────────────────────────

def extract_deliveries(filepath: str) -> tuple[list[dict], list[dict]]:
    """Extract PO deliveries + contacts from delivery tracking file."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    fname = os.path.basename(filepath)
    deliveries = []
    contacts = []

    # Main sheet (THONG KE PO)
    for sname in wb.sheetnames:
        if "PO" in sname.upper() or "THỐNG KÊ" in sname.upper():
            ws = wb[sname]
            seen = set()
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                if not row or len(row) < 5:
                    continue
                po = _s(_cell(row, 1))
                bqms = _s(_cell(row, 4))
                if not po and not bqms:
                    continue
                h = _hash(po, bqms, _s(_cell(row, 5)), _d(_cell(row, 0)))
                if h in seen:
                    continue
                seen.add(h)

                deliveries.append({
                    "po_date": _d(_cell(row, 0)),
                    "po_number": po[:50] or None,
                    "shipping_no": _s(_cell(row, 2))[:50] or None,
                    "rfq_no": _s(_cell(row, 3))[:100] or None,
                    "bqms_code": bqms[:100] if bqms else None,
                    "spec": _s(_cell(row, 5))[:500] or None,
                    "quantity": _n(_cell(row, 6)),
                    "unit": _s(_cell(row, 7))[:20] or None,
                    "unit_price": _n(_cell(row, 8)),
                    "total_amount": _n(_cell(row, 9)),
                    "buyer": _s(_cell(row, 10))[:50] or None,
                    "mail_pur": _s(_cell(row, 11))[:100] or None,
                    "receiver_name": _s(_cell(row, 12))[:100] or None,
                    "warehouse": _s(_cell(row, 13))[:100] or None,
                    "pur_phone": _s(_cell(row, 14))[:50] or None,
                    "status": _s(_cell(row, 15))[:100] or None,
                    "delivery_date": _d(_cell(row, 16)),
                    "actual_qty": _n(_cell(row, 17)),
                    "delivery_info": _s(_cell(row, 18))[:300] or None,
                    "delivery_method": _s(_cell(row, 19))[:100] or None,
                    "origin": _s(_cell(row, 20))[:100] or None,
                    "total_delivered_vnd": _n(_cell(row, 21)),
                    "source_file": fname,
                    "source_row": i + 2,
                    "row_hash": h,
                })
            break

    # DANH BA sheet (contacts)
    for sname in wb.sheetnames:
        if "DANH" in sname.upper():
            ws = wb[sname]
            for i, row in enumerate(ws.iter_rows(min_row=4, values_only=True)):
                if not row or len(row) < 2:
                    continue
                email = _s(_cell(row, 0))
                name = _s(_cell(row, 1))
                if not email and not name:
                    continue
                contacts.append({
                    "email": email[:100] or None,
                    "full_name": name[:100] or None,
                    "delivery_info": _s(_cell(row, 2))[:300] or None,
                    "phone": _s(_cell(row, 3))[:50] or None,
                    "source_file": fname,
                })
            break

    wb.close()
    return deliveries, contacts


# ─── 3. Thong ke dat hang ────────────────────────────────

def extract_orders(filepath: str) -> list[dict]:
    """Extract order tracking data."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    fname = os.path.basename(filepath)
    orders = []
    ws = wb[wb.sheetnames[0]]
    seen = set()

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or len(row) < 4:
            continue
        rfq = _s(_cell(row, 1))
        bqms = _s(_cell(row, 2))
        if not rfq and not bqms:
            continue
        h = _hash("ORD", rfq, bqms, _s(_cell(row, 3)))
        if h in seen:
            continue
        seen.add(h)

        orders.append({
            "order_no": int(_n(_cell(row, 0)) or 0) or None,
            "rfq_no": rfq[:100] if rfq else None,
            "bqms_code": bqms[:100] if bqms else None,
            "spec": _s(_cell(row, 3))[:500] or None,
            "customer": _s(_cell(row, 4))[:200] or None,
            "quantity_planned": _n(_cell(row, 5)),
            "quantity_ordered": _n(_cell(row, 6)),
            "unit": _s(_cell(row, 7))[:20] or None,
            "order_date": _d(_cell(row, 8)),
            "validity_period": _s(_cell(row, 9))[:100] or None,
            "status": _s(_cell(row, 10))[:100] or None,
            "quantity_delivered": _n(_cell(row, 11)),
            "delivery_date": _d(_cell(row, 12)),
            "notes": _s(_cell(row, 13))[:500] or None,
            "source_file": fname,
            "source_row": i + 2,
            "row_hash": h,
        })

    wb.close()
    return orders


# ─── 4. BC BQMS monthly (INPUT DATA sheets) ─────────────

def extract_monthly_report(filepath: str, year: int, month: int) -> list[dict]:
    """Extract INPUT DATA from BC BQMS monthly report."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    fname = os.path.basename(filepath)
    rows_out = []

    # Find INPUT DATA sheet
    target_sheet = None
    for sname in wb.sheetnames:
        if "INPUT" in sname.upper():
            target_sheet = sname
            break

    if not target_sheet:
        wb.close()
        return []

    ws = wb[target_sheet]
    seen = set()

    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
        if not row or len(row) < 6:
            continue
        quot_no = _s(_cell(row, 2))
        desc = _s(_cell(row, 3))
        spec = _s(_cell(row, 5))
        if not quot_no and not desc and not spec:
            continue

        h = _hash("BC", year, month, quot_no, desc, spec)
        if h in seen:
            continue
        seen.add(h)

        product_code = _s(_cell(row, 14))
        withdrew = _s(_cell(row, 16)).upper() in ("Y", "YES", "X", "TRUE", "1")
        free_sample = _s(_cell(row, 17)).upper() in ("Y", "YES", "X", "TRUE", "1")

        rows_out.append({
            "year": year,
            "month": month,
            "row_no": int(_n(_cell(row, 0)) or 0) or None,
            "folder_name": _s(_cell(row, 1))[:200] or None,
            "quotation_no": quot_no[:100] if quot_no else None,
            "description": desc[:300] if desc else None,
            "quantity_forecast": _n(_cell(row, 4)),
            "tech_specs": spec[:500] if spec else None,
            "unit": _s(_cell(row, 6))[:20] or None,
            "quote_price": _n(_cell(row, 7)),
            "quote_value": _n(_cell(row, 8)),
            "currency": _s(_cell(row, 9))[:10] or None,
            "moq": _n(_cell(row, 10)),
            "product_code": product_code[:100] if product_code else None,
            "image_ref": _s(_cell(row, 15))[:200] or None,
            "withdrew": withdrew,
            "free_sample": free_sample,
            "cis_code": _s(_cell(row, 18))[:50] or None,
            "supplier": _s(_cell(row, 19))[:200] or None,
            "part_number": _s(_cell(row, 20))[:100] or None,
            "leadtime": _s(_cell(row, 21))[:100] or None,
            "input_date": _d(_cell(row, 22)),
            "end_date": _d(_cell(row, 23)),
            "source_file": fname,
            "source_row": i + 2,
            "row_hash": h,
        })

    wb.close()
    return rows_out
