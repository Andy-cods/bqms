"""
tools/tool1_autofill/engine.py
Auto Fill engine â€” pure logic, zero Streamlit dependency.
Extracted from app.py (v1 monolith).
"""

import os
import re
import io
import json
import unicodedata
import copy as copy_module
import shutil
import subprocess
import tempfile
import time
from decimal import Decimal, InvalidOperation
from datetime import datetime, timedelta

import pandas as pd
import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

COL_KEYWORDS = {
    "don_hang":  ["Ä‘Æ¡n hÃ ng", "don hang", "rfq"],
    "bqms":      ["bqms"],
    "spec":      ["tÃªn hÃ ng", "ten hang", "spec"],
    "short_name":["explain", "tÃªn ngáº¯n", "ten ngan"],
    "loai_hang": ["loáº¡i hÃ ng", "loai hang"],
    "maker":     ["maker"],
    "mark":      ["mark"],
    "hinh_anh":  ["hÃ¬nh áº£nh", "hinh anh", "image"],
    "don_vi":    ["Ä‘Æ¡n vá»‹", "don vi", "unit"],
    "so_luong":  ["sá»‘ lÆ°á»£ng", "so luong", "qty", "quantity"],
    "han_bg":    ["háº¡n bg", "han bg", "deadline", "háº¡n"],
    "hien_trang":["hiá»‡n tráº¡ng", "hien trang"],
    "ghi_chu":   ["ghi chÃº", "ghi chu", "note"],
}

# ---------------------------------------------------------------------------
# Logger â€” receives a callback so UI can intercept, or falls back to print
# ---------------------------------------------------------------------------

_log_callback = None


def set_log_callback(fn):
    """Register a callback: fn(msg, level) â€” level: 'info'|'ok'|'warn'|'error'."""
    global _log_callback
    _log_callback = fn


def _log(msg: str, level: str = "info") -> None:
    if _log_callback:
        _log_callback(msg, level)


# ---------------------------------------------------------------------------
# String utilities
# ---------------------------------------------------------------------------

def clean_cell(val) -> str:
    """Strip non-breaking spaces and whitespace from a cell value."""
    if val is None:
        return ""
    return str(val).replace("\xa0", " ").strip()


def sanitize_filename(name: str) -> str:
    """Remove characters invalid in Windows filenames."""
    return re.sub(r'[\\/*?:"<>|]', "_", str(name))


def fuzzy_match_col(header_str: str, keywords: list) -> bool:
    """Return True if any keyword appears in header_str (case-insensitive)."""
    h = header_str.lower().replace("\xa0", " ")
    return any(kw.lower() in h for kw in keywords)


def classify_loai_hang(val: str) -> str:
    """Classify 'Loáº¡i hÃ ng' value to 'gc', 'tm', or 'unknown'."""
    if not val:
        return "unknown"
    lh = val.lower().replace("\xa0", " ").strip()
    if any(x in lh for x in ("gia cÃ´ng", "gia cong", "gc")):
        return "gc"
    if any(x in lh for x in ("thÆ°Æ¡ng máº¡i", "thuong mai", "tm")):
        return "tm"
    abbr = lh[:2]
    if abbr == "gc":
        return "gc"
    if abbr == "tm":
        return "tm"
    return "unknown"


# ---------------------------------------------------------------------------
# BC BQMS parser
# ---------------------------------------------------------------------------

def _find_bc_sheet_name(file_bytes: bytes) -> tuple:
    """
    Scan sheets to find BC BQMS data sheet.
    Returns (sheet_name, header_row_0based).
    """
    xl = pd.ExcelFile(io.BytesIO(file_bytes), engine="openpyxl")
    _log(f"File cÃ³ sheet(s): {', '.join(xl.sheet_names)}")

    best_sheet, best_header, best_score = None, 0, 0

    for name in xl.sheet_names:
        try:
            preview = xl.parse(name, header=None, nrows=15, dtype=str)
        except Exception:
            continue
        for i, row in preview.iterrows():
            row_text = " ".join(str(v) for v in row if pd.notna(v)).lower()
            score = (
                (2 if "Ä‘Æ¡n hÃ ng" in row_text else 0) +
                (2 if "bqms" in row_text else 0) +
                (1 if "maker" in row_text else 0) +
                (1 if "loáº¡i hÃ ng" in row_text else 0)
            )
            if score > best_score:
                best_score, best_sheet, best_header = score, name, i
                break

    if best_sheet is None:
        best_sheet = xl.sheet_names[0]
        _log("KhÃ´ng tÃ¬m tháº¥y header chuáº©n, dÃ¹ng sheet Ä‘áº§u tiÃªn", "warn")
    else:
        _log(f"DÃ¹ng sheet '{best_sheet}', header row {best_header + 1}")

    return best_sheet, best_header


def parse_deadline(deadline_str: str):
    """
    Parse Vietnamese deadline like '2/4 17h' â†’ datetime.
    Returns None if parsing fails.
    """
    if not deadline_str:
        return None
    try:
        s = deadline_str.strip()
        now = datetime.now()
        m = re.match(r"(\d{1,2})/(\d{1,2})\s*(\d{1,2})h?", s)
        if m:
            day, month, hour = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return datetime(now.year, month, day, hour, 0)
        m = re.match(r"(\d{1,2})/(\d{1,2})", s)
        if m:
            day, month = int(m.group(1)), int(m.group(2))
            return datetime(now.year, month, day, 17, 0)
    except Exception:
        pass
    return None


def is_deadline_urgent(deadline_dt) -> bool:
    """Return True if deadline is within 24 hours from now."""
    if deadline_dt is None:
        return False
    return (deadline_dt - datetime.now()) < timedelta(hours=24)


def parse_bc_bqms(file_input) -> list:
    """
    Parse BC BQMS Excel file.
    file_input: file path (str) or bytes.
    Returns list of order dicts.
    """
    if isinstance(file_input, (str, os.PathLike)):
        with open(file_input, "rb") as f:
            file_bytes = f.read()
    else:
        file_bytes = file_input

    try:
        sheet_name, header_idx = _find_bc_sheet_name(file_bytes)
        df = pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name=sheet_name,
            header=header_idx,
            dtype=str,
            engine="openpyxl",
        )
    except Exception as e:
        _log(f"KhÃ´ng thá»ƒ Ä‘á»c file BC BQMS: {e}", "error")
        return []

    df.columns = [clean_cell(c) for c in df.columns]

    col_map = {}
    for col in df.columns:
        for field, keywords in COL_KEYWORDS.items():
            if field not in col_map and fuzzy_match_col(col, keywords):
                col_map[field] = col

    _log(f"Cá»™t map: {col_map}")

    def get_field(row, field):
        col = col_map.get(field)
        if col is None or col not in row.index:
            return ""
        v = row[col]
        return clean_cell(v) if pd.notna(v) else ""

    orders = []
    for idx, row in df.iterrows():
        don_hang = get_field(row, "don_hang")
        bqms     = get_field(row, "bqms")
        if not don_hang and not bqms:
            continue
        if don_hang.lower() in ("Ä‘Æ¡n hÃ ng", "don hang", "nan"):
            continue

        loai_hang_raw = get_field(row, "loai_hang")
        loai_norm     = classify_loai_hang(loai_hang_raw).upper()  # 'GC'|'TM'|'UNKNOWN'

        so_luong_raw = get_field(row, "so_luong")
        try:
            so_luong = int(float(so_luong_raw)) if so_luong_raw and so_luong_raw not in ("0", "") else ""
        except Exception:
            so_luong = so_luong_raw

        han_bg      = get_field(row, "han_bg")
        deadline_dt = parse_deadline(han_bg)

        orders.append({
            "id":          f"{don_hang}_{bqms}",
            "don_hang":    don_hang,
            "bqms":        bqms,
            "spec":        get_field(row, "spec"),
            "short_name":  get_field(row, "short_name"),
            "loai_hang":   loai_norm,
            "maker":       get_field(row, "maker"),
            "mark":        get_field(row, "mark"),
            "don_vi":      get_field(row, "don_vi") or "EA",
            "so_luong":    so_luong,
            "han_bg":      han_bg,
            "deadline_dt": deadline_dt,
            "is_urgent":   is_deadline_urgent(deadline_dt),
            "ghi_chu":     get_field(row, "ghi_chu"),
        })

    _log(f"Äá»c Ä‘Æ°á»£c {len(orders)} dÃ²ng tá»« BC BQMS.", "ok")
    return orders


# ---------------------------------------------------------------------------
# Image utilities
# ---------------------------------------------------------------------------

def match_images_to_orders(uploaded_files: list, orders: list) -> dict:
    """
    Auto-match image files to BQMS codes by filename.
    uploaded_files: list of objects with .name and .read() (Streamlit UploadedFile)
                   OR list of file paths (str).
    Returns dict: bqms_code -> image bytes.
    """
    images_map = {}
    bqms_codes = [o["bqms"] for o in orders if o.get("bqms")]

    for uf in uploaded_files:
        if isinstance(uf, (str, os.PathLike)):
            fname = os.path.basename(str(uf))
            with open(uf, "rb") as f:
                img_bytes = f.read()
        else:
            fname     = uf.name
            img_bytes = uf.read()

        fname_clean = os.path.splitext(fname)[0].strip()
        matched = None
        for code in bqms_codes:
            if code and (code in fname_clean or fname_clean in code):
                matched = code
                break

        if matched:
            images_map[matched] = img_bytes
            _log(f"áº¢nh '{fname}' ghÃ©p vá»›i BQMS {matched}", "ok")
        else:
            _log(f"áº¢nh '{fname}' khÃ´ng khá»›p vá»›i BQMS nÃ o (bá» qua)", "warn")

    return images_map


def pil_image_to_openpyxl_image(img_bytes: bytes, max_w: int = 80, max_h: int = 80):
    """Convert image bytes â†’ openpyxl Image, resized."""
    try:
        pil_img = PILImage.open(io.BytesIO(img_bytes))
        if pil_img.mode not in ("RGB", "RGBA"):
            pil_img = pil_img.convert("RGBA")
        pil_img.thumbnail((max_w, max_h), PILImage.LANCZOS)
        w, h = pil_img.size
        buf  = io.BytesIO()
        pil_img.save(buf, format="PNG")
        buf.seek(0)
        xl_img        = XLImage(buf)
        xl_img.width  = w
        xl_img.height = h
        return xl_img
    except Exception as e:
        _log(f"Lá»—i chuyá»ƒn áº£nh: {e}", "warn")
        return None


# ---------------------------------------------------------------------------
# openpyxl cell/row helpers
# ---------------------------------------------------------------------------

def get_top_left_of_merged(ws, row: int, col: int):
    for merged in ws.merged_cells.ranges:
        if (merged.min_row <= row <= merged.max_row and
                merged.min_col <= col <= merged.max_col):
            return ws.cell(row=merged.min_row, column=merged.min_col)
    return ws.cell(row=row, column=col)


def safe_set_cell(ws, row: int, col: int, value) -> None:
    get_top_left_of_merged(ws, row, col).value = value


def copy_row_style(ws, src_row: int, dst_row: int, max_col: int = None) -> None:
    if max_col is None:
        max_col = ws.max_column
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, max_col + 1):
        src = ws.cell(row=src_row, column=col)
        dst = ws.cell(row=dst_row, column=col)
        if src.has_style:
            dst.font         = copy_module.copy(src.font)
            dst.fill         = copy_module.copy(src.fill)
            dst.alignment    = copy_module.copy(src.alignment)
            dst.border       = copy_module.copy(src.border)
            dst.number_format = src.number_format


def copy_merged_ranges_for_row(ws, src_row: int, dst_row: int) -> None:
    new_merges = [
        (m.min_col, m.max_col)
        for m in list(ws.merged_cells.ranges)
        if m.min_row == src_row == m.max_row
    ]
    for min_col, max_col in new_merges:
        ws.merge_cells(
            start_row=dst_row, start_column=min_col,
            end_row=dst_row,   end_column=max_col
        )


def insert_rows_copy_style(ws, after_row: int, count: int, template_row: int) -> int:
    ws.insert_rows(after_row + 1, amount=count)
    for i in range(count):
        new_row = after_row + 1 + i
        copy_row_style(ws, template_row, new_row)
        copy_merged_ranges_for_row(ws, template_row, new_row)
    return after_row + 1


# ---------------------------------------------------------------------------
# CAM KET filler
# ---------------------------------------------------------------------------

def fill_cam_ket(template_path: str, products: list, images_map: dict, output_path: str) -> bool:
    """
    Fill CAM KET template.
    products: list of order dicts (TM only).
    Returns True on success.
    """
    try:
        wb = openpyxl.load_workbook(template_path)
    except PermissionError:
        _log(f"CAM KET template Ä‘ang bá»‹ khÃ³a: {template_path}", "error")
        return False
    except Exception as e:
        _log(f"Lá»—i má»Ÿ CAM KET template: {e}", "error")
        return False

    ws = wb.active
    SP_START    = 16
    ROWS_PER_SP = 3

    C = column_index_from_string("C")
    D = column_index_from_string("D")
    F = column_index_from_string("F")
    J = column_index_from_string("J")
    L = column_index_from_string("L")

    def _fill_product(start_row: int, idx: int, product: dict) -> None:
        bqms  = product.get("bqms", "")
        spec  = product.get("spec", "")
        maker = product.get("maker", "")
        safe_set_cell(ws, start_row, C, idx)
        safe_set_cell(ws, start_row, D, bqms)
        safe_set_cell(ws, start_row, F, spec)
        safe_set_cell(ws, start_row, J, maker)
        safe_set_cell(ws, start_row, L, "")  # Unit Price â€” always blank
        if bqms in images_map:
            xl_img = pil_image_to_openpyxl_image(images_map[bqms])
            if xl_img:
                try:
                    ws.add_image(xl_img, f"N{start_row}")
                except Exception as e:
                    _log(f"KhÃ´ng thÃªm áº£nh N{start_row}: {e}", "warn")
        else:
            _log(f"KhÃ´ng cÃ³ áº£nh cho BQMS {bqms}", "warn")

    # First 2 products in template rows
    for i, product in enumerate(products[:2]):
        _fill_product(SP_START + i * ROWS_PER_SP, i + 1, product)

    # Products 3+: insert rows
    if len(products) > 2:
        insert_after = SP_START + 2 * ROWS_PER_SP - 1
        template_src = SP_START + ROWS_PER_SP
        for j, product in enumerate(products[2:]):
            insert_rows_copy_style(ws, insert_after, ROWS_PER_SP, template_src)
            new_start = insert_after + 1
            _fill_product(new_start, j + 3, product)
            insert_after = new_start + ROWS_PER_SP - 1

    # Date cell L31 (adjusted for inserted rows)
    extra_rows = max(0, len(products) - 2) * ROWS_PER_SP
    date_row   = 31 + extra_rows
    now        = datetime.now()
    safe_set_cell(ws, date_row, L, f"NgÃ y {now.day} ThÃ¡ng {now.month} nÄƒm {now.year}")

    try:
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        _log(f"LÆ°u CAM KET: {output_path}", "ok")
        return True
    except PermissionError:
        _log(f"KhÃ´ng thá»ƒ lÆ°u (file Ä‘ang má»Ÿ?): {output_path}", "error")
        return False
    except Exception as e:
        _log(f"Lá»—i lÆ°u CAM KET: {e}", "error")
        return False


# ---------------------------------------------------------------------------
# Commercial Quotation filler
# ---------------------------------------------------------------------------

def fill_quotation(template_path: str, products: list, images_map: dict,
                   rfq_no: str, output_path: str) -> bool:
    """
    Fill Commercial Quotation Form template.
    Returns True on success.
    """
    try:
        wb = openpyxl.load_workbook(template_path)
    except PermissionError:
        _log(f"Quotation template Ä‘ang bá»‹ khÃ³a: {template_path}", "error")
        return False
    except Exception as e:
        _log(f"Lá»—i má»Ÿ Quotation template: {e}", "error")
        return False

    # Find BQMS CODE sheet or fall back to active
    ws = None
    for name in wb.sheetnames:
        if "bqms" in name.lower() or "code" in name.lower():
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    now = datetime.now()

    # Header cells
    safe_set_cell(ws, 4, column_index_from_string("C"), now.strftime("%d/%m/%Y"))
    safe_set_cell(ws, 6, column_index_from_string("H"), rfq_no)
    safe_set_cell(ws, 7, column_index_from_string("C"),
                  f"QTAMABN-SEV {now.strftime('%d%m%Y')} - {rfq_no}")
    short_name = products[0].get("short_name", "") if products else ""
    safe_set_cell(ws, 14, column_index_from_string("A"),
                  f"Product description/ TÃªn hÃ ng: {short_name}")

    DATA_START   = 17
    TEMPLATE_ROW = 17

    A = column_index_from_string("A")
    B = column_index_from_string("B")
    C = column_index_from_string("C")
    D = column_index_from_string("D")
    F = column_index_from_string("F")
    G = column_index_from_string("G")
    H = column_index_from_string("H")
    I = column_index_from_string("I")
    J = column_index_from_string("J")

    def _fill_data_row(row: int, idx: int, product: dict) -> None:
        bqms     = product.get("bqms", "")
        spec     = product.get("spec", "")
        safe_set_cell(ws, row, A, idx)
        safe_set_cell(ws, row, B, product.get("don_hang", ""))
        safe_set_cell(ws, row, C, f"{bqms}\n{spec}")
        safe_set_cell(ws, row, D, product.get("maker", ""))
        if bqms in images_map:
            xl_img = pil_image_to_openpyxl_image(images_map[bqms])
            if xl_img:
                try:
                    ws.add_image(xl_img, f"E{row}")
                except Exception as e:
                    _log(f"KhÃ´ng thÃªm áº£nh E{row}: {e}", "warn")
        safe_set_cell(ws, row, F, product.get("don_vi", "EA") or "EA")
        safe_set_cell(ws, row, G, product.get("so_luong", ""))
        safe_set_cell(ws, row, H, "")               # Unit Price â€” always blank
        ws.cell(row=row, column=I).value = f"=G{row}*H{row}"
        safe_set_cell(ws, row, J, "")               # Notes

    # First product in template row
    _fill_data_row(DATA_START, 1, products[0])
    last_data_row = DATA_START

    # Additional products: insert rows
    for i, product in enumerate(products[1:], start=2):
        ws.insert_rows(last_data_row + 1)
        new_row = last_data_row + 1
        copy_row_style(ws, TEMPLATE_ROW, new_row)
        copy_merged_ranges_for_row(ws, TEMPLATE_ROW, new_row)
        _fill_data_row(new_row, i, product)
        last_data_row = new_row

    # Grand Total row
    total_row = last_data_row + 1
    ws.insert_rows(total_row)
    copy_row_style(ws, TEMPLATE_ROW, total_row)
    safe_set_cell(ws, total_row, A, "Grand Total (vnd)")
    ws.cell(row=total_row, column=G).value = f"=SUM(G{DATA_START}:G{last_data_row})"
    ws.cell(row=total_row, column=I).value = f"=SUM(I{DATA_START}:I{last_data_row})"

    try:
        os.makedirs(os.path.dirname(os.path.abspath(output_path)), exist_ok=True)
        wb.save(output_path)
        _log(f"LÆ°u Quotation: {output_path}", "ok")
        return True
    except PermissionError:
        _log(f"KhÃ´ng thá»ƒ lÆ°u (file Ä‘ang má»Ÿ?): {output_path}", "error")
        return False
    except Exception as e:
        _log(f"Lá»—i lÆ°u Quotation: {e}", "error")
        return False


# ---------------------------------------------------------------------------
# PDF export
# ---------------------------------------------------------------------------

def export_pdf(xlsx_path: str, output_dir: str, libreoffice_path: str = None) -> str:
    """
    Export xlsx to PDF via LibreOffice headless. Retries 3x.
    Returns pdf path on success, None on failure.
    """
    if libreoffice_path is None:
        # Try both standard paths
        for candidate in [
            "C:/Program Files/LibreOffice/program/soffice.exe",
            "C:/Program Files (x86)/LibreOffice/program/soffice.exe",
            "soffice",
        ]:
            if candidate == "soffice" or os.path.exists(candidate):
                libreoffice_path = candidate
                break

    if not libreoffice_path:
        _log("LibreOffice khÃ´ng tÃ¬m tháº¥y â€” bá» qua PDF.", "warn")
        return None

    base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    pdf_path  = os.path.join(output_dir, f"{base_name}.pdf")

    for attempt in range(3):
        try:
            result = subprocess.run(
                [libreoffice_path, "--headless", "--convert-to", "pdf",
                 "--outdir", output_dir, xlsx_path],
                timeout=60,
                capture_output=True,
                text=True,
            )
            if result.returncode == 0 and os.path.exists(pdf_path):
                _log(f"PDF xuáº¥t thÃ nh cÃ´ng: {base_name}.pdf", "ok")
                return pdf_path
            else:
                _log(f"LibreOffice láº§n {attempt+1} lá»—i: {result.stderr[:100]}", "warn")
        except FileNotFoundError:
            _log("LibreOffice (soffice) khÃ´ng tÃ¬m tháº¥y â€” bá» qua PDF.", "warn")
            return None
        except subprocess.TimeoutExpired:
            _log(f"LibreOffice timeout láº§n {attempt+1}.", "warn")

    return None


# ---------------------------------------------------------------------------
# Output folder management
# ---------------------------------------------------------------------------

def create_output_folder(base_path: str, rfq_no: str, on_conflict: str = "version") -> str:
    """
    Create output folder for rfq_no under base_path.
    on_conflict: 'version' (add _v2, _v3...) | 'overwrite' | 'skip' (returns None)
    Returns folder path or None if skipped.
    """
    folder = os.path.join(base_path, sanitize_filename(rfq_no))
    if not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)
        return folder

    if on_conflict == "overwrite":
        return folder
    elif on_conflict == "version":
        for v in range(2, 100):
            candidate = f"{folder}_v{v}"
            if not os.path.exists(candidate):
                os.makedirs(candidate)
                return candidate
    else:  # skip
        return None

    return folder


# ---------------------------------------------------------------------------
# Config helper  (delegates to modules/config for DRY)
# ---------------------------------------------------------------------------
from modules.config import get_config as _mc_get, save_config as _mc_save


def load_config(config_path: str = None) -> dict:
    """Load config.json. Returns dict with defaults for missing keys."""
    return _mc_get()


def save_config(cfg: dict, config_path: str = None) -> bool:
    result = _mc_save(cfg)
    if not result:
        _log("Loi luu config", "error")
    return result


# ---------------------------------------------------------------------------
# Main orchestration
# ---------------------------------------------------------------------------

def run_auto_fill_job(
    products: list,
    config: dict,
    images_map: dict,
    progress_callback=None
) -> dict:
    """
    Orchestrate: folder â†’ cam_ket â†’ quotation â†’ PDFs.
    progress_callback(step, total, message) â€” optional.

    products: list of order dicts (same RFQ No., already filtered TM).
    Returns: {success, rfq_no, files, errors}
    """
    def _progress(step, total, msg):
        _log(f"[{step}/{total}] {msg}")
        if progress_callback:
            progress_callback(step, total, msg)

    rfq_no   = products[0].get("don_hang", "unknown") if products else "unknown"
    result   = {"success": False, "rfq_no": rfq_no, "files": [], "errors": []}
    app_root = os.path.normpath(os.path.join(os.path.dirname(__file__), "..", ".."))

    # Resolve templates dir
    tmpl_dir_raw = config.get("templates_dir", "./templates")
    tmpl_dir     = tmpl_dir_raw if os.path.isabs(tmpl_dir_raw) else os.path.join(app_root, tmpl_dir_raw)
    cam_ket_tmpl = os.path.join(tmpl_dir, config["cam_ket_template"])
    quot_tmpl    = os.path.join(tmpl_dir, config["quotation_template"])

    # Validate templates exist
    for p in [cam_ket_tmpl, quot_tmpl]:
        if not os.path.exists(p):
            result["errors"].append(f"Template khÃ´ng tÃ¬m tháº¥y: {p}")
            _log(f"Template khÃ´ng tÃ¬m tháº¥y: {p}", "error")
            return result

    # Resolve output dir
    rfq_folder = config.get("rfq_folder") or config.get("output_base_path", "./outputs")
    if not os.path.isabs(rfq_folder):
        rfq_folder = os.path.join(app_root, rfq_folder)
    on_conflict = config.get("on_conflict", "version")

    _progress(0, 4, f"Táº¡o thÆ° má»¥c cho {rfq_no}...")
    out_dir = create_output_folder(rfq_folder, rfq_no, on_conflict=on_conflict)
    if out_dir is None:
        result["errors"].append(f"Bá» qua {rfq_no} (cháº¿ Ä‘á»™ skip)")
        return result

    safe_dh     = sanitize_filename(rfq_no)
    cam_ket_out = os.path.join(out_dir, f"CAM_KET_{safe_dh}.xlsx")
    quot_out    = os.path.join(out_dir, f"QUOTATION_{safe_dh}.xlsx")

    # Step 1: Fill CAM KET
    _progress(1, 4, "Äang Ä‘iá»n Cam Káº¿t...")
    cam_ok = fill_cam_ket(cam_ket_tmpl, products, images_map, cam_ket_out)
    if cam_ok:
        result["files"].append(cam_ket_out)

    # Step 2: Fill Quotation
    _progress(2, 4, "Äang Ä‘iá»n BÃ¡o GiÃ¡...")
    quot_ok = fill_quotation(quot_tmpl, products, images_map, rfq_no, quot_out)
    if quot_ok:
        result["files"].append(quot_out)

    # Step 3: Export PDFs
    libre_path = config.get("libreoffice_path")
    if not libre_path or not os.path.exists(libre_path):
        libre_path = config.get("libreoffice_path_alt")

    _progress(3, 4, "Äang xuáº¥t PDF...")
    if cam_ok:
        pdf = export_pdf(cam_ket_out, out_dir, libre_path)
        if pdf:
            result["files"].append(pdf)
    if quot_ok:
        pdf = export_pdf(quot_out, out_dir, libre_path)
        if pdf:
            result["files"].append(pdf)

    _progress(4, 4, "HoÃ n thÃ nh.")
    result["success"] = cam_ok or quot_ok
    result["output_dir"] = out_dir
    return result


def _norm_key(v: str) -> str:
    s = clean_cell(v).lower()
    s = s.replace(" ", "")
    return re.sub(r"[^a-z0-9]", "", s)


def _find_rfq_folder(root_rfq_dir: str, rfq_no: str) -> str:
    if not root_rfq_dir or not os.path.isdir(root_rfq_dir):
        return ""
    rfq_key = _norm_key(rfq_no)
    if not rfq_key:
        return ""

    candidates = []
    now = datetime.now()
    month_hints = [f"thang{now.month}", f"thang{max(now.month - 1, 1)}", f"thang{min(now.month + 1, 12)}"]
    year_hint = str(now.year)

    for cur, dirs, _ in os.walk(root_rfq_dir):
        base = os.path.basename(cur)
        bkey = _norm_key(base)
        if bkey.startswith(rfq_key):
            score = 0
            pkey = _norm_key(cur)
            if year_hint in pkey:
                score += 2
            if any(mh in pkey for mh in month_hints):
                score += 2
            try:
                mt = os.path.getmtime(cur)
            except Exception:
                mt = 0
            candidates.append((score, mt, cur))

    if not candidates:
        return ""
    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return candidates[0][2]


def _is_marker_cell(v) -> bool:
    s = clean_cell(v)
    if not s:
        return False
    compact = re.sub(r"\s+", "", s)
    # Accept wider marker variants: "절사 (-)", "절사(-)", "절사"
    return ("\uc808\uc0ac" in compact)


def _to_decimal_num(v):
    s = clean_cell(v)
    if not s:
        return None
    s = s.replace(" ", "").replace(",", "")
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        m = re.search(r"[-+]?\d+(?:\.\d+)?", s)
        if not m:
            return None
        try:
            return Decimal(m.group(0))
        except (InvalidOperation, ValueError):
            return None


def _add_to_cell_k(ws, row: int, delta_value) -> None:
    col_k = column_index_from_string("K")
    cur = _to_decimal_num(ws.cell(row=row, column=col_k).value)
    delta = _to_decimal_num(delta_value)
    cur = abs(cur) if cur is not None else Decimal(0)
    delta = abs(delta) if delta is not None else Decimal(0)
    # Business rule: this K value must remain positive and accumulative.
    new_val = cur + delta
    if new_val == new_val.to_integral_value():
        ws.cell(row=row, column=col_k).value = int(new_val)
    else:
        ws.cell(row=row, column=col_k).value = float(new_val)


def _calc_new_positive_k_value(current_value, delta_value):
    cur = _to_decimal_num(current_value)
    delta = _to_decimal_num(delta_value)
    cur = abs(cur) if cur is not None else Decimal(0)
    delta = abs(delta) if delta is not None else Decimal(0)
    new_val = cur + delta
    if new_val == new_val.to_integral_value():
        return int(new_val)
    return float(new_val)


def _extract_level_from_name(name: str) -> int:
    n = clean_cell(name).lower()
    n_fold = unicodedata.normalize("NFKD", n)
    n_fold = "".join(ch for ch in n_fold if not unicodedata.combining(ch))
    n_fold = n_fold.replace("đ", "d")
    m = re.search(r"\bl\s*([0-9]{1,2})\b", n_fold)
    if m:
        return int(m.group(1))
    m = re.search(r"lan[^0-9]*([0-9]{1,2})", n_fold)
    if m:
        return int(m.group(1))
    return -1


def _clone_l_folder_for_next_round(rfq_folder: str, target_level: int = 2) -> tuple[str, int, str]:
    subdirs = []
    for name in os.listdir(rfq_folder):
        p = os.path.join(rfq_folder, name)
        if os.path.isdir(p):
            lv = _extract_level_from_name(name)
            if lv > 0:
                subdirs.append((lv, p, name))

    target_level = int(target_level or 2)
    if target_level < 1:
        target_level = 1

    if not subdirs:
        # Fallback: infer "round folders" by content (must contain at least one Excel file),
        # but never copy from loose files in RFQ parent.
        inferred = []
        for name in os.listdir(rfq_folder):
            p = os.path.join(rfq_folder, name)
            if not os.path.isdir(p):
                continue
            excels = [
                f for f in os.listdir(p)
                if os.path.isfile(os.path.join(p, f))
                and f.lower().endswith((".xlsx", ".xlsm"))
                and not f.startswith("~$")
            ]
            if not excels:
                continue
            lv = _extract_level_from_name(name)
            if lv < 0:
                # unknown level folder; keep with lv=0 and use as fallback source.
                lv = 0
            try:
                mt = os.path.getmtime(p)
            except Exception:
                mt = 0
            inferred.append((lv, mt, p, name))

        if not inferred:
            raise RuntimeError(
                "No existing quote round folder (L1/L2/...) with Excel files found under RFQ folder."
            )

        inferred.sort(key=lambda x: (x[0], x[1]), reverse=True)
        if target_level > 1:
            src = next((x for x in inferred if x[0] == (target_level - 1)), inferred[0])
        else:
            src = inferred[0]
        src_path = src[2]
        src_name = src[3]
        next_lv = target_level

        dst_name = _rename_round_name(src_name, next_lv)
        dst_path = os.path.join(rfq_folder, sanitize_filename(dst_name))
        if os.path.exists(dst_path):
            for v in range(2, 100):
                cand = f"{dst_path}_v{v}"
                if not os.path.exists(cand):
                    dst_path = cand
                    break
        shutil.copytree(src_path, dst_path)
        _rename_excel_files_for_level(dst_path, next_lv)
        _log(f"Inferred source round folder: {src_path} -> {dst_path}", "ok")
        return src_path, next_lv, dst_path

    subdirs.sort(key=lambda x: x[0])
    by_lv = {x[0]: x for x in subdirs}
    if target_level > 1 and (target_level - 1) in by_lv:
        source = by_lv[target_level - 1]
    elif 1 in by_lv:
        source = by_lv[1]
    else:
        source = subdirs[-1]
    next_lv = target_level

    src_path = source[1]
    src_name = source[2]
    dst_name = _rename_round_name(src_name, next_lv)
    dst_path = os.path.join(rfq_folder, sanitize_filename(dst_name))

    if os.path.exists(dst_path):
        for v in range(2, 100):
            cand = f"{dst_path}_v{v}"
            if not os.path.exists(cand):
                dst_path = cand
                break
    shutil.copytree(src_path, dst_path)
    _rename_excel_files_for_level(dst_path, next_lv)
    return src_path, next_lv, dst_path


def _rename_round_name(src_name: str, target_level: int) -> str:
    """
    Rename round marker in folder name to target Lx.
    Must replace existing round token, not append (avoid L2_L3).
    """
    s = clean_cell(src_name)
    pats = [
        r"(?i)l\s*[_-]?\s*\d{1,2}",
        r"(?i)lan\s*bao\s*gia\s*\d{1,2}",
        r"(?i)lan[^0-9]{0,6}\d{1,2}",
        r"(?i)_l\d{1,2}",
    ]
    for p in pats:
        out, n = re.subn(p, f"L{int(target_level)}", s, count=1)
        if n > 0 and out != s:
            return out
    return f"Lan bao gia L{int(target_level)}"


def _rename_excel_files_for_level(folder_path: str, target_level: int) -> int:
    changed = 0
    for fn in os.listdir(folder_path):
        src = os.path.join(folder_path, fn)
        if not os.path.isfile(src):
            continue
        if not fn.lower().endswith((".xlsx", ".xlsm", ".xls")):
            continue
        if fn.startswith("~$"):
            continue
        base, ext = os.path.splitext(fn)
        new_base = _rename_round_name(base, target_level)
        if new_base == base:
            continue
        dst = os.path.join(folder_path, sanitize_filename(new_base) + ext)
        if os.path.normcase(dst) == os.path.normcase(src):
            continue
        if os.path.exists(dst):
            # avoid overwrite
            for i in range(2, 100):
                cand = os.path.join(folder_path, sanitize_filename(new_base) + f"_v{i}" + ext)
                if not os.path.exists(cand):
                    dst = cand
                    break
        try:
            os.rename(src, dst)
            changed += 1
        except Exception:
            pass
    return changed


def _delete_all_pdfs(folder: str) -> int:
    c = 0
    for cur, _, files in os.walk(folder):
        for fn in files:
            if fn.lower().endswith(".pdf"):
                try:
                    os.remove(os.path.join(cur, fn))
                    c += 1
                except Exception:
                    pass
    return c


def _load_v2_map(history_path: str, rfq_no: str) -> dict:
    from tools.tool2_pricetracker.engine import load_bqms_analytics

    r = load_bqms_analytics(history_path)
    if not r.get("ok"):
        raise RuntimeError(r.get("error") or "Cannot load BQMS history file.")
    df = r["df"].copy()
    if df is None or df.empty:
        raise RuntimeError("BQMS history is empty.")

    rfq_key = _norm_key(rfq_no)
    df = df[df["rfq_no"].map(_norm_key) == rfq_key]
    if df.empty:
        return {}
    df = df[df["quote_v2"].notna()].copy()
    if df.empty:
        return {}

    if "date" in df.columns:
        df = df.sort_values("date", ascending=False, na_position="last")

    out = {}
    for _, row in df.iterrows():
        code = clean_cell(row.get("bqms_code", ""))
        if not code:
            continue
        ckey = _norm_key(code)
        if not ckey or ckey in out:
            continue
        out[ckey] = {
            "code": code,
            "price": row.get("quote_v2"),
        }
    return out


def _find_code_for_sheet(ws, v2_map: dict) -> str:
    if not v2_map:
        return ""
    title_key = _norm_key(ws.title)
    for ckey, info in v2_map.items():
        if ckey and ckey in title_key:
            return info["code"]

    max_r = min(ws.max_row, 220)
    max_c = min(ws.max_column, 26)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            t = _norm_key(ws.cell(row=r, column=c).value)
            if t in v2_map:
                return v2_map[t]["code"]
    return ""


_MARKER_TEXT = "\uc808\uc0ac (-)"


def _fill_marker_in_col_k(ws, value) -> tuple[int, list[str]]:
    filled = 0
    addresses = []
    # Business rule: target row is always the row right below "Profit".
    fr = _fallback_cut_row(ws)
    target_rows = [fr] if fr else []
    # Keep marker rows as fallback only when profit anchor is missing.
    if not target_rows:
        target_rows = _marker_rows_in_sheet(ws)
    for r in target_rows:
        try:
            _add_to_cell_k(ws, r, value)
            filled += 1
            addresses.append(f"K{r}")
        except Exception:
            pass
    return filled, addresses


def _fill_first_empty_in_col_k(ws, value) -> tuple[int, list[str]]:
    col_k = column_index_from_string("K")
    last = max(ws.max_row, 1)
    for r in range(2, last + 1):
        cur = clean_cell(ws.cell(row=r, column=col_k).value)
        if cur == "":
            _add_to_cell_k(ws, r, value)
            return 1, [f"K{r}"]
    _add_to_cell_k(ws, last + 1, value)
    return 1, [f"K{last + 1}"]


def _pending_marker_count(ws) -> int:
    fr = _fallback_cut_row(ws)
    if fr:
        return 1
    return len(_marker_rows_in_sheet(ws))


def _marker_rows_in_col_k(ws) -> list[int]:
    col_k = column_index_from_string("K")
    rows = []
    for r in range(1, ws.max_row + 1):
        if _is_marker_cell(ws.cell(row=r, column=col_k).value):
            rows.append(r)
    return rows


def _marker_rows_in_sheet(ws) -> list[int]:
    rows = []
    max_r = min(ws.max_row, 350)
    max_c = min(ws.max_column, 14)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            if _is_marker_cell(ws.cell(row=r, column=c).value):
                rows.append(r)
                break
    return rows


def _fallback_cut_row(ws) -> int:
    r_profit = _find_row_by_text(ws, ["profit"])
    r_result = _find_row_by_text(ws, ["result", "total", "amount"])
    if r_profit and r_result and (r_profit + 1) < r_result:
        return r_profit + 1
    if r_profit:
        return r_profit + 1
    return 0


def _is_summary_sheet(ws) -> bool:
    t = clean_cell(ws.title).lower()
    fold = unicodedata.normalize("NFKD", t)
    fold = "".join(ch for ch in fold if not unicodedata.combining(ch))
    fold = fold.replace("đ", "d")
    if any(k in fold for k in ["material", "process", "summary", "tong hop", "total"]):
        return True
    # Heuristic: no BQMS-like pattern in first area
    max_r = min(ws.max_row, 80)
    max_c = min(ws.max_column, 12)
    code_pat = re.compile(r"Z\d{6,}-\d{3,}", re.IGNORECASE)
    for r in range(1, max_r + 1):
        for c in range(1, max_c + 1):
            v = clean_cell(ws.cell(row=r, column=c).value)
            if code_pat.search(v):
                return False
    return "material" in fold or "process" in fold


def _find_row_by_text(ws, keywords: list[str]) -> int:
    max_r = min(ws.max_row, 350)
    max_c = min(ws.max_column, 12)
    for r in range(1, max_r + 1):
        row_text = " ".join(clean_cell(ws.cell(row=r, column=c).value).lower() for c in range(1, max_c + 1))
        if all(k.lower() in row_text for k in keywords):
            return r
    return 0


def _find_rows_by_text(ws, keywords: list[str]) -> list[int]:
    rows = []
    max_r = min(ws.max_row, 350)
    max_c = min(ws.max_column, 12)
    for r in range(1, max_r + 1):
        row_text = " ".join(clean_cell(ws.cell(row=r, column=c).value).lower() for c in range(1, max_c + 1))
        if all(k.lower() in row_text for k in keywords):
            rows.append(r)
    return rows


def _find_material_process_row(ws, r_profit: int) -> int:
    rows = _find_rows_by_text(ws, ["material", "process", "total"])
    if not rows:
        return 0
    # Prefer the nearest row above profit in the same block.
    above = [r for r in rows if r <= r_profit] if r_profit else rows
    return max(above) if above else rows[-1]


def _find_management_row_near_profit(ws, r_profit: int) -> int:
    if not r_profit:
        return 0
    # Prefer explicit text hit first.
    r_mgmt = _find_row_by_text(ws, ["management", "expenses"])
    if r_mgmt:
        return r_mgmt
    # Fallback: closest non-empty row above Profit (within 6 rows).
    for r in range(max(1, r_profit - 6), r_profit):
        txt = " ".join(clean_cell(ws.cell(row=r, column=c).value).lower() for c in range(1, 8))
        if txt.strip():
            return r
    return max(1, r_profit - 1)


def _find_result_row_near_cut(ws, r_cut: int) -> int:
    # Must use the last "Result Total Amount" row in sheet.
    result_rows = _find_rows_by_text(ws, ["result", "total", "amount"])
    if result_rows:
        return result_rows[-1]
    if not r_cut:
        return 0
    # Fallback: next non-empty title-like row below cut.
    max_r = min(ws.max_row, r_cut + 12)
    for rr in range(r_cut + 1, max_r + 1):
        txt = " ".join(clean_cell(ws.cell(row=rr, column=c).value).lower() for c in range(1, 8))
        if "result" in txt or "total" in txt:
            return rr
    return min(ws.max_row, r_cut + 1)


def _verify_or_fix_result_total_formula(ws) -> tuple[str, str]:
    """
    Ensure K[row_result_total] has expected formula:
    =ROUNDUP(SUM(K[mat]:K[profit])-K[cut],0)
    Returns (status, formula_or_reason)
    """
    col_k = column_index_from_string("K")
    r_profit = _find_row_by_text(ws, ["profit"])
    r_cut = _fallback_cut_row(ws)
    r_mgmt = _find_management_row_near_profit(ws, r_profit)
    r_mat = _find_material_process_row(ws, r_profit)
    r_result = _find_result_row_near_cut(ws, r_cut)

    if not all([r_profit, r_cut, r_result]):
        return "skip", "missing anchors"

    # Standardized formula requested by business:
    # =ROUNDUP((SUM(K[Material+ProcessTotal or Management]:K[Profit]) - K[cut]), -3)
    sum_start = r_mat if r_mat else r_mgmt
    if not sum_start:
        return "skip", "missing sum-start anchors"
    expected = f"=ROUNDUP((SUM(K{sum_start}:K{r_profit})-K{r_cut}),-3)"
    # Do not write by openpyxl here (to preserve drawings/external links).
    return "fixed", expected


def _plan_fill_marker_in_col_k(ws, value) -> tuple[int, list[str], list]:
    filled = 0
    addresses = []
    updates = []
    fr = _fallback_cut_row(ws)
    target_rows = [fr] if fr else []
    if not target_rows:
        target_rows = _marker_rows_in_sheet(ws)
    col_k = column_index_from_string("K")
    for r in target_rows:
        try:
            cur = ws.cell(row=r, column=col_k).value
            new_val = _calc_new_positive_k_value(cur, value)
            addr = f"K{r}"
            filled += 1
            addresses.append(addr)
            updates.append((addr, new_val))
        except Exception:
            pass
    return filled, addresses, updates


def _plan_fill_first_empty_in_col_k(ws, value) -> tuple[int, list[str], list]:
    col_k = column_index_from_string("K")
    last = max(ws.max_row, 1)
    for r in range(2, last + 1):
        cur_txt = clean_cell(ws.cell(row=r, column=col_k).value)
        if cur_txt == "":
            cur = ws.cell(row=r, column=col_k).value
            new_val = _calc_new_positive_k_value(cur, value)
            addr = f"K{r}"
            return 1, [addr], [(addr, new_val)]
    cur = ws.cell(row=last + 1, column=col_k).value
    new_val = _calc_new_positive_k_value(cur, value)
    addr = f"K{last + 1}"
    return 1, [addr], [(addr, new_val)]


def _apply_excel_edits_via_com(workbook_path: str, updates: list[dict]) -> str:
    if not updates:
        return ""

    def _esc_ps(s: str) -> str:
        return str(s).replace("'", "''")

    payload = json.dumps(updates, ensure_ascii=False)
    script = f"""
$ErrorActionPreference = 'Stop'
$workbook = '{_esc_ps(workbook_path)}'
$updates = @'
{payload}
'@ | ConvertFrom-Json

function Set-CellValueSafe($ws, $cell, $rawValue) {{
  if ($null -eq $rawValue) {{
    $ws.Range($cell).Value2 = $null
    return
  }}

  if ($rawValue -is [string]) {{
    $txt = $rawValue.Trim()
    if ($txt -eq '') {{
      $ws.Range($cell).Value2 = ''
      return
    }}
    $num = 0.0
    $txtNorm = $txt.Replace(',', '')
    if ([double]::TryParse($txtNorm, [System.Globalization.NumberStyles]::Any, [System.Globalization.CultureInfo]::InvariantCulture, [ref]$num)) {{
      $ws.Range($cell).Value2 = [double]$num
      return
    }}
    $ws.Range($cell).Value2 = $txt
    return
  }}

  if (
    ($rawValue -is [byte]) -or ($rawValue -is [int16]) -or ($rawValue -is [int32]) -or ($rawValue -is [int64]) -or
    ($rawValue -is [single]) -or ($rawValue -is [double]) -or ($rawValue -is [decimal])
  ) {{
    $ws.Range($cell).Value2 = [double]$rawValue
    return
  }}

  try {{
    $ws.Range($cell).Value2 = [double]$rawValue
  }} catch {{
    $ws.Range($cell).Value2 = [string]$rawValue
  }}
}}

$excel = New-Object -ComObject Excel.Application
$wb = $null

$excel.Visible = $false
$excel.DisplayAlerts = $false
$excel.AskToUpdateLinks = $false

for($i=0; $i -lt 3; $i++){{
  try {{
    $wb = $excel.Workbooks.Open($workbook, 0, $false)
    break
  }} catch {{
    Start-Sleep -Milliseconds 700
  }}
}}

if(-not $wb){{ throw "Cannot open workbook via Excel COM for edit: $workbook" }}

try {{
  foreach($u in $updates){{
    $sheetName = [string]$u.sheet
    $cell = [string]$u.cell
    $ws = $wb.Worksheets.Item($sheetName)
    if($u.PSObject.Properties.Name -contains 'formula' -and -not [string]::IsNullOrWhiteSpace([string]$u.formula)){{
      $ws.Range($cell).Formula = [string]$u.formula
    }} else {{
      Set-CellValueSafe $ws $cell $u.value
    }}
  }}
  $wb.Save()
}}
finally {{
  if($wb){{ $wb.Close($true) }}
  if($excel){{ $excel.Quit() }}
  if($wb){{ [System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb) | Out-Null }}
  if($excel){{ [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null }}
}}
"""
    try:
        r = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            timeout=240,
            capture_output=True,
            text=True,
        )
    except Exception as e:
        return str(e)

    if r.returncode != 0:
        return (r.stderr or r.stdout or f"returncode={r.returncode}")[:600]
    return ""


def _export_sheet_pdfs_excel_com(
    workbook_path: str,
    out_dir: str,
    sheet_items: list,
    rfq_no: str,
    level: int,
    use_sheet_name: bool = False,
) -> tuple[list, str]:
    if not sheet_items:
        return [], "No sheet items."
    items_json = json.dumps(sheet_items, ensure_ascii=False)

    def _esc_ps(s: str) -> str:
        return str(s).replace("'", "''")

    script = f"""
$ErrorActionPreference = 'Stop'
$workbook = '{_esc_ps(workbook_path)}'
$outDir = '{_esc_ps(out_dir)}'
$rfq = '{_esc_ps(rfq_no)}'
$lv = {int(level)}
$useSheet = @'
{json.dumps(bool(use_sheet_name)).lower()}
'@ | ConvertFrom-Json
$items = @'
{items_json}
'@ | ConvertFrom-Json

$excel = New-Object -ComObject Excel.Application
$wb = $null

$excel.Visible = $false
$excel.DisplayAlerts = $false
$excel.AskToUpdateLinks = $false

for($i=0; $i -lt 3; $i++){{
  try {{
    $wb = $excel.Workbooks.Open($workbook, 0, $true)
    break
  }} catch {{
    Start-Sleep -Milliseconds 700
  }}
}}

if(-not $wb){{ throw "Cannot open workbook via Excel COM: $workbook" }}
$outs = New-Object System.Collections.Generic.List[string]
try {{
  $cnt = @($items).Count
  foreach($it in $items){{
    $sheetName = [string]$it.sheet
    $code = [string]$it.code
    if([string]::IsNullOrWhiteSpace($code)){{ $code = $sheetName }}
    if($useSheet -eq $true){{
      $name = "$($sheetName).pdf"
    }} elseif($cnt -eq 1){{
      $name = "$($rfq)_$($code)_AMABACNINH_L$($lv).pdf"
    }} else {{
      $name = "$($sheetName).pdf"
    }}
    foreach($ch in [System.IO.Path]::GetInvalidFileNameChars()){{ $name = $name.Replace($ch, '_') }}
    $pdfPath = Join-Path $outDir $name
    $ws = $wb.Worksheets.Item($sheetName)
    # Keep template print settings, only normalize scaling to avoid horizontal cut.
    try {{ $ws.ResetAllPageBreaks() }} catch {{}}
    $null = $ws.UsedRange
    try {{ $excel.PrintCommunication = $false }} catch {{}}
    $ps = $ws.PageSetup
    $ps.Zoom = $false
    $ps.FitToPagesWide = 1
    $ps.FitToPagesTall = $false
    try {{ $excel.PrintCommunication = $true }} catch {{}}
    # Use template-defined print area/layout for stable output.
    $ws.ExportAsFixedFormat(0, $pdfPath, 0, $true, $false)
    $outs.Add($pdfPath) | Out-Null
  }}
  $outs | ForEach-Object {{ Write-Output $_ }}
}}
finally {{
  if($wb){{ $wb.Close($false) }}
  if($excel){{ $excel.Quit() }}
  if($wb){{ [System.Runtime.InteropServices.Marshal]::ReleaseComObject($wb) | Out-Null }}
  if($excel){{ [System.Runtime.InteropServices.Marshal]::ReleaseComObject($excel) | Out-Null }}
}}
"""
    try:
        r = subprocess.run(
            ["powershell", "-NoProfile", "-ExecutionPolicy", "Bypass", "-Command", script],
            timeout=240,
            capture_output=True,
            text=True,
        )
    except Exception as e:
        return [], str(e)
    if r.returncode != 0:
        return [], (r.stderr or r.stdout or f"returncode={r.returncode}")[:500]
    outs = []
    cnt = len(sheet_items or [])
    for it in (sheet_items or []):
        sheet_name = str(it.get("sheet", "") or "")
        code = str(it.get("code", "") or "") or sheet_name
        if use_sheet_name:
            name = f"{sheet_name}.pdf"
        elif cnt == 1:
            name = f"{rfq_no}_{code}_AMABACNINH_L{int(level)}.pdf"
        else:
            name = f"{sheet_name}.pdf"
        name = sanitize_filename(name)
        p = os.path.join(out_dir, name)
        if os.path.isfile(p):
            outs.append(p)
    if not outs:
        diag_out = (r.stdout or "").strip().replace("\n", " | ")[:300]
        diag_err = (r.stderr or "").strip().replace("\n", " | ")[:300]
        return [], f"no_pdf_output stdout=[{diag_out}] stderr=[{diag_err}]"
    return outs, ""


def run_gia_cong_prepare_job(
    rfq_no: str,
    history_path: str,
    rfq_root_dir: str,
    rfq_folder_override: str = "",
    manual_timeout_sec: int = 300,
    selected_v2_map: dict | None = None,
    quote_level: int = 2,
    workflow_type: str = "gia_cong",
    progress_callback=None,
) -> dict:
    result = {
        "success": False,
        "rfq_no": rfq_no,
        "errors": [],
        "warnings": [],
        "output_dir": "",
        "level": None,
        "manual_required": [],
        "workflow_type": workflow_type,
        "edit_report": [],
        "excel_path": "",
        "sheet_items": [],
        "sheet_count": 0,
        "code_count": 0,
        "marker_total": 0,
    }

    def _progress(step, total, msg):
        _log(f"[{step}/{total}] {msg}")
        if progress_callback:
            progress_callback(step, total, msg)

    total = 6
    rfq_no = clean_cell(rfq_no)
    if not rfq_no:
        result["errors"].append("RFQ No. is empty.")
        return result
    if not os.path.isfile(history_path):
        result["errors"].append(f"BQMS history file not found: {history_path}")
        return result

    _progress(1, total, "Finding RFQ folder...")
    rfq_folder = ""
    if rfq_folder_override and os.path.isdir(rfq_folder_override):
        rfq_folder = os.path.normpath(rfq_folder_override)
    else:
        rfq_folder = _find_rfq_folder(rfq_root_dir, rfq_no)
    if not rfq_folder:
        result["errors"].append(f"Cannot find RFQ folder for: {rfq_no}")
        return result

    _progress(2, total, f"Preparing folder for L{int(quote_level)}...")
    try:
        _, lv, new_round_dir = _clone_l_folder_for_next_round(rfq_folder, target_level=int(quote_level))
    except Exception as e:
        result["errors"].append(str(e))
        return result
    result["level"] = lv
    result["output_dir"] = new_round_dir

    _progress(3, total, "Removing old PDFs in target folder...")
    _delete_all_pdfs(new_round_dir)

    _progress(4, total, "Loading price map...")
    try:
        if isinstance(selected_v2_map, dict) and selected_v2_map:
            v2_map = {}
            for k, v in selected_v2_map.items():
                nk = _norm_key(k)
                if nk:
                    v2_map[nk] = {"code": str(k), "price": v}
        else:
            v2_map = _load_v2_map(history_path, rfq_no)
    except Exception as e:
        result["errors"].append(str(e))
        return result
    if not v2_map:
        result["errors"].append(f"No price map found for RFQ: {rfq_no}")
        return result

    excel_files = [
        os.path.join(new_round_dir, f)
        for f in os.listdir(new_round_dir)
        if f.lower().endswith((".xlsx", ".xlsm")) and not f.startswith("~$")
    ]
    if not excel_files:
        result["errors"].append("No Excel file found in round folder.")
        return result
    excel_files.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    excel_path = excel_files[0]
    result["excel_path"] = excel_path

    _progress(5, total, "Editing Excel cells...")
    # Read by openpyxl only for mapping. Write via Excel COM to preserve drawing/external links.
    wb = openpyxl.load_workbook(excel_path, keep_links=True)
    result["sheet_count"] = int(len(wb.worksheets))
    sheet_items = []
    unresolved = []
    edit_report = []
    code_set = set()
    marker_total = 0
    com_updates = []

    for ws in wb.worksheets:
        marker_rows = _marker_rows_in_sheet(ws)
        marker_total += len(marker_rows)
        fstat, fval = _verify_or_fix_result_total_formula(ws)
        if _is_summary_sheet(ws):
            edit_report.append(
                {
                    "sheet": ws.title,
                    "sheet_type": "summary",
                    "code": "",
                    "price": None,
                    "marker_rows": ",".join(str(x) for x in marker_rows[:8]),
                    "marker_count": len(marker_rows),
                    "filled_cells": 0,
                    "addresses": "",
                    "formula_status": fstat,
                    "formula": fval,
                    "status": "summary_skip",
                }
            )
            continue
        code = _find_code_for_sheet(ws, v2_map)
        if not code:
            unresolved.append({"sheet": ws.title, "reason": "No BQMS code mapping"})
            edit_report.append(
                {
                    "sheet": ws.title,
                    "sheet_type": "product",
                    "code": "",
                    "price": None,
                    "marker_rows": ",".join(str(x) for x in marker_rows[:8]),
                    "marker_count": len(marker_rows),
                    "filled_cells": 0,
                    "addresses": "",
                    "formula_status": fstat,
                    "formula": fval,
                    "status": "no_code",
                }
            )
            continue
        code_set.add(_norm_key(code))
        info = v2_map.get(_norm_key(code))
        if not info:
            unresolved.append({"sheet": ws.title, "reason": f"No price for code {code}"})
            edit_report.append(
                {
                    "sheet": ws.title,
                    "sheet_type": "product",
                    "code": code,
                    "price": None,
                    "marker_rows": ",".join(str(x) for x in marker_rows[:8]),
                    "marker_count": len(marker_rows),
                    "filled_cells": 0,
                    "addresses": "",
                    "formula_status": fstat,
                    "formula": fval,
                    "status": "no_price",
                }
            )
            continue
        price = info.get("price")
        if str(workflow_type or "").lower().startswith("thuong"):
            filled_cells, addresses, planned_updates = _plan_fill_first_empty_in_col_k(ws, price)
        else:
            filled_cells, addresses, planned_updates = _plan_fill_marker_in_col_k(ws, price)
        for addr, val in planned_updates:
            com_updates.append({"sheet": ws.title, "cell": addr, "value": val})
        if fstat == "fixed" and fval and _find_result_row_near_cut(ws, _fallback_cut_row(ws)):
            r_result = _find_result_row_near_cut(ws, _fallback_cut_row(ws))
            com_updates.append({"sheet": ws.title, "cell": f"K{r_result}", "formula": fval})
        if filled_cells <= 0:
            unresolved.append({"sheet": ws.title, "reason": f"Marker '{_MARKER_TEXT}' row not found in sheet"})
            edit_report.append(
                {
                    "sheet": ws.title,
                    "sheet_type": "product",
                    "code": code,
                    "price": price,
                    "marker_rows": ",".join(str(x) for x in marker_rows[:8]),
                    "marker_count": len(marker_rows),
                    "filled_cells": 0,
                    "addresses": "",
                    "formula_status": fstat,
                    "formula": fval,
                    "status": "not_found",
                }
            )
        else:
            edit_report.append(
                {
                    "sheet": ws.title,
                    "sheet_type": "product",
                    "code": code,
                    "price": price,
                    "marker_rows": ",".join(str(x) for x in marker_rows[:8]),
                    "marker_count": len(marker_rows),
                    "filled_cells": int(filled_cells),
                    "addresses": ", ".join(addresses[:8]),
                    "formula_status": fstat,
                    "formula": fval,
                    "status": "edited",
                }
            )
        sheet_items.append({"sheet": ws.title, "code": code})

    wb.close()

    # Apply edits on a temp local copy to avoid OneDrive/Unicode path lock issues.
    src_ext = os.path.splitext(excel_path)[1] or ".xlsx"
    tmp_edit = ""
    try:
        fd, tmp_edit = tempfile.mkstemp(prefix="bsmq_edit_", suffix=src_ext)
        os.close(fd)
        shutil.copy2(excel_path, tmp_edit)
    except Exception:
        tmp_edit = excel_path

    com_err = _apply_excel_edits_via_com(tmp_edit, com_updates)
    if com_err:
        result["errors"].append(f"Excel COM edit failed: {com_err}")
        return result
    if os.path.normcase(tmp_edit) != os.path.normcase(excel_path):
        try:
            shutil.copy2(tmp_edit, excel_path)
        except Exception as e:
            result["errors"].append(f"Cannot copy edited workbook back to target folder: {e}")
            return result
        finally:
            try:
                os.remove(tmp_edit)
            except Exception:
                pass

    result["edit_report"] = edit_report
    result["manual_required"] = unresolved
    result["sheet_items"] = sheet_items
    result["code_count"] = int(len([x for x in code_set if x]))
    result["marker_total"] = int(marker_total)

    if unresolved:
        _progress(6, total, f"Manual review needed for {len(unresolved)} sheet(s).")
    else:
        _progress(6, total, "Excel edit completed.")
    result["success"] = True
    return result


def run_gia_cong_export_job(
    rfq_no: str,
    excel_path: str,
    output_dir: str,
    level: int,
    sheet_items: list,
    progress_callback=None,
) -> dict:
    result = {
        "success": False,
        "files": [],
        "errors": [],
        "warnings": [],
        "sheet_report": [],
        "export_profile": "locked_template_fit_width",
    }

    def _progress(step, total, msg):
        _log(f"[{step}/{total}] {msg}")
        if progress_callback:
            progress_callback(step, total, msg)

    if not excel_path or not os.path.isfile(excel_path):
        result["errors"].append(f"Excel file missing: {excel_path}")
        return result
    if not os.path.isdir(output_dir):
        result["errors"].append(f"Output folder missing: {output_dir}")
        return result
    if not sheet_items:
        result["warnings"].append("No sheet mapping available, fallback to whole workbook PDF.")

    # Export from a temporary snapshot to avoid lock conflicts when user keeps Excel open.
    src_ext = os.path.splitext(excel_path)[1] or ".xlsx"
    try:
        fd, tmp_export = tempfile.mkstemp(prefix="bsmq_export_", suffix=src_ext)
        os.close(fd)
        shutil.copy2(excel_path, tmp_export)
    except Exception:
        tmp_export = excel_path

    # Sequential export per sheet with progress; do not fail whole batch on one sheet.
    # Prefer original file first (can attach to opened workbook in Excel),
    # then fallback to temp snapshot.
    candidates = [tmp_export]
    if os.path.normcase(tmp_export) != os.path.normcase(excel_path):
        candidates.append(excel_path)

    ok_files = []
    failed = []

    if sheet_items:
        total = len(sheet_items)
        for idx, it in enumerate(sheet_items, start=1):
            sheet_name = str(it.get("sheet", "") or "")
            code = str(it.get("code", "") or "")
            one = [{"sheet": sheet_name, "code": code}]
            _progress(idx, total, f"Exporting {sheet_name} ({idx}/{total})...")

            done = False
            err_msg = ""
            for wb_path in candidates:
                files, err = _export_sheet_pdfs_excel_com(
                    wb_path,
                    output_dir,
                    one,
                    rfq_no,
                    level,
                    use_sheet_name=True,
                )
                if files:
                    ok_files.extend(files)
                    result["sheet_report"].append(
                        {
                            "sheet": sheet_name,
                            "code": code,
                            "status": "ok",
                            "pdf_file": files[0],
                            "error": "",
                        }
                    )
                    done = True
                    break
                if err:
                    err_msg = f"{err_msg} | {err}".strip(" |")

            if not done:
                result["sheet_report"].append(
                    {
                        "sheet": sheet_name,
                        "code": code,
                        "status": "failed",
                        "pdf_file": "",
                        "error": err_msg or "unknown COM error",
                    }
                )
                failed.append({"sheet": sheet_name, "error": err_msg or "unknown COM error"})

        if failed:
            result["warnings"].append(f"Failed sheets: {len(failed)}")
            for x in failed[:10]:
                result["warnings"].append(f"{x['sheet']}: {x['error'][:180]}")

    # If sequential COM export got nothing, fallback to whole-workbook LibreOffice.
    if not ok_files:
        _progress(1, 1, "COM export failed, trying LibreOffice fallback...")
        libre = load_config().get("libreoffice_path")
        fallback = export_pdf(excel_path, output_dir, libre)
        if fallback:
            ok_files = [fallback]
            result["warnings"].append("Used LibreOffice fallback (whole workbook PDF).")

    if not ok_files:
        result["errors"].append("Cannot export PDF. Check Excel COM / LibreOffice.")
        return result

    result["files"] = ok_files
    result["success"] = True
    _progress(1, 1, "Export completed.")
    return result


def run_gia_cong_l2_job(
    rfq_no: str,
    history_path: str,
    rfq_root_dir: str,
    manual_timeout_sec: int = 300,
    selected_v2_map: dict | None = None,
    quote_level: int = 2,
    workflow_type: str = "gia_cong",
    progress_callback=None,
) -> dict:
    prep = run_gia_cong_prepare_job(
        rfq_no=rfq_no,
        history_path=history_path,
        rfq_root_dir=rfq_root_dir,
        manual_timeout_sec=manual_timeout_sec,
        selected_v2_map=selected_v2_map,
        quote_level=quote_level,
        workflow_type=workflow_type,
        progress_callback=progress_callback,
    )
    if not prep.get("success"):
        prep["files"] = []
        return prep

    exp = run_gia_cong_export_job(
        rfq_no=rfq_no,
        excel_path=prep.get("excel_path", ""),
        output_dir=prep.get("output_dir", ""),
        level=int(prep.get("level") or quote_level),
        sheet_items=prep.get("sheet_items", []),
        progress_callback=progress_callback,
    )
    out = {**prep}
    out["files"] = exp.get("files", [])
    out["errors"] = prep.get("errors", []) + exp.get("errors", [])
    out["warnings"] = prep.get("warnings", []) + exp.get("warnings", [])
    out["success"] = bool(exp.get("success"))
    return out

