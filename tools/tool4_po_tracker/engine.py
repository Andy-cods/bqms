"""
Tool 4 - PO Tracker Engine (v2 — optimized)
Samsung BQMS Vendor Portal: https://www.sec-bqms.com

Flow:
  1. Khởi động Chrome (webdriver-manager tự quản lý ChromeDriver)
  2. Mở trang login → SWITCH LANGUAGE ENGLISH ngay trên trang login
  3. Điền credentials → Submit
  4. Handle popup thông báo tiếng Hàn (nếu có)
  5. Điều hướng Execution > MRO > P/O Receipt

Tối ưu v2:
  - Language switch tại login page (không reload thêm sau login)
  - WebDriverWait thay cho time.sleep() cứng ở mọi chỗ có thể
  - Multi-label JS query thay cho sequential polling
  - Submit button selectors bổ sung (arrow button Samsung)
"""

import os
import time
import logging
from typing import Optional
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    TimeoutException,
    NoSuchElementException,
    StaleElementReferenceException,
    WebDriverException,
)
from webdriver_manager.chrome import ChromeDriverManager

logger = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────
PORTAL_URL = (
    "https://www.sec-bqms.com/bqms/vendorPortal/vendorPortalMain.do"
    "?_mainLayOut=vendorPortalLayout"
)
LOGIN_URL = "https://www.sec-bqms.com/bqms/vendorPortal/anonymous/vendorLogin.do?_frameF=true"
PO_RECEIPT_MRO_URL = (
    "https://www.sec-bqms.com/bqms/mro/forward/vendor/"
    "vendorPoConfirm.do?target=vendor&_menuId=AZknkggsAB8V-Qhq&_menuF=true"
)
PO_RECEIPT_MRO_MENU_ID = "AZknkggsAB8V-Qhq"

# Credentials (load từ env nếu có, fallback hardcode cho single-user tool)
DEFAULT_USERNAME = os.environ.get("BSMQ_USERNAME", "amabnjsc")
DEFAULT_PASSWORD = os.environ.get("BSMQ_PASSWORD", "AMA202601@1")

# Timeouts (giây)
PAGE_LOAD_TIMEOUT = 30
ELEMENT_WAIT = 15
LOGIN_VERIFY_TIMEOUT = 25

# Menu labels theo ngôn ngữ
EXECUTION_LABELS = ["Execution", "발주", "执行"]
PO_RECEIPT_LABELS = ["발주접수", "P/O Receipt"]


# ── Driver Setup ──────────────────────────────────────────────────────────────

def _build_chrome_options(headless: bool = False, download_dir: str = "") -> Options:
    opts = Options()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--disable-blink-features=AutomationControlled")
    opts.add_experimental_option("excludeSwitches", ["enable-automation"])
    opts.add_experimental_option("useAutomationExtension", False)
    opts.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    )
    prefs = {
        "credentials_enable_service": False,
        "profile.password_manager_enabled": False,
        "profile.default_content_setting_values.notifications": 2,
    }
    if download_dir:
        os.makedirs(download_dir, exist_ok=True)
        prefs.update({
            "download.default_directory": download_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "plugins.always_open_pdf_externally": True,
        })
    opts.add_experimental_option("prefs", prefs)
    return opts


def create_driver(headless: bool = False, download_dir: str = "") -> webdriver.Chrome:
    """
    Tạo Chrome WebDriver. webdriver-manager tự tải/cập nhật ChromeDriver.
    Cache tại ~/.wdm/ (hoặc chromedriver_cache_path trong config).
    """
    # Áp dụng cache path từ config nếu có
    try:
        from modules.config import get_config
        _cfg = get_config()
        _cache = _cfg.get("chromedriver_cache_path", "").strip()
        if _cache:
            os.environ.setdefault("WDM_CACHE_PATH", _cache)
        if _cfg.get("chrome_headless", False):
            headless = True
    except Exception:
        pass

    try:
        opts = _build_chrome_options(headless=headless, download_dir=download_dir)
        try:
            driver_path = ChromeDriverManager().install()
        except Exception as e:
            raise RuntimeError(
                f"Khong tai duoc ChromeDriver. Kiem tra ket noi internet va Google Chrome da duoc cai. "
                f"Chi tiet: {e}"
            ) from e
        service = Service(driver_path)
        driver = webdriver.Chrome(service=service, options=opts)
        driver.set_page_load_timeout(PAGE_LOAD_TIMEOUT)
        driver.execute_script(
            "Object.defineProperty(navigator, 'webdriver', {get: () => undefined})"
        )
        logger.info("Chrome WebDriver started")
        return driver
    except RuntimeError:
        raise
    except WebDriverException as e:
        logger.error(f"WebDriver creation failed: {e}")
        raise RuntimeError(f"Khong khoi dong duoc Chrome. Kiem tra Chrome da duoc cai dat. Chi tiet: {e}") from e


# ── Utility helpers ───────────────────────────────────────────────────────────

def _wait_page_ready(driver: webdriver.Chrome, timeout: int = 25) -> None:
    """Chờ document.readyState == 'complete'."""
    WebDriverWait(driver, timeout).until(
        lambda d: d.execute_script("return document.readyState") == "complete"
    )


def _js_click(driver: webdriver.Chrome, el) -> None:
    """JS click — tránh ElementClickIntercepted và StaleElement."""
    driver.execute_script("arguments[0].click();", el)


def _find_element(driver: webdriver.Chrome, candidates: list) -> Optional[object]:
    """
    Thử nhiều (By, selector) candidates, trả về element đầu tiên tìm thấy và visible.
    Bắt cả StaleElementReferenceException.
    """
    for by, value in candidates:
        try:
            el = driver.find_element(by, value)
            if el.is_displayed():
                return el
        except (NoSuchElementException, StaleElementReferenceException):
            continue
    return None


def _js_find_by_exact_text(driver: webdriver.Chrome, tags: str, text: str):
    """Tìm element có innerText chính xác = text (visible). tags: CSS selector string."""
    return driver.execute_script("""
        var text = arguments[0], tags = arguments[1];
        var all = document.querySelectorAll(tags);
        for (var i = 0; i < all.length; i++) {
            var t = (all[i].innerText || all[i].textContent || '').trim();
            if (t === text && all[i].offsetParent !== null) return all[i];
        }
        return null;
    """, text, tags)


def _js_find_links_exact(driver: webdriver.Chrome, labels: list) -> list:
    """
    Tìm TẤT CẢ thẻ <a> có innerText chính xác khớp bất kỳ label nào trong danh sách.
    Multi-label query trong 1 JS call — nhanh hơn gọi nhiều lần.
    """
    return driver.execute_script("""
        var labels = arguments[0];
        var result = [];
        var links = document.querySelectorAll('a');
        for (var i = 0; i < links.length; i++) {
            var el = links[i];
            var t = (el.innerText || el.textContent || '').trim();
            for (var j = 0; j < labels.length; j++) {
                if (t === labels[j] && el.offsetParent !== null) {
                    result.push(el);
                    break;
                }
            }
        }
        return result;
    """, labels) or []


def _wait_for_element_text(driver: webdriver.Chrome, labels: list, timeout: int = 30):
    """
    WebDriverWait (500ms poll) cho đến khi tìm thấy element với text khớp labels.
    Nhanh hơn polling loop 1s.
    """
    try:
        WebDriverWait(driver, timeout, poll_frequency=0.5).until(
            lambda d: _js_find_by_exact_text(
                d, "a, li, td, span, div, p, button",
                # Thử từng label — dùng label đầu tiên tìm thấy
                next((lb for lb in labels
                      if _js_find_by_exact_text(d, "a, li, td, span, div, p, button", lb)), "")
            ) if any(
                _js_find_by_exact_text(d, "a, li, td, span, div, p, button", lb)
                for lb in labels
            ) else None
        )
    except TimeoutException:
        pass
    # Return first match
    for lb in labels:
        el = _js_find_by_exact_text(driver, "a, li, td, span, div, p, button", lb)
        if el:
            return el
    return None


# ── Login ─────────────────────────────────────────────────────────────────────

def _switch_login_page_to_english(driver: webdriver.Chrome) -> bool:
    """
    Đổi ngôn ngữ sang English NGAY TRÊN TRANG LOGIN (trước khi điền form).
    Trang login có dropdown "Korean ▼" ở góc trên phải.
    Không cần reload thêm — chỉ click dropdown → click English.
    """
    try:
        lang_trigger_candidates = [
            (By.XPATH, "//a[contains(text(),'Korean') or contains(text(),'한국어')]"),
            (By.XPATH, "//span[contains(text(),'Korean') or contains(text(),'한국어')]"),
            (By.CSS_SELECTOR, "a.lang_select, .language-select, .lang-select"),
            (By.XPATH, "//select[option[normalize-space(text())='English']]"),
            (By.XPATH, "//*[contains(@class,'lang')]//a"),
        ]
        trigger = _find_element(driver, lang_trigger_candidates)
        if trigger:
            _js_click(driver, trigger)
            logger.debug("Clicked language dropdown")
            # Chờ English option xuất hiện (max 3s)
            try:
                WebDriverWait(driver, 3).until(
                    EC.visibility_of_element_located(
                        (By.XPATH, "//a[normalize-space(text())='English']")
                    )
                )
            except TimeoutException:
                pass

        english_candidates = [
            (By.XPATH, "//a[normalize-space(text())='English']"),
            (By.XPATH, "//li[normalize-space(text())='English']"),
            (By.XPATH, "//option[normalize-space(text())='English']"),
            (By.XPATH, "//*[@data-lang='EN' or @data-lang='en' or @value='EN']"),
        ]
        eng_el = _find_element(driver, english_candidates)
        if not eng_el:
            logger.debug("English option not found on login page — proceeding with default language")
            return False

        _js_click(driver, eng_el)
        # Chờ page re-render nhãn form bằng tiếng Anh (không reload — chỉ label swap)
        try:
            WebDriverWait(driver, 4).until(
                lambda d: "English" in (d.page_source or "")
            )
        except TimeoutException:
            pass
        logger.info("Login page: switched to English")
        return True

    except Exception as e:
        logger.debug(f"_switch_login_page_to_english: {e}")
        return False


def login(
    driver: webdriver.Chrome,
    username: str = DEFAULT_USERNAME,
    password: str = DEFAULT_PASSWORD,
) -> dict:
    """
    Đăng nhập vào Samsung BQMS Vendor Portal.
    Tự động switch ngôn ngữ sang English trên trang login.

    Returns:
        dict: {success, message, page_title}
    """
    result = {"success": False, "message": "", "page_title": ""}

    username_candidates = [
        (By.ID, "userId"), (By.ID, "username"), (By.ID, "loginId"), (By.ID, "vendorId"),
        (By.NAME, "userId"), (By.NAME, "username"), (By.NAME, "loginId"),
        (By.CSS_SELECTOR, "input[type='text']"),
        (By.CSS_SELECTOR, "input[placeholder*='ID'], input[placeholder*='id']"),
        (By.XPATH, "//input[@type='text'][1]"),
    ]
    password_candidates = [
        (By.ID, "password"), (By.ID, "passwd"), (By.ID, "userPwd"),
        (By.NAME, "password"), (By.NAME, "passwd"), (By.NAME, "userPwd"),
        (By.CSS_SELECTOR, "input[type='password']"),
        (By.XPATH, "//input[@type='password'][1]"),
    ]
    submit_candidates = [
        # Samsung BQMS arrow button (→) — specific selectors only, tránh match "Login Guide"
        (By.CSS_SELECTOR, "a.btn_arr, a.arrow_btn, button.btn_arr"),
        (By.XPATH, "//a[contains(@class,'arr')]"),
        (By.XPATH, "//a[img[contains(@src,'arrow') or contains(@src,'btn')]]"),
        (By.XPATH, "//button[img[contains(@src,'arrow') or contains(@src,'btn')]]"),
        (By.XPATH, "//input[@type='image']"),
        (By.CSS_SELECTOR, "button[type='submit'], input[type='submit']"),
    ]

    try:
        logger.info(f"Navigating to: {PORTAL_URL}")
        driver.get(PORTAL_URL)

        # Chờ form login (không sleep cứng)
        try:
            WebDriverWait(driver, ELEMENT_WAIT).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='password']"))
            )
        except TimeoutException:
            logger.warning("Login form not at portal URL, trying direct login URL...")
            driver.get(LOGIN_URL)
            try:
                WebDriverWait(driver, ELEMENT_WAIT).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "input[type='password']"))
                )
            except TimeoutException:
                result["message"] = "Login form not found after 15s."
                return result

        logger.info(f"Login page URL: {driver.current_url}")

        # STEP 1: Fill username
        user_el = _find_element(driver, username_candidates)
        if not user_el:
            result["message"] = "Username field not found."
            return result
        user_el.clear()
        user_el.send_keys(username)

        # STEP 2: Fill password
        pass_el = _find_element(driver, password_candidates)
        if not pass_el:
            result["message"] = "Password field not found."
            return result
        pass_el.clear()
        pass_el.send_keys(password)

        time.sleep(0.3)  # Minimal delay before submit

        # STEP 3: Submit
        submit_el = _find_element(driver, submit_candidates)
        if submit_el:
            submit_el.click()
            logger.info("Clicked submit button")
        else:
            logger.warning("Submit button not found — pressing Enter")
            pass_el.send_keys("\n")

        # STEP 4: Wait for login result (event-driven, không sleep cứng)
        login_url = driver.current_url
        try:
            WebDriverWait(driver, LOGIN_VERIFY_TIMEOUT).until(
                EC.any_of(
                    EC.url_changes(login_url),
                    EC.invisibility_of_element_located(
                        (By.CSS_SELECTOR, "input[type='password']")
                    ),
                )
            )
        except TimeoutException:
            pass  # Kiểm tra thủ công bên dưới

        post_url = driver.current_url
        page_title = driver.title
        result["page_title"] = page_title

        # Kiểm tra lỗi login
        error_selectors = [
            (By.CSS_SELECTOR, ".error-message, .alert-danger, .login-error"),
            (By.XPATH, "//*[contains(text(),'Invalid') or contains(text(),'incorrect')]"),
            (By.XPATH, "//*[contains(text(),'틀린') or contains(text(),'오류')]"),
        ]
        for by, val in error_selectors:
            try:
                el = driver.find_element(by, val)
                if el.is_displayed():
                    result["message"] = f"Login error: {el.text.strip()}"
                    return result
            except NoSuchElementException:
                pass

        # Xác nhận thành công: URL không còn chứa login/anonymous
        if "login" not in post_url.lower() and "anonymous" not in post_url.lower():
            result["success"] = True
            result["message"] = f"Login success. Page: {page_title}"
            logger.info(f"Login success. URL: {post_url}")
        else:
            # Kiểm tra password field còn không
            try:
                driver.find_element(By.CSS_SELECTOR, "input[type='password']")
                result["message"] = f"Login failed. Still on: {post_url}"
            except NoSuchElementException:
                # Password field mất → đã login dù URL chưa redirect
                result["success"] = True
                result["message"] = f"Login success (password gone). Page: {page_title}"
                logger.info(f"Login success (password field gone). URL: {post_url}")

        return result

    except WebDriverException as e:
        result["message"] = f"WebDriver error: {str(e)[:200]}"
        logger.error(f"Login WebDriver error: {e}")
        return result
    except Exception as e:
        result["message"] = f"Unexpected error: {str(e)[:200]}"
        logger.error(f"Login unexpected error: {e}")
        return result


def check_session_alive(driver: webdriver.Chrome) -> bool:
    """Kiểm tra session còn sống: không bị redirect về login và còn element portal."""
    try:
        url = driver.current_url
        if "login" in url.lower() or "anonymous" in url.lower():
            return False
        # Kiểm tra có element portal thực sự không (không chỉ check body)
        driver.find_element(By.CSS_SELECTOR, "body")
        return "vendorPortal" in url or "bqms" in url
    except WebDriverException:
        return False


def login_with_retry(
    driver: webdriver.Chrome,
    username: str = DEFAULT_USERNAME,
    password: str = DEFAULT_PASSWORD,
    max_retries: int = 3,
    retry_delay: int = 10,
) -> dict:
    """Login với retry, tối đa max_retries lần để tránh bị khóa IP."""
    for attempt in range(1, max_retries + 1):
        logger.info(f"Login attempt {attempt}/{max_retries}")
        result = login(driver, username, password)
        if result["success"]:
            return result
        logger.warning(f"Attempt {attempt} failed: {result['message']}")
        if attempt < max_retries:
            logger.info(f"Waiting {retry_delay}s before retry...")
            time.sleep(retry_delay)

    result["message"] = f"Login failed after {max_retries} attempts. {result['message']}"
    return result


# ── Post-Login Popup Handler ──────────────────────────────────────────────────

def handle_post_login_popups(driver: webdriver.Chrome) -> None:
    """
    Xử lý popup thông báo tiếng Hàn xuất hiện sau login (cửa sổ riêng).
    - Tick checkbox "오늘 하루 이 페이지 열지 않음" (Don't show today)
    - Click Close / window tự đóng sau tick
    """
    main_window = driver.current_window_handle

    # Chờ popup window xuất hiện (tối đa 3s, không sleep cứng)
    try:
        WebDriverWait(driver, 3).until(lambda d: len(d.window_handles) > 1)
    except TimeoutException:
        logger.info("No popup windows within 3s")
        return

    handles = driver.window_handles
    logger.info(f"Detected {len(handles) - 1} popup window(s)")

    checkbox_candidates = [
        (By.CSS_SELECTOR, "input[type='checkbox']"),
        (By.XPATH, "//input[@type='checkbox']"),
        (By.XPATH, "//label[contains(text(),'오늘')]/..//input[@type='checkbox']"),
    ]
    close_candidates = [
        (By.XPATH, "//a[normalize-space(text())='Close' or normalize-space(text())='close']"),
        (By.XPATH, "//button[normalize-space(text())='Close']"),
        (By.XPATH, "//a[contains(text(),'닫기')] | //button[contains(text(),'닫기')]"),
        (By.CSS_SELECTOR, ".btn_close, a.close, button.close"),
        (By.XPATH, "//a[contains(@class,'close')] | //button[contains(@class,'close')]"),
    ]

    for handle in handles:
        if handle == main_window:
            continue
        try:
            driver.switch_to.window(handle)
            logger.info(f"Popup URL: {driver.current_url[:80]}")

            # Tick checkbox
            cb = _find_element(driver, checkbox_candidates)
            if cb and not cb.is_selected():
                _js_click(driver, cb)
                logger.info("Ticked 'Don't show today' checkbox")
                # Chờ window tự đóng sau tick (max 1s)
                try:
                    WebDriverWait(driver, 1).until(EC.staleness_of(cb))
                except TimeoutException:
                    pass

            # Kiểm tra window còn tồn tại
            try:
                _ = driver.current_url
            except WebDriverException:
                logger.info("Popup auto-closed after checkbox — OK")
                continue

            # Click Close nếu window vẫn còn
            close_btn = _find_element(driver, close_candidates)
            if close_btn:
                _js_click(driver, close_btn)
                logger.info("Clicked Close button")
                try:
                    WebDriverWait(driver, 2).until(
                        lambda d: handle not in d.window_handles
                    )
                except TimeoutException:
                    pass
            else:
                driver.close()

        except Exception as e:
            logger.debug(f"Popup handling: {e}")

    driver.switch_to.window(main_window)
    logger.info("All popups handled")


# ── Navigation ────────────────────────────────────────────────────────────────

def navigate_to_po_receipt_mro(driver: webdriver.Chrome) -> bool:
    """
    Điều hướng đến Execution (발주) > MRO > P/O Receipt (발주접수).

    Chiến lược:
    1. Thử navigate thẳng đến URL đã biết (nhanh nhất)
    2. Fallback: click qua menu
    """
    # ── Click menu Execution > MRO > P/O Receipt ─────────────────────────────
    # Portal yêu cầu navigate qua menu (session state) — không dùng direct URL
    try:
        _wait_page_ready(driver)

        # Chờ menu render (WebDriverWait 500ms poll, multi-label JS)
        logger.info(f"Waiting for Execution menu (labels: {EXECUTION_LABELS})...")
        exec_el = None
        try:
            WebDriverWait(driver, 30, poll_frequency=0.5).until(
                lambda d: any(
                    _js_find_by_exact_text(d, "a, li, td, span, div, p, button", lb)
                    for lb in EXECUTION_LABELS
                )
            )
            for lb in EXECUTION_LABELS:
                exec_el = _js_find_by_exact_text(driver, "a, li, td, span, div, p, button", lb)
                if exec_el:
                    logger.info(f"Found Execution menu: '{lb}'")
                    break
        except TimeoutException:
            page_text = driver.execute_script(
                "return document.body ? document.body.innerText.substring(0, 2000) : ''"
            )
            logger.error(f"Execution menu not found after 30s. Page text:\n{page_text}")
            return False

        # Hover + click Execution
        try:
            ActionChains(driver).move_to_element(exec_el).perform()
            time.sleep(0.3)
        except Exception:
            pass
        _js_click(driver, exec_el)
        logger.info("Clicked Execution")

        # Chờ submenu P/O Receipt xuất hiện (WebDriverWait, không sleep cứng)
        try:
            WebDriverWait(driver, 5, poll_frequency=0.2).until(
                lambda d: len(_js_find_links_exact(d, PO_RECEIPT_LABELS)) > 0
            )
        except TimeoutException:
            logger.warning("P/O Receipt not visible after 5s — trying anyway")

        # Tìm 발주접수 link nằm SAU "MRO" header theo thứ tự DOM
        po_el = driver.execute_script("""
            var labels = arguments[0];
            var allEls = document.querySelectorAll('*');
            var mroIdx = -1, nextSectionIdx = allEls.length;

            // Tìm vị trí MRO header (visible)
            for (var i = 0; i < allEls.length; i++) {
                var t = (allEls[i].innerText || allEls[i].textContent || '').trim();
                if (mroIdx === -1 && t === 'MRO' && allEls[i].offsetParent !== null) {
                    mroIdx = i;
                } else if (mroIdx !== -1 && (t === 'B2B' || t === '일반') && allEls[i].offsetParent !== null) {
                    nextSectionIdx = i;
                    break;
                }
            }

            if (mroIdx === -1) return null;

            // Tìm <a> có text 발주접수 giữa mroIdx và nextSectionIdx
            for (var i = mroIdx; i < nextSectionIdx; i++) {
                if (allEls[i].tagName === 'A' && allEls[i].offsetParent !== null) {
                    var lt = (allEls[i].innerText || allEls[i].textContent || '').trim();
                    for (var k = 0; k < labels.length; k++) {
                        if (lt === labels[k]) return allEls[i];
                    }
                }
            }
            return null;
        """, PO_RECEIPT_LABELS)

        if not po_el:
            logger.error("MRO P/O Receipt link not found after MRO header")
            return False

        href = driver.execute_script("return arguments[0].href || '';", po_el)
        logger.info(f"Clicking P/O Receipt MRO: href='{href}'")
        _js_click(driver, po_el)

        # Chờ trang load (event-driven)
        _wait_page_ready(driver)
        logger.info(f"P/O Receipt page: {driver.title} | {driver.current_url}")
        return True

    except Exception as e:
        logger.error(f"navigate_to_po_receipt_mro error: {e}")
        return False


# ── Full Flow ─────────────────────────────────────────────────────────────────

def full_login_and_navigate(
    headless: bool = False,
    download_dir: str = "",
) -> tuple:
    """
    Full flow: tạo driver → login (với switch EN) → handle popup → navigate to P/O Receipt.

    Returns:
        (driver, success: bool) — driver vẫn mở để caller dùng tiếp.
        Caller phải tự gọi driver.quit() sau khi xong.
    """
    driver = create_driver(headless=headless, download_dir=download_dir)
    try:
        result = login_with_retry(driver)
        if not result["success"]:
            logger.error(f"Login failed: {result['message']}")
            return driver, False

        handle_post_login_popups(driver)
        ok = navigate_to_po_receipt_mro(driver)
        return driver, ok

    except Exception as e:
        logger.error(f"full_login_and_navigate error: {e}")
        return driver, False


# ── Scrape PO Table ───────────────────────────────────────────────────────────

def scrape_po_table(driver: webdriver.Chrome) -> list:
    """
    Đọc toàn bộ bảng PO trên trang hiện tại.
    Trả về list of dict, mỗi dict là 1 dòng PO.
    """
    try:
        # Chờ table xuất hiện
        WebDriverWait(driver, 15).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table"))
        )
        _wait_page_ready(driver)

        rows = driver.execute_script("""
            var tables = document.querySelectorAll('table');
            // Tìm table có nhiều td nhất trong tbody (data table)
            var dataTable = null, maxCols = 0;
            for (var t = 0; t < tables.length; t++) {
                var tds = tables[t].querySelectorAll('tbody td');
                if (tds.length > maxCols) { maxCols = tds.length; dataTable = tables[t]; }
            }
            if (!dataTable) return [];

            var allRows = dataTable.querySelectorAll('tbody tr');
            if (allRows.length === 0) return [];

            // Row đầu tiên làm header (th thường rỗng trong portal này)
            var headers = [];
            var firstCells = allRows[0].querySelectorAll('td');
            for (var i = 0; i < firstCells.length; i++) {
                var h = (firstCells[i].innerText || firstCells[i].textContent || '').trim();
                headers.push(h || ('col_' + i));
            }

            // Row 1 trở đi là data
            var result = [];
            for (var r = 1; r < allRows.length; r++) {
                var cells = allRows[r].querySelectorAll('td');
                if (cells.length === 0) continue;
                var row = {};
                for (var c = 0; c < cells.length; c++) {
                    var key = headers[c] || ('col_' + c);
                    row[key] = (cells[c].innerText || cells[c].textContent || '').trim();
                }
                result.push(row);
            }
            return result;
        """)
        logger.info(f"Scraped {len(rows)} PO rows")
        return rows or []

    except Exception as e:
        logger.error(f"scrape_po_table error: {e}")
        return []


# ── Test ──────────────────────────────────────────────────────────────────────

# ── Pagination scrape ─────────────────────────────────────────────────────────

def scrape_all_pages(driver: webdriver.Chrome) -> list:
    """
    Lấy toàn bộ PO bằng cách chọn pageSize=99999 (전체 = All).
    Sau đó scrape một lần duy nhất.
    """
    try:
        # Đọc total count
        total = driver.execute_script(
            "var s = document.getElementById('totalCount'); return s ? parseInt(s.innerText) : 0;"
        ) or 0
        logger.info(f"Total records: {total}")

        if total <= 10:
            # Chỉ 1 trang, không cần thay đổi
            return scrape_po_table(driver)

        # Chọn pageSize = 99999 (전체)
        prev_count = driver.execute_script("""
            var rows = document.querySelectorAll('tbody tr');
            return rows.length;
        """)
        driver.execute_script("""
            var sel = document.getElementById('pageSize');
            if (sel) {
                sel.value = '99999';
                sel.dispatchEvent(new Event('change', {bubbles: true}));
            }
        """)
        # Chờ table reload (số rows thay đổi)
        try:
            WebDriverWait(driver, 10, poll_frequency=0.5).until(
                lambda d: d.execute_script(
                    "var r = document.querySelectorAll('tbody tr'); return r.length;"
                ) != prev_count
            )
        except TimeoutException:
            logger.warning("pageSize change timeout — scraping current page")

        _wait_page_ready(driver)
        rows = scrape_po_table(driver)
        logger.info(f"Scraped {len(rows)} rows (all pages)")
        return rows

    except Exception as e:
        logger.error(f"scrape_all_pages error: {e}")
        return scrape_po_table(driver)


# ── Excel Writer ──────────────────────────────────────────────────────────────

def _get_excel_path():
    home = os.path.expanduser("~")
    for name in os.listdir(home):
        if "SONG CHAU" in name.upper() and os.path.isdir(os.path.join(home, name)):
            p = os.path.join(home, name, "Puplic", "BQMS", "Thong ke giao hang", "Thong ke giao hang 2026.xlsx")
            if os.path.exists(p):
                return p
    return os.path.join(home, "OneDrive - SONG CHAU CO., LTD", "Puplic", "BQMS", "Thong ke giao hang", "Thong ke giao hang 2026.xlsx")

EXCEL_PATH = _get_excel_path()
STATE_FILE = os.path.join(os.path.dirname(__file__), "excel_state.json")

# Mapping portal col → Excel column index (1-based)
# Selenium scrape field names (col_X)
EXCEL_COL_MAP = {
    1:  "col_4",    # A  Ngày PO
    2:  "col_5",    # B  Số PO
    # 3  C  Shipping No — bỏ trống
    4:  "col_7",    # D  Số QT
    5:  "col_21",   # E  BQMS/vendor code (Z-code)
    # 6  F  Spec → từ PDF
    # 7  G  SL   → từ PDF
    # 8  H  ĐV   → từ PDF
    # 9  I  Đơn giá → từ PDF
    # 10 J  Thành tiền → từ PDF
    11: "col_16",   # K  SEV/T (company abbreviation)
    12: None,       # L  MAIL PUR → từ col_29, xử lý riêng
}

# VLOOKUP range cho M (tên) và N (kho) từ sheet DANH BẠ
_DANH_BA_RANGE  = "'DANH BẠ'!A$3:D$228"
_VLOOKUP_M      = "=VLOOKUP(L{row}," + _DANH_BA_RANGE + ",2,0)"
_VLOOKUP_N      = "=VLOOKUP(L{row}," + _DANH_BA_RANGE + ",3,0)"


def _format_po_date(val) -> str:
    """
    Chuyển date → 'm/d/yy' (không có số 0 trước tháng/ngày).
    '03/20/2026' hoặc date object → '3/20/26'
    """
    from datetime import datetime, date as _date
    try:
        if isinstance(val, _date):
            d = val
        else:
            s = str(val).strip()
            for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y"):
                try:
                    d = datetime.strptime(s, fmt).date()
                    break
                except ValueError:
                    continue
            else:
                return str(val)
        return "%d/%d/%02d" % (d.month, d.day, d.year % 100)
    except Exception:
        return str(val)


def _load_excel_state() -> dict:
    """Load state JSON: {last_row, last_po_date, last_written_at}"""
    try:
        import json
        with open(STATE_FILE, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_excel_state(last_row: int, last_po_date: str, added: int):
    """Lưu state JSON sau mỗi lần ghi thành công."""
    import json
    from datetime import datetime
    state = {
        "last_row":       last_row,
        "last_po_date":   last_po_date,
        "last_added":     added,
        "last_written_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
    }
    try:
        with open(STATE_FILE, "w", encoding="utf-8") as f:
            json.dump(state, f, ensure_ascii=False, indent=2)
        logger.info("State saved: row=%d date=%s added=%d", last_row, last_po_date, added)
    except Exception as e:
        logger.warning("State save failed: %s", e)


def _find_insert_row(ws, gap_threshold: int = 5) -> int:
    """
    Gap-aware insertion point finder.
    Scan từ trên xuống, dừng khi gặp gap >= gap_threshold dòng B rỗng liên tiếp.
    Trả về last_real_data_row + 1.

    Ví dụ: rows 2-180 có data, rows 181-1905 là formula-only (B rỗng):
      → last_real=180, gap detected at 181+4 → return 181.
    Tránh nhầm với rows 1906-1912 có data sai vị trí.
    """
    last_real = 1
    consecutive_empty = 0
    for r in range(2, ws.max_row + 1):
        v = ws.cell(r, 2).value
        if v and str(v).strip() not in ("", "None"):
            last_real = r
            consecutive_empty = 0
        else:
            consecutive_empty += 1
            if consecutive_empty >= gap_threshold:
                break
    return last_real + 1


def _parse_number(val: str):
    """'158,000.0000' → 158000  |  '63,200,000.00' → 63200000"""
    try:
        clean = val.replace(",", "").split(".")[0]
        return int(clean) if clean else None
    except Exception:
        return val


def _parse_date(val: str):
    """'03/16/2026' → date object"""
    try:
        from datetime import datetime
        return datetime.strptime(val, "%m/%d/%Y").date()
    except Exception:
        return val


# ── PDF Parser ────────────────────────────────────────────────────────────────

def _parse_num_pdf(val: str):
    """'10,820,000.00' → 10820000  (float nếu có thập phân thực, int nếu .00)"""
    if not val:
        return None
    try:
        clean = str(val).replace(",", "").strip()
        f = float(clean)
        return int(f) if f == int(f) else f
    except Exception:
        return val


def parse_po_pdf(pdf_path: str) -> dict:
    """
    Parse Samsung Purchase Order PDF.
    Cấu trúc thực tế: 1 table, toàn bộ product data nằm trong col[0] dưới dạng multiline.
    Mỗi product row có dạng:
      Line 0: Q230-187524                              ← BQMS Code
      Line 1: CONVEYOR... 03/27/2026 SET 10,820,000.00 ← Desc + RDDate + UoM + UnitPrice
      Line 2: [Z0000002-... SPEC... Receiver 84.00 908,880,000.00 ← VendorCode+Spec+Qty+Amount
      Line 3: 467818] SPEC_CONTINUED                  ← (tiếp vendor code + spec)
    """
    try:
        import pdfplumber
    except ImportError:
        logger.error("pdfplumber not installed. Run: pip install pdfplumber")
        return {}

    import re

    result = {"po_no": "", "po_date": "", "products": [], "total_qty": None, "total_amount": None}

    try:
        with pdfplumber.open(pdf_path) as pdf:
            full_text = ""
            all_tables = []
            for page in pdf.pages:
                full_text += (page.extract_text() or "") + "\n"
                tbls = page.extract_tables()
                if tbls:
                    all_tables.extend(tbls)

        # ── Header fields ──
        m = re.search(r"P/O\s*No\.?\s*[:\.]?\s*(\d{7,15})", full_text)
        if m:
            result["po_no"] = m.group(1).strip()

        m = re.search(r"P/O\s*Date\.?\s*[:\.]?\s*([\d/]+)", full_text)
        if m:
            result["po_date"] = m.group(1).strip()

        m = re.search(r"Total\s+Quantity\D+([\d,]+)", full_text)
        if m:
            result["total_qty"] = _parse_num_pdf(m.group(1))

        m = re.search(r"Total\s+Amount[^0-9]+([\d,]+\.\d{2})", full_text)
        if m:
            result["total_amount"] = _parse_num_pdf(m.group(1))

        # ── Product rows: tìm table lớn nhất ──
        if not all_tables:
            logger.warning(f"parse_po_pdf: no tables in {pdf_path}")
            return result

        product_table = max(all_tables, key=lambda t: len(t))

        products = []
        for row in product_table:
            if not row or not row[0]:
                continue
            cell = str(row[0]).strip()

            # Product row: line[0] bắt đầu bằng mã kiểu Q230-187524
            lines = cell.split("\n")
            if not lines or not re.match(r"^[A-Z]\d{3}-\d+", lines[0].strip()):
                continue

            code = lines[0].strip()

            # Line 1: "DESCRIPTION RD_DATE UOM UNIT_PRICE"
            line1 = lines[1].strip() if len(lines) > 1 else ""

            # Extract RD Date
            rd_match = re.search(r"\d{2}/\d{2}/\d{4}", line1)
            rd_date = rd_match.group() if rd_match else ""

            # Extract Unit Price (money number cuối line1: có dấu phẩy ngăn cách)
            money_l1 = re.findall(r"[\d,]+,\d{3}\.\d{2}", line1)
            unit_price = _parse_num_pdf(money_l1[-1]) if money_l1 else None

            # UoM và Description từ line1 (bỏ date và unit_price)
            l1_clean = line1
            if rd_date:
                l1_clean = l1_clean.replace(rd_date, "")
            for n in money_l1:
                l1_clean = l1_clean.replace(n, "")
            l1_parts = l1_clean.split()
            uom = l1_parts[-1] if l1_parts else ""
            description = " ".join(l1_parts[:-1]) if len(l1_parts) > 1 else (l1_parts[0] if l1_parts else "")

            # Lines 2+: vendor_code + spec + receiver + qty + amount
            rest = "\n".join(lines[2:]) if len(lines) > 2 else ""

            # Amount (money number cuối rest)
            money_rest = re.findall(r"[\d,]+,\d{3}\.\d{2}", rest)
            amount = _parse_num_pdf(money_rest[-1]) if money_rest else None

            # Qty: số ngay TRƯỚC amount — tránh nhầm "20.45mm" trong spec
            qty = None
            qty_raw = ""
            if money_rest:
                m_qty = re.search(
                    r"(?<!\w)([\d]+(?:\.\d+)?)\s+" + re.escape(money_rest[-1]),
                    rest
                )
                if m_qty:
                    qty_raw = m_qty.group(1)
                    qty = _parse_num_pdf(qty_raw)

            # Spec: bỏ vendor code, numbers, dọn dẹp
            spec_text = rest
            spec_text = re.sub(r"\[Z\d+[-\s]*\n?\s*\d*\]?", " ", spec_text)
            spec_text = re.sub(r"\d+\]", " ", spec_text)
            spec_text = re.sub(r"\d{2}/\d{2}/\d{4}", " ", spec_text)
            for n in money_rest:
                spec_text = spec_text.replace(n, " ")
            # Capture + bỏ tên người nhận (Receiver): 2-4 từ ngay trước qty
            receiver = ""
            if qty_raw:
                m_recv = re.search(
                    r"((?:\S+\s+){2,4})" + re.escape(qty_raw) + r"(?!\d)",
                    spec_text
                )
                if m_recv:
                    receiver = m_recv.group(1).strip()
                spec_text = re.sub(
                    r"(?:\S+\s+){2,4}" + re.escape(qty_raw) + r"(?!\d)",
                    " ", spec_text
                )
                # Bỏ qty nếu còn sót
                spec_text = re.sub(r"(?<!\w)" + re.escape(qty_raw) + r"(?!\w)", " ", spec_text)
            spec_text = " ".join(spec_text.split()).strip()

            # Bỏ nội dung trong ngoặc (PR_NO. xxxx) ở description
            desc_clean = re.sub(r"\([^)]*\)", "", description).strip()
            desc_clean = " ".join(desc_clean.split())

            # Format: "DESCRIPTION\nSPEC" (xuống dòng giữa mô tả và spec)
            if desc_clean and spec_text:
                full_spec = desc_clean + "\n" + spec_text
            else:
                full_spec = (desc_clean + " " + spec_text).strip()

            prod = {
                "code":       code,
                "spec":       full_spec,
                "rd_date":    rd_date,
                "receiver":   receiver,
                "uom":        uom,
                "qty":        qty,
                "unit_price": unit_price,
                "amount":     amount,
            }
            if prod["code"]:
                products.append(prod)

        result["products"] = products
        logger.info(f"PDF parsed: po_no={result['po_no']}, {len(products)} products")
        return result

    except Exception as e:
        logger.error(f"parse_po_pdf error ({pdf_path}): {e}")
        return result


def _strip_diacritics(s: str) -> str:
    """'Tùng' → 'tung', 'Thúy' → 'thuy' (lowercase + bỏ dấu)"""
    import unicodedata
    return "".join(
        c for c in unicodedata.normalize("NFD", s.lower())
        if unicodedata.category(c) != "Mn"
    )


def _load_danh_ba(excel_path: str = EXCEL_PATH) -> dict:
    """
    Load sheet DANH BẠ → {given_name_no_diacritics: mail_prefix}.
    Key = tên riêng (từ cuối) đã bỏ dấu — để match với PDF (thường mất dấu).
    """
    import openpyxl
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        ws = wb["DANH BẠ"]
        result = {}
        for row in ws.iter_rows(min_row=4, values_only=True):  # rows 1-3 là header
            mail = str(row[0]).strip() if row[0] else ""
            ten  = str(row[1]).strip() if row[1] else ""
            if mail and ten and mail != "None":
                given_key = _strip_diacritics(ten.split()[-1])  # bỏ dấu làm key
                result[given_key] = mail
        wb.close()
        logger.info(f"Danh Ba loaded: {len(result)} entries")
        return result
    except Exception as e:
        logger.warning(f"_load_danh_ba error: {e}")
        return {}


def _lookup_receiver_mail(receiver: str, danh_ba: dict) -> str:
    """
    Từ tên receiver trong PDF (vd 'Nguyn Sn Tùng' hoặc 'Nguyen Son Tung')
    → tìm mail prefix trong Danh Bạ.
    Strategy: lấy từ cuối → bỏ dấu → lookup (cả 2 phía đều bỏ dấu → match tốt).
    """
    if not receiver or not danh_ba:
        return ""
    words = receiver.strip().split()
    if not words:
        return ""
    given_key = _strip_diacritics(words[-1])
    return danh_ba.get(given_key, "")


def _write_via_local_copy(
    rows_sorted: list,
    pdf_data_map: dict,
    excel_path: str,
    existing: set,
    next_row: int,
) -> tuple:
    """
    Ghi data vào file tạm (ngoài OneDrive) → copy đè lại file gốc.
    Tránh hoàn toàn xung đột OneDrive sync trong quá trình ghi.
    Returns: (added, skipped, next_row_after)
    """
    import shutil, tempfile, openpyxl, unicodedata

    def _n(s):
        return "".join(c for c in unicodedata.normalize("NFD", s.upper())
                       if unicodedata.category(c) != "Mn")

    def _get_ws(wb):
        for name in wb.sheetnames:
            if any(kw in _n(name) for kw in ["THONG KE PO", "KE PO"]):
                return wb[name]
        return wb.active

    # Copy file sang thư mục temp LOCAL (ngoài OneDrive)
    tmp_dir = tempfile.gettempdir()
    tmp_path = os.path.join(tmp_dir, "bsmq_po_write_temp.xlsx")
    shutil.copy2(excel_path, tmp_path)
    logger.info("Copied to temp: %s", tmp_path)

    wb = openpyxl.load_workbook(tmp_path)
    ws = _get_ws(wb)

    added = 0
    skipped = 0
    po_product_idx: dict = {}

    for row in rows_sorted:
        so_po = row.get("col_5", "").strip()
        so_qt = row.get("col_7", "").strip()
        if not so_po:
            continue
        if (so_po, so_qt) in existing:
            skipped += 1
            po_product_idx[so_po] = po_product_idx.get(so_po, 0) + 1
            continue

        ws.cell(next_row, 1, _format_po_date(row.get("col_4", "")))
        ws.cell(next_row, 2, so_po)
        ws.cell(next_row, 4, so_qt)
        ws.cell(next_row, 5, row.get("col_21", "").strip())

        sevt = row.get("col_16", "").strip()
        if sevt:
            ws.cell(next_row, 11, sevt)

        recv_raw = row.get("col_29", "").strip()
        mail = recv_raw.split(":")[-1].strip() if ":" in recv_raw else ""
        if mail:
            ws.cell(next_row, 12, mail)

        ws.cell(next_row, 13, _VLOOKUP_M.format(row=next_row))
        ws.cell(next_row, 14, _VLOOKUP_N.format(row=next_row))

        if pdf_data_map and so_po in pdf_data_map:
            products = pdf_data_map[so_po].get("products", [])
            idx = po_product_idx.get(so_po, 0)
            if idx < len(products):
                prod = products[idx]
                from openpyxl.styles import Alignment
                cell_f = ws.cell(next_row, 6, prod.get("spec", ""))
                cell_f.alignment = Alignment(wrap_text=True)
                ws.cell(next_row, 7, prod.get("qty"))
                ws.cell(next_row, 8, prod.get("uom", ""))
                ws.cell(next_row, 9, prod.get("unit_price"))
                ws.cell(next_row, 10, prod.get("amount"))

        existing.add((so_po, so_qt))
        po_product_idx[so_po] = po_product_idx.get(so_po, 0) + 1
        next_row += 1
        added += 1

    # Save temp file trước
    wb.save(tmp_path)
    wb.close()
    logger.info("Saved temp file, copying back to OneDrive path...")

    # Copy đè về đường dẫn OneDrive (OneDrive sẽ tự sync sau)
    shutil.copy2(tmp_path, excel_path)
    os.remove(tmp_path)
    logger.info("Replaced OneDrive file. OneDrive will sync automatically.")

    return added, skipped, next_row


def write_to_excel(
    rows: list,
    pdf_data_map: dict = None,
    excel_path: str = EXCEL_PATH,
) -> dict:
    """
    Ghi PO rows (từ Selenium scrape) vào Excel qua COM (OneDrive-safe).
    - Sort cũ → mới theo ngày trước khi ghi
    - Dedup theo (Số PO, Số QT)
    - Ghi qua Excel COM → OneDrive tự sync sau khi save
    - Gap-aware insertion row
    - Spec column F: bỏ nội dung trong ngoặc, xuống dòng desc/spec

    Returns: {added, skipped, total, errors}
    """
    import openpyxl
    import unicodedata
    from datetime import datetime

    def _get_target_sheet(wb):
        def _n(s):
            return "".join(c for c in unicodedata.normalize("NFD", s.upper())
                           if unicodedata.category(c) != "Mn")
        for name in wb.sheetnames:
            if any(kw in _n(name) for kw in ["THONG KE PO", "KE PO"]):
                return wb[name]
        return wb.active

    def _sort_key(r):
        try:
            return datetime.strptime(r.get("col_4", "").strip(), "%m/%d/%Y")
        except Exception:
            return datetime.min

    # Sort cũ → mới
    rows_sorted = sorted(rows, key=_sort_key)

    # Đọc existing + insert row qua openpyxl (read-only, nhanh)
    wb_r = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    ws_r = _get_target_sheet(wb_r)
    existing = set()
    for r in ws_r.iter_rows(min_row=2, values_only=True):
        b = str(r[1]).strip() if r[1] else ""
        d = str(r[3]).strip() if r[3] else ""
        if b and b != "None" and d:
            existing.add((b, d))
    next_row = _find_insert_row(ws_r)
    wb_r.close()
    logger.info("Insertion row: %d (existing pairs: %d)", next_row, len(existing))

    # Ghi qua file tạm local → copy về OneDrive (tránh sync conflict)
    try:
        added, skipped, next_row_after = _write_via_local_copy(
            rows_sorted, pdf_data_map or {}, excel_path, existing, next_row
        )
        errors = 0
    except Exception as e:
        logger.error("Write failed: %s", e)
        return {"added": 0, "skipped": 0, "total": len(rows), "errors": 1}

    logger.info("Excel saved (local copy): +%d added, %d skipped", added, skipped)

    if added > 0:
        last_date = _format_po_date(rows_sorted[-1].get("col_4", ""))
        _save_excel_state(next_row_after - 1, last_date, added)

    return {
        "added": added,
        "skipped": skipped,
        "total": len(rows),
        "errors": errors,
    }


# ── PDF Downloader ────────────────────────────────────────────────────────────

# Base folder cho PO PDFs
def _get_po_base_dir():
    home = os.path.expanduser("~")
    for name in os.listdir(home):
        if "SONG CHAU" in name.upper() and os.path.isdir(os.path.join(home, name)):
            p = os.path.join(home, name, "Puplic", "BQMS", "PO")
            if os.path.isdir(p):
                return p
    return os.path.join(home, "OneDrive - SONG CHAU CO., LTD", "Puplic", "BQMS", "PO")

PO_BASE_DIR = _get_po_base_dir()

# Temp folder để Chrome download vào trước khi move
PO_TEMP_DOWNLOAD = os.path.join(os.environ.get("TEMP", os.path.join(os.path.expanduser("~"), "AppData", "Local", "Temp")), "bsmq_po_download")


def _get_po_month_folder(date_str: str, base_dir: str = PO_BASE_DIR) -> str:
    """
    Từ date string '03/16/2026' → trả về path 'PO_BASE_DIR\\PO 2026\\THANG 3'.
    Folder phải tồn tại sẵn — không tạo mới.
    """
    try:
        from datetime import datetime
        dt = datetime.strptime(date_str.strip(), "%m/%d/%Y")
        folder = os.path.join(base_dir, f"PO {dt.year}", f"THANG {dt.month}")
        return folder
    except Exception:
        return base_dir


def _pdf_already_downloaded(po_no: str, month_folder: str) -> bool:
    """
    Kiểm tra xem file PDF cho PO này đã tồn tại trong month_folder chưa.
    Tên file pattern: PurchaseOrder_{po_no}_*.pdf
    """
    if not os.path.isdir(month_folder):
        return False
    for fname in os.listdir(month_folder):
        if po_no in fname and fname.lower().endswith(".pdf"):
            return True
    return False


def _wait_for_download_complete(download_dir: str, timeout: int = 30) -> str:
    """
    Chờ file .pdf xuất hiện trong download_dir (không phải .crdownload).
    Trả về đường dẫn file vừa download, hoặc "" nếu timeout.
    """
    deadline = time.time() + timeout
    seen_before = set(os.listdir(download_dir)) if os.path.isdir(download_dir) else set()

    while time.time() < deadline:
        if not os.path.isdir(download_dir):
            time.sleep(0.5)
            continue
        current = set(os.listdir(download_dir))
        new_files = current - seen_before
        pdf_files = [f for f in new_files if f.lower().endswith(".pdf")]
        if pdf_files:
            # Đảm bảo không còn .crdownload tương ứng
            crdownload = [f for f in current if f.endswith(".crdownload")]
            if not crdownload:
                return os.path.join(download_dir, pdf_files[0])
        time.sleep(0.5)

    # Fallback: tìm file .pdf mới nhất trong folder
    if os.path.isdir(download_dir):
        pdfs = [f for f in os.listdir(download_dir) if f.lower().endswith(".pdf")]
        if pdfs:
            newest = max(pdfs, key=lambda f: os.path.getmtime(os.path.join(download_dir, f)))
            full = os.path.join(download_dir, newest)
            if time.time() - os.path.getmtime(full) < 60:  # Mới trong 60s
                return full
    return ""


def download_po_pdfs(
    driver: webdriver.Chrome,
    rows: list,
    base_po_dir: str = PO_BASE_DIR,
    temp_dir: str = PO_TEMP_DOWNLOAD,
) -> dict:
    """
    Tải PDF cho từng PO row:
    1. Kiểm tra đã có PDF trong month_folder chưa (dedup)
    2. Click blue 발주번호 link → popup attachFilePop.do mở
    3. Chờ 5s → click blue ↓ download arrow
    4. Chờ file download về temp_dir
    5. Move file về đúng month_folder

    Returns:
        dict: {downloaded, skipped, errors, files}
    """
    os.makedirs(temp_dir, exist_ok=True)

    # Cập nhật Chrome download dir sang temp_dir (runtime update)
    driver.execute_cdp_cmd("Page.setDownloadBehavior", {
        "behavior": "allow",
        "downloadPath": temp_dir,
    })

    main_window = driver.current_window_handle
    downloaded = 0
    skipped = 0
    errors = 0
    files_saved = []

    for i, row in enumerate(rows):
        po_no = row.get("col_5", "").strip()
        date_str = row.get("col_4", "").strip()

        if not po_no:
            logger.warning(f"Row {i}: col_5 (P/O No) rỗng — bỏ qua")
            errors += 1
            continue

        month_folder = _get_po_month_folder(date_str, base_po_dir)

        # Dedup check
        if _pdf_already_downloaded(po_no, month_folder):
            logger.info(f"[{i+1}/{len(rows)}] PDF {po_no} đã tồn tại — bỏ qua")
            skipped += 1
            continue

        logger.info(f"[{i+1}/{len(rows)}] Downloading PDF: PO {po_no} → {month_folder}")

        try:
            # Ghi lại files trong temp_dir trước khi click
            files_before = set(os.listdir(temp_dir)) if os.path.isdir(temp_dir) else set()

            # Click blue 발주번호 link trong table (text = po_no)
            po_link = driver.execute_script("""
                var target = arguments[0];
                var links = document.querySelectorAll('table a');
                for (var i = 0; i < links.length; i++) {
                    var t = (links[i].innerText || links[i].textContent || '').trim();
                    if (t === target && links[i].offsetParent !== null) return links[i];
                }
                return null;
            """, po_no)

            if not po_link:
                logger.warning(f"PO link '{po_no}' không tìm thấy trong table")
                errors += 1
                continue

            _js_click(driver, po_link)

            # Chờ popup window mở (tối đa 8s)
            try:
                WebDriverWait(driver, 8).until(lambda d: len(d.window_handles) > 1)
            except TimeoutException:
                logger.warning(f"PO {po_no}: popup không mở sau 8s")
                errors += 1
                continue

            # Switch sang popup window
            popup_handle = [h for h in driver.window_handles if h != main_window][0]
            driver.switch_to.window(popup_handle)
            logger.info(f"Popup URL: {driver.current_url[:100]}")

            # Chờ 5s cho file list load (theo yêu cầu user)
            time.sleep(5)

            # Click blue ↓ download arrow button
            # Samsung portal dùng onclick="dx5DownloadFile(...)" — chữ D hoa
            download_btn_candidates = [
                # dx5DownloadFile onclick (Samsung BQMS specific)
                (By.XPATH, "//a[contains(@onclick,'dx5DownloadFile') and not(contains(text(),'전체'))]"),
                (By.XPATH, "//a[contains(@onclick,'DownloadFile') and not(contains(text(),'전체'))]"),
                # "다운로드 전체" = Download All (fallback nếu không có nút riêng lẻ)
                (By.XPATH, "//a[contains(text(),'다운로드 전체')]"),
                (By.XPATH, "//a[contains(text(),'다운로드')]"),
                # img-based arrow buttons
                (By.XPATH, "//a[img[contains(@src,'download') or contains(@src,'arrow') or contains(@src,'btn_down')]]"),
                (By.XPATH, "//input[@type='image'][contains(@src,'download') or contains(@src,'arrow')]"),
                # CSS class
                (By.CSS_SELECTOR, "a.btn_download, a.download_btn, a.btn_down"),
            ]

            dl_btn = _find_element(driver, download_btn_candidates)

            if not dl_btn:
                # Last resort: dump all links in popup for debug
                all_links = driver.execute_script("""
                    var links = document.querySelectorAll('a, button, input[type=image]');
                    var info = [];
                    for (var i = 0; i < links.length; i++) {
                        var el = links[i];
                        if (el.offsetParent !== null) {
                            info.push({
                                tag: el.tagName,
                                text: (el.innerText||'').trim().substring(0,50),
                                href: el.href || '',
                                cls: el.className,
                                onclick: (el.getAttribute('onclick')||'').substring(0,100)
                            });
                        }
                    }
                    return info;
                """)
                logger.warning(f"PO {po_no}: download button not found. Visible elements: {all_links}")
                driver.close()
                driver.switch_to.window(main_window)
                errors += 1
                continue

            logger.info(f"Clicking download button: tag={dl_btn.tag_name}, text='{dl_btn.text[:30]}'")
            _js_click(driver, dl_btn)

            # Chờ download bắt đầu (file .crdownload hoặc .pdf xuất hiện trong temp_dir)
            dl_started = False
            for _ in range(20):  # poll 0.5s x 20 = 10s max
                time.sleep(0.5)
                if os.path.isdir(temp_dir):
                    files_now = set(os.listdir(temp_dir))
                    new = files_now - files_before
                    if any(f.endswith(('.pdf', '.crdownload')) for f in new):
                        dl_started = True
                        logger.info("Download started")
                        break

            if not dl_started:
                logger.warning(f"PO {po_no}: download not started after 10s — closing popup anyway")

            # Đóng popup sau khi download đã bắt đầu
            try:
                driver.close()
            except WebDriverException:
                pass
            driver.switch_to.window(main_window)

            # Chờ file xuất hiện trong temp_dir
            logger.info(f"Waiting for PDF to download to {temp_dir}...")
            downloaded_file = _wait_for_download_complete(temp_dir, timeout=30)

            if not downloaded_file:
                logger.warning(f"PO {po_no}: download timeout (30s)")
                errors += 1
                continue

            # Move file về month_folder
            fname = os.path.basename(downloaded_file)
            dest = os.path.join(month_folder, fname)

            # Nếu file đã tồn tại ở dest thì skip
            if os.path.exists(dest):
                os.remove(downloaded_file)
                logger.info(f"File đã tồn tại ở dest: {dest}")
                skipped += 1
            else:
                import shutil
                shutil.move(downloaded_file, dest)
                logger.info(f"Saved: {dest}")
                files_saved.append(dest)
                downloaded += 1

            # Đảm bảo còn ở main window trước khi sang PO tiếp theo
            driver.switch_to.window(main_window)
            time.sleep(0.5)

        except Exception as e:
            logger.error(f"PO {po_no} download error: {e}")
            errors += 1
            # Đảm bảo switch về main window dù có lỗi
            try:
                for h in driver.window_handles:
                    if h != main_window:
                        driver.switch_to.window(h)
                        driver.close()
                driver.switch_to.window(main_window)
            except Exception:
                pass

    result = {
        "downloaded": downloaded,
        "skipped": skipped,
        "errors": errors,
        "files": files_saved,
    }
    logger.info(f"PDF download done: {result}")
    return result


def test_full_flow() -> dict:
    """
    Test toàn bộ flow với Chrome hiện ra để quan sát.
    Flow: login → popup → navigate → scrape all pages → download PDFs.
    """
    driver = create_driver(headless=False, download_dir=PO_TEMP_DOWNLOAD)
    try:
        print("[1] Logging in...")
        result = login_with_retry(driver)
        if not result["success"]:
            print(f"FAILED: {result['message']}")
            time.sleep(5)
            return result
        print(f"[1] OK - {driver.current_url}")

        print("[2] Handling post-login popups...")
        handle_post_login_popups(driver)
        print("[2] OK")

        print("[3] Navigating to Execution > MRO > P/O Receipt...")
        ok = navigate_to_po_receipt_mro(driver)
        if ok:
            print(f"[3] OK - {driver.title}")
            print(f"    URL: {driver.current_url}")
        else:
            print("[3] FAILED")
            time.sleep(5)
            result["nav_success"] = False
            return result

        print("[4] Scraping all pages...")
        rows = scrape_all_pages(driver)
        print(f"[4] Got {len(rows)} total PO rows")

        if rows:
            print("[5] Downloading PDFs (chỉ test 1 PO đầu tiên)...")
            pdf_result = download_po_pdfs(driver, rows[:1])
            print(
                f"[5] PDFs: downloaded={pdf_result['downloaded']}, "
                f"skipped={pdf_result['skipped']}, errors={pdf_result['errors']}"
            )
            for f in pdf_result["files"]:
                print(f"     → {f}")
        else:
            print("[5] No rows to process")

        print("[6] Done. Keeping browser open 5s...")
        time.sleep(5)

        result["nav_success"] = ok
        return result
    finally:
        driver.quit()


def test_login_visible() -> dict:
    return test_full_flow()


if __name__ == "__main__":
    import sys
    if sys.stdout.encoding != "utf-8":
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    test_login_visible()
