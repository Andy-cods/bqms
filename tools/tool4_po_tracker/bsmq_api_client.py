# -*- coding: utf-8 -*-
"""
BSMQ API Client — Samsung BQMS Vendor Portal (requests-based, no Selenium)

Login flow (4 steps via HTTP):
  1. GET vendorLogin.do  → JSESSIONID cookie
  2. POST loginCertificatePage.do  { loginPasswordHash: SHA256(pw) }
     → { result:"SUCCESS", userId, certCode, certName }
  3. POST mfaCreateToken.do  { userId, passwordHash }
     → { result:"DIRECT" }
  4. POST authLoginPage.do  (certForm fields)
     → redirect → vendor portal (Set-Cookie: session confirmed)

PO list: POST selectPOAcceptList.do  Content-Type: application/json;charset=UTF-8
  body: { srchStDate, srchEdDate, srchStatusCode:["N"],
          mroPageVO:{ pageInfos:[{id,pageIndex,location,pageScript,pageSize,originalPageSize}] } }

API response fields used:
  PO_NO, REQ_NO, ITEM_CODE, SPECIFICATION, PO_QTY, UNIT_CODE,
  RECEIVER_NAME  ("Full Name: mail_prefix"),
  PO_CONFIRM_DT (Unix ms), REQ_DELIVERY_DATE ("YYYYMMDD")
"""

import os
import hashlib
import logging
import unicodedata
from datetime import datetime, date, timedelta
from typing import Optional

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

logger = logging.getLogger(__name__)

# ── Constants ──────────────────────────────────────────────────────────────────
BASE_URL = "https://www.sec-bqms.com"
LOGIN_PAGE_URL     = f"{BASE_URL}/bqms/vendorPortal/anonymous/vendorLogin.do?_frameF=true"
CERT_PAGE_URL      = f"{BASE_URL}/bqms/vendorPortal/anonymous/loginCertificatePage.do"
MFA_TOKEN_URL      = f"{BASE_URL}/bqms/partnerLogin/anonymous/mfaCreateToken.do"
AUTH_LOGIN_URL     = f"{BASE_URL}/bqms/partnerLogin/anonymous/authLoginPage.do"
PO_LIST_URL        = f"{BASE_URL}/bqms/mro/vendor/selectPOAcceptList.do"
PO_NAV_URL         = f"{BASE_URL}/bqms/mro/forward/vendor/vendorPoConfirm.do?target=vendor&_menuId=AZknkggsAB8V-Qhq&_menuF=true"

DEFAULT_USERNAME = os.environ.get("BSMQ_USERNAME", "amabnjsc")
DEFAULT_PASSWORD = os.environ.get("BSMQ_PASSWORD", "AMA202601@1")

def _get_excel_path():
    import os
    home = os.path.expanduser("~")
    for name in os.listdir(home):
        if "SONG CHAU" in name.upper() and os.path.isdir(os.path.join(home, name)):
            p = os.path.join(home, name, "Puplic", "BQMS", "Thong ke giao hang", "Thong ke giao hang 2026.xlsx")
            if os.path.exists(p):
                return p
    return os.path.join(home, "OneDrive - SONG CHAU CO., LTD", "Puplic", "BQMS", "Thong ke giao hang", "Thong ke giao hang 2026.xlsx")

EXCEL_PATH = _get_excel_path()

_HEADERS_BASE = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": BASE_URL,
}


# ── Session builder ────────────────────────────────────────────────────────────

def _build_session() -> requests.Session:
    s = requests.Session()
    retry = Retry(total=3, backoff_factor=0.5,
                  status_forcelist=[500, 502, 503, 504])
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.headers.update(_HEADERS_BASE)
    return s


# ── SHA-256 helper ─────────────────────────────────────────────────────────────

def _sha256(text: str) -> str:
    return hashlib.sha256(text.encode("utf-8")).hexdigest()


# ── Login ──────────────────────────────────────────────────────────────────────

def login_api(
    username: str = DEFAULT_USERNAME,
    password: str = DEFAULT_PASSWORD,
) -> Optional[requests.Session]:
    """
    4-step login to Samsung BQMS.
    Returns an authenticated requests.Session or None on failure.
    """
    session = _build_session()
    pw_hash = _sha256(password)

    # ── Step 1: GET login page → JSESSIONID ────────────────────────────────
    try:
        r = session.get(LOGIN_PAGE_URL, timeout=20)
        r.raise_for_status()
        logger.info("Step 1 OK — JSESSIONID: %s", session.cookies.get("JSESSIONID", "?"))
    except Exception as e:
        logger.error("Step 1 GET login page failed: %s", e)
        return None

    # ── Step 2: POST loginCertificatePage.do ──────────────────────────────
    try:
        payload = {
            "loginId": username,
            "loginPasswordHash": pw_hash,
        }
        r = session.post(CERT_PAGE_URL, data=payload, timeout=20)
        r.raise_for_status()
        data = r.json()
        logger.info("Step 2 response: %s", data)

        if data.get("result") != "SUCCESS":
            logger.error("Step 2 failed: result=%s", data.get("result"))
            return None

        user_id   = data["userId"]
        cert_code = data.get("certCode", "")
        cert_name = data.get("certName", "")
        logger.info("Step 2 OK — userId=%s, certCode=%s", user_id, cert_code)
    except Exception as e:
        logger.error("Step 2 POST cert page failed: %s", e)
        return None

    # ── Step 3: POST mfaCreateToken.do (optional — 404 = skip) ───────────
    try:
        payload = {
            "userId": user_id,
            "passwordHash": pw_hash,
        }
        r = session.post(MFA_TOKEN_URL, data=payload, timeout=20)
        if r.status_code == 404:
            logger.info("Step 3 skipped (endpoint not found — OK, proceeding)")
        else:
            r.raise_for_status()
            data = r.json()
            logger.info("Step 3 response: %s", data)
    except requests.exceptions.HTTPError as e:
        logger.info("Step 3 HTTP error (non-fatal, skipping): %s", e)
    except Exception as e:
        logger.info("Step 3 error (non-fatal, skipping): %s", e)

    # ── Step 4: POST authLoginPage.do (certForm hidden fields) ────────────
    try:
        payload = {
            "userId":       user_id,
            "certId":       cert_code,    # form field name is certId
            "certName":     cert_name,
            "passwordHash": pw_hash,
            "certPassword": "",           # empty for DIRECT MFA flow
            "signsrc":      "",
            "signdata":     "",
            "signcert":     "",
        }
        r = session.post(AUTH_LOGIN_URL, data=payload,
                         timeout=20, allow_redirects=True)
        r.raise_for_status()
        final_url = r.url
        logger.info("Step 4 OK — final URL: %s", final_url)

        # Confirm we're no longer on an anonymous/login page
        if "anonymous" in final_url or "login" in final_url.lower():
            logger.warning("Step 4 may have failed — still on login URL: %s", final_url)
        else:
            logger.info("Login confirmed — session authenticated")

    except Exception as e:
        logger.error("Step 4 POST auth login failed: %s", e)
        return None

    return session


# ── Fetch PO list ──────────────────────────────────────────────────────────────

def fetch_po_list(
    session: requests.Session,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
) -> list:
    """
    Fetch all PO rows from the BQMS JSON API.
    Dates: "YYYYMMDD" strings.  Defaults to current month (1st → today).

    Returns a list of dicts with raw API field names.
    Key fields: PO_NO, REQ_NO, CIS_CODE, ITEM_CODE, SPECIFICATION,
                PO_QTY, UNIT_CODE, RECEIVER_NAME, PO_CONFIRM_DT,
                REQ_DELIVERY_DATE, COMPANY_NAME, secureKey, PO_SEQ, SP_CODE
    """
    today = date.today()
    if not start_date:
        start_date = today.replace(day=1).strftime("%Y%m%d")
    if not end_date:
        end_date = today.strftime("%Y%m%d")

    # Navigate to PO receipt page first to establish page context
    try:
        session.get(PO_NAV_URL, timeout=20)
    except Exception as e:
        logger.warning("Could not navigate to PO page (non-fatal): %s", e)

    body = {
        "srchStDate":      start_date,
        "srchEdDate":      end_date,
        "srchPoNo":        None,
        "srchCompanyCode": None,
        "srchStatusCode":  ["N"],
        "mroPageVO": {
            "pageInfos": [
                {
                    "id":               "page1",
                    "pageIndex":        1,
                    "location":         ".paginate",
                    "pageScript":       "search()",
                    "pageSize":         "9999",
                    "originalPageSize": None,
                }
            ]
        },
    }

    headers = {
        "Content-Type":    "application/json; charset=UTF-8",
        "Accept":          "application/json, text/plain, */*",
        "X-Requested-With": "XMLHttpRequest",
        "Referer":         PO_NAV_URL,
        "Origin":          BASE_URL,
    }

    try:
        r = session.post(PO_LIST_URL, json=body, headers=headers, timeout=30)
        r.raise_for_status()
        data = r.json()
    except Exception as e:
        logger.error("fetch_po_list request failed: %s", e)
        return []

    # Response: {"page1_result": {..., "totalCnt": "34"}, "poList": [...]}
    rows = data.get("poList") or data.get("list") or data.get("rows") or (
        data if isinstance(data, list) else []
    )
    total = data.get("page1_result", {}).get("totalCnt", "?")
    logger.info("Fetched %d PO rows (total=%s, %s→%s)", len(rows), total, start_date, end_date)
    return rows


# ── Date helpers ───────────────────────────────────────────────────────────────

def _ms_to_date(ms_val) -> Optional[date]:
    """Unix milliseconds → Python date"""
    try:
        return date.fromtimestamp(int(ms_val) / 1000)
    except Exception:
        return None


def _yyyymmdd_to_date(val: str) -> Optional[date]:
    """'20260316' → date(2026,3,16)"""
    try:
        return datetime.strptime(str(val).strip(), "%Y%m%d").date()
    except Exception:
        return None


def _parse_api_date(val) -> Optional[date]:
    """Try Unix ms first, then YYYYMMDD string."""
    if val is None:
        return None
    s = str(val).strip()
    if s.isdigit() and len(s) >= 10:
        return _ms_to_date(s) if len(s) == 13 else _yyyymmdd_to_date(s)
    return _yyyymmdd_to_date(s)


# ── Mail prefix extractor ──────────────────────────────────────────────────────

def _extract_mail_prefix(receiver_name: str) -> str:
    """
    'Duong Minh Tuyen: minh.tuyen1' → 'minh.tuyen1'
    'Full Name'                     → ''
    """
    if ":" in receiver_name:
        return receiver_name.split(":")[-1].strip()
    return ""


def _extract_receiver_display(receiver_name: str) -> str:
    """'Duong Minh Tuyen: minh.tuyen1' → 'Duong Minh Tuyen'"""
    if ":" in receiver_name:
        return receiver_name.split(":")[0].strip()
    return receiver_name.strip()


# ── Excel writer ───────────────────────────────────────────────────────────────

def _get_target_sheet(wb):
    """
    Find the 'THỐNG KÊ PO' sheet (Vietnamese name, tolerant of diacritics).
    Falls back to wb.active if not found.
    """
    def _normalize(s: str) -> str:
        return "".join(
            c for c in unicodedata.normalize("NFD", s.upper())
            if unicodedata.category(c) != "Mn"
        )

    target_kw = ["THONG KE PO", "KE PO", "THONG KE"]
    for name in wb.sheetnames:
        norm = _normalize(name)
        if any(kw in norm for kw in target_kw):
            logger.info("Target sheet: '%s'", name)
            return wb[name]

    logger.warning("Sheet 'THỐNG KÊ PO' not found by name — using wb.active. Sheets: %s", wb.sheetnames)
    return wb.active


def write_api_pos_to_excel(
    po_rows: list,
    excel_path: str = EXCEL_PATH,
) -> dict:
    """
    Write new PO rows (from API) to the Excel tracking file.
    Deduplicates by (PO_NO, REQ_NO) — skips rows already present.

    Column mapping (1-indexed):
      A=1  Ngày PO         ← PO_CONFIRM_DT (Unix ms → date)
      B=2  Số PO           ← PO_NO
      D=4  Số QT           ← REQ_NO
      E=5  BQMS code       ← ITEM_CODE
      F=6  Spec            ← SPECIFICATION
      G=7  SL              ← PO_QTY
      H=8  Đơn vị          ← UNIT_CODE
      K=11 SEV/T           ← "SEVT" (constant)
      L=12 MAIL PUR        ← mail prefix from RECEIVER_NAME
      M=13 TÊN NGƯỜI NHẬN  ← display name from RECEIVER_NAME
      Q=17 NGÀY GIAO HÀNG  ← REQ_DELIVERY_DATE (YYYYMMDD → date)

    Returns: {added, skipped, total, errors}
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return {"added": 0, "skipped": 0, "total": 0, "errors": 1}

    if not po_rows:
        return {"added": 0, "skipped": 0, "total": 0, "errors": 0}

    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        logger.error("Cannot open Excel file: %s", e)
        return {"added": 0, "skipped": 0, "total": len(po_rows), "errors": 1}

    ws = _get_target_sheet(wb)

    # ── Read existing (PO_NO, REQ_NO) pairs to dedup ──────────────────────
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        so_po = str(row[1]).strip() if row[1] else ""   # col B = index 1
        so_qt = str(row[3]).strip() if row[3] else ""   # col D = index 3
        if so_po and so_po != "None" and so_qt and so_qt != "None":
            existing.add((so_po, so_qt))

    logger.info("Existing rows in Excel: %d unique (PO_NO, REQ_NO) pairs", len(existing))

    # Find insertion point: last data row + 1 (before any "Total" row)
    insert_row = 2  # fallback
    for r_idx in range(ws.max_row, 1, -1):
        cell_a = ws.cell(row=r_idx, column=1).value
        cell_b = ws.cell(row=r_idx, column=2).value
        # Skip empty rows and "Total" / formula rows
        if cell_b and str(cell_b).strip() and str(cell_b).strip().lower() not in ("none", ""):
            insert_row = r_idx + 1
            break

    logger.info("Inserting new rows starting at row %d", insert_row)

    added = 0
    skipped = 0
    errors = 0

    for row in po_rows:
        po_no  = str(row.get("PO_NO",  "") or "").strip()
        req_no = str(row.get("REQ_NO", "") or "").strip()

        if not po_no:
            errors += 1
            continue

        if (po_no, req_no) in existing:
            skipped += 1
            continue

        try:
            # A=1: Ngày PO (date)
            po_date = _parse_api_date(row.get("PO_CONFIRM_DT"))
            ws.cell(row=insert_row, column=1, value=po_date)

            # B=2: Số PO
            ws.cell(row=insert_row, column=2, value=po_no)

            # C=3: Shipping No — blank

            # D=4: Số QT (Request No)
            ws.cell(row=insert_row, column=4, value=req_no)

            # E=5: BQMS code (CIS_CODE = Q-prefix catalog code, e.g. "Q311-205111")
            item_code = str(row.get("CIS_CODE", "") or row.get("ITEM_CODE", "") or "").strip()
            ws.cell(row=insert_row, column=5, value=item_code)

            # F=6: Specification
            spec = str(row.get("SPECIFICATION", "") or "").strip()
            ws.cell(row=insert_row, column=6, value=spec)

            # G=7: Qty
            qty = row.get("PO_QTY")
            try:
                qty = int(float(str(qty).replace(",", ""))) if qty else None
            except Exception:
                qty = None
            ws.cell(row=insert_row, column=7, value=qty)

            # H=8: Unit
            unit = str(row.get("UNIT_CODE", "") or "").strip()
            ws.cell(row=insert_row, column=8, value=unit)

            # I=9, J=10: Unit price, Amount — blank (filled manually)

            # K=11: SEV/T constant
            ws.cell(row=insert_row, column=11, value="SEVT")

            # L=12: MAIL PUR (mail prefix from RECEIVER_NAME)
            receiver_raw = str(row.get("RECEIVER_NAME", "") or "").strip()
            mail_prefix = _extract_mail_prefix(receiver_raw)
            if mail_prefix:
                ws.cell(row=insert_row, column=12, value=mail_prefix)

            # M=13: TÊN NGƯỜI NHẬN (display name)
            receiver_display = _extract_receiver_display(receiver_raw)
            if receiver_display:
                ws.cell(row=insert_row, column=13, value=receiver_display)

            # Q=17: NGÀY GIAO HÀNG (delivery date)
            rd_raw = row.get("REQ_DELIVERY_DATE")
            rd_date = _parse_api_date(rd_raw)
            if rd_date:
                ws.cell(row=insert_row, column=17, value=rd_date)

            existing.add((po_no, req_no))
            insert_row += 1
            added += 1

        except Exception as e:
            logger.error("Error writing row PO=%s REQ=%s: %s", po_no, req_no, e)
            errors += 1

    try:
        wb.save(excel_path)
        logger.info("Excel saved: +%d added, %d skipped, %d errors", added, skipped, errors)
    except Exception as e:
        logger.error("Failed to save Excel: %s", e)
        errors += 1

    return {
        "added":   added,
        "skipped": skipped,
        "total":   len(po_rows),
        "errors":  errors,
    }


# ── Full API scan flow ─────────────────────────────────────────────────────────

def run_po_scan(
    username: str = DEFAULT_USERNAME,
    password: str = DEFAULT_PASSWORD,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    excel_path: str = EXCEL_PATH,
    log_fn=None,
) -> dict:
    """
    Full flow: login → fetch PO list → write new rows to Excel.

    log_fn: optional callable(msg: str) for live progress updates (used by app.py)
    Returns: {success, po_count, added, skipped, errors, message}
    """

    def _log(msg: str):
        logger.info(msg)
        if log_fn:
            log_fn(msg)

    result = {
        "success": False,
        "po_count": 0,
        "added": 0,
        "skipped": 0,
        "errors": 0,
        "message": "",
    }

    _log("[1/3] Dang nhap BQMS API...")
    session = login_api(username, password)
    if not session:
        result["message"] = "Login that bai. Kiem tra tai khoan/mat khau."
        return result
    _log("[1/3] Login thanh cong")

    _log("[2/3] Dang tai danh sach PO...")
    po_rows = fetch_po_list(session, start_date, end_date)
    result["po_count"] = len(po_rows)
    if not po_rows:
        result["message"] = "Khong lay duoc PO nao tu API."
        return result
    _log("[2/3] Da tai %d dong PO" % len(po_rows))

    _log("[3/3] Dang ghi vao Excel...")
    write_result = write_api_pos_to_excel(po_rows, excel_path)
    result.update(write_result)
    result["success"] = write_result.get("errors", 0) == 0

    _log(
        "[3/3] Hoan tat: +%d moi, %d da co, %d loi"
        % (write_result["added"], write_result["skipped"], write_result["errors"])
    )
    result["message"] = (
        "Hoan tat: %d PO moi ghi vao Excel, %d da co san, %d loi"
        % (write_result["added"], write_result["skipped"], write_result["errors"])
    )
    return result


# ── CLI test ───────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    import sys
    import io

    # Fix console encoding on Windows
    if hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=[logging.StreamHandler(sys.stdout)],
    )

    print("=== BSMQ API Client Test ===")
    print("Connecting to Samsung BQMS...")

    res = run_po_scan(
        log_fn=print,
        # Scan current month by default
    )

    print("\n=== RESULT ===")
    for k, v in res.items():
        print("  %s: %s" % (k, v))
