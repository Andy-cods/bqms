"""Extract transactions from TT XNK BQMS Excel files.

Full column mapping (28 columns A-AB):
  A: TT (row number)
  B: Ngay Thang (transaction date)
  C: Don hang (RFQ no, e.g. QT26000091)
  D: BMSQ (BQMS code)
  E: Ten hang hoa (product name/spec)
  F: Explain/Description
  G: Loai hang (Thuong mai / Gia cong)
  H: Maker
  I: Ghi chu (note 1)
  J: Ghi chu 2 (note 2)
  K: Don vi tinh (unit)
  L: So luong (quantity)
  M: Quote Deadline
  N: Ngay (import date)
  O: BMSQ3 (import BQMS code)
  P: Mieu ta hang hoa (import description)
  Q: Ma HS (HS code)
  R: DVT (import unit)
  S: SL (import quantity)
  T: Tong cong USD (total USD)
  U: Don gia USD (unit price USD)
  V: Don gia VND (unit price VND)
  W: Ben mua (buyer: SEVT, SEV, etc.)
  X: Ben ban (seller name)
  Y: NCC khac / Ghi chu (DATA) or older year column
  Z-AB: Year-specific columns (DATA sheet only)
"""

import hashlib
import os
from datetime import datetime
from typing import Optional

import openpyxl


def _str(v) -> str:
    if v is None:
        return ""
    s = str(v).replace("\xa0", " ").strip()
    if s in ("#VALUE!", "#REF!", "#N/A", "#DIV/0!"):
        return ""
    return s


def _num(v) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace(",", "").replace("\xa0", "").strip()
    if not s or s in ("#VALUE!", "#REF!"):
        return None
    try:
        return float(s)
    except (ValueError, TypeError):
        return None


def _date(v) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    if not s or len(s) < 6:
        return None
    return s[:10]


def _cell(row: tuple, idx: int) -> any:
    """Safe cell access."""
    if idx < len(row):
        return row[idx]
    return None


def _hash_row(rfq: str, bqms: str, date_str: str, spec: str) -> str:
    key = f"{rfq}|{bqms}|{date_str}|{spec}"
    return hashlib.sha256(key.encode("utf-8")).hexdigest()[:32]


def _classify_type(v: str) -> Optional[str]:
    lh = v.lower().replace("\xa0", " ").strip()
    if any(x in lh for x in ("gia cong", "gia công", "gc")):
        return "gc"
    if any(x in lh for x in ("thuong mai", "thương mại", "tm")):
        return "tm"
    return None


def extract_file(filepath: str, sheets: list[str] | None = None) -> list[dict]:
    """Extract all 28 columns from TT XNK BQMS file."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)

    if sheets is None:
        target_sheets = []
        for s in wb.sheetnames:
            sl = s.upper()
            if sl in ("DATA", "JAN", "FEB", "MAR", "APR", "MAY", "JUN",
                       "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"):
                target_sheets.append(s)
        if not target_sheets:
            target_sheets = wb.sheetnames[:1]
    else:
        target_sheets = sheets

    source_file = os.path.basename(filepath)
    results = []
    seen_hashes = set()

    for sheet_name in target_sheets:
        ws = wb[sheet_name]
        row_idx = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            row_idx += 1
            if not row or len(row) < 5:
                continue

            rfq = _str(_cell(row, 2))       # C: Don hang
            bqms = _str(_cell(row, 3))      # D: BMSQ
            spec = _str(_cell(row, 4))      # E: Ten hang hoa

            if not rfq and not bqms and not spec:
                continue

            date_str = _date(_cell(row, 1))  # B: Ngay Thang

            h = _hash_row(rfq, bqms, date_str or "", spec)
            if h in seen_hashes:
                continue
            seen_hashes.add(h)

            # Extract ALL columns
            description = _str(_cell(row, 5))   # F
            type_raw = _str(_cell(row, 6))      # G: Loai hang
            maker = _str(_cell(row, 7))         # H
            note1 = _str(_cell(row, 8))         # I: Ghi chu
            note2 = _str(_cell(row, 9))         # J: Ghi chu 2
            unit = _str(_cell(row, 10))         # K: DVT
            qty = _num(_cell(row, 11))          # L: So luong
            deadline = _str(_cell(row, 12))     # M: Quote Deadline
            import_date = _date(_cell(row, 13)) # N: Ngay
            bqms_import = _str(_cell(row, 14))  # O: BMSQ3
            desc_import = _str(_cell(row, 15))  # P: Mieu ta hang hoa
            hs_code = _str(_cell(row, 16))      # Q: Ma HS
            unit2 = _str(_cell(row, 17))        # R: DVT import
            qty2 = _num(_cell(row, 18))         # S: SL import
            total_usd = _num(_cell(row, 19))    # T: Tong cong USD
            unit_price_usd = _num(_cell(row, 20))  # U: Don gia USD
            unit_price_vnd = _num(_cell(row, 21))   # V: Don gia VND
            buyer = _str(_cell(row, 22))        # W: Ben mua
            seller = _str(_cell(row, 23))       # X: Ben ban

            # Y or AB: supplier alt / notes (varies by sheet)
            supplier_alt = _str(_cell(row, 24))  # Y
            if not supplier_alt and len(row) > 27:
                supplier_alt = _str(_cell(row, 27))  # AB in DATA sheet

            # Combine notes
            notes_parts = [p for p in [note1, note2] if p]
            notes = " | ".join(notes_parts) if notes_parts else None

            results.append({
                "transaction_date": date_str,
                "rfq_no": rfq[:100] if rfq else None,
                "bqms_code": bqms[:100] if bqms else None,
                "spec": spec[:500] if spec else None,
                "description": description[:300] if description else None,
                "type": _classify_type(type_raw) if type_raw else None,
                "maker": maker[:200] if maker else None,
                "quantity": qty,
                "unit": unit[:20] if unit else "EA",
                "deadline_text": deadline[:100] if deadline else None,
                "import_date": import_date,
                "bqms_import": bqms_import[:100] if bqms_import else None,
                "desc_import": desc_import[:500] if desc_import else None,
                "hs_code": hs_code[:20] if hs_code else None,
                "unit2": unit2[:20] if unit2 else None,
                "qty2": qty2,
                "price_quoted": total_usd,
                "unit_price_usd": unit_price_usd,
                "unit_price_vnd": unit_price_vnd,
                "buyer": buyer[:100] if buyer else None,
                "seller": seller[:200] if seller else None,
                "supplier_alt": supplier_alt[:300] if supplier_alt else None,
                "notes": notes[:500] if notes else None,
                "price_rmb": None,
                "price_vnd": unit_price_vnd,
                "version": None,
                "result": None,
                "person_in_charge": None,
                "source_file": source_file,
                "source_sheet": sheet_name,
                "source_row": row_idx + 1,
                "row_hash": h,
            })

    wb.close()
    return results
