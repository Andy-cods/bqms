"""
modules/data_validation.py
Input file validation (F06–F12). Pure Python — no Streamlit dependency.
"""

import os
import unicodedata
from dataclasses import dataclass, field
from typing import Optional

import pandas as pd
import openpyxl


# ---------------------------------------------------------------------------
# Result type
# ---------------------------------------------------------------------------

@dataclass
class ValidationResult:
    ok: bool
    errors: list = field(default_factory=list)
    warnings: list = field(default_factory=list)

    def add_error(self, msg: str) -> None:
        self.errors.append(msg)
        self.ok = False

    def add_warning(self, msg: str) -> None:
        self.warnings.append(msg)

    def merge(self, other: 'ValidationResult') -> None:
        self.errors.extend(other.errors)
        self.warnings.extend(other.warnings)
        if not other.ok:
            self.ok = False


# ---------------------------------------------------------------------------
# F06 — Validate file can be opened
# ---------------------------------------------------------------------------

def validate_input_file(path: str) -> ValidationResult:
    """Try opening file with openpyxl. Catches corrupt/locked/missing files."""
    result = ValidationResult(ok=True)
    if not path:
        result.add_error("Không có đường dẫn file.")
        return result
    if not os.path.exists(path):
        result.add_error(f"File không tồn tại: {path}")
        return result
    if os.path.getsize(path) == 0:
        result.add_error("File rỗng (0 bytes).")
        return result
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        if not wb.sheetnames:
            result.add_error("File không có sheet nào.")
        wb.close()
    except PermissionError:
        result.add_error("File đang bị mở hoặc bị khóa — đóng Excel rồi thử lại.")
    except Exception as e:
        result.add_error(f"Không đọc được file: {e}")
    return result


# ---------------------------------------------------------------------------
# F07 — Column schema fuzzy check
# ---------------------------------------------------------------------------

def _normalize(s: str) -> str:
    """Lowercase + strip accents for fuzzy matching."""
    nfkd = unicodedata.normalize('NFKD', str(s).lower())
    return ''.join(c for c in nfkd if not unicodedata.combining(c)).strip()


def check_column_schema(df: pd.DataFrame, required_keywords: dict) -> ValidationResult:
    """
    required_keywords: {field_name: [keyword1, keyword2, ...]}
    Fuzzy-matches column headers. Errors on required fields not found.
    """
    result = ValidationResult(ok=True)
    normalized_cols = [_normalize(str(c)) for c in df.columns]
    REQUIRED_FIELDS = {'don_hang', 'bqms', 'loai_hang'}

    for field_name, keywords in required_keywords.items():
        found = any(
            any(_normalize(kw) in col for kw in keywords)
            for col in normalized_cols
        )
        if not found:
            if field_name in REQUIRED_FIELDS:
                result.add_error(f"Không tìm thấy cột bắt buộc: '{field_name}' (keywords: {keywords})")
            else:
                result.add_warning(f"Không tìm thấy cột: '{field_name}' (keywords: {keywords})")
    return result


# ---------------------------------------------------------------------------
# F08 — File lock check
# ---------------------------------------------------------------------------

def check_file_lock(path: str) -> bool:
    """Returns True if file is locked (cannot be opened exclusively)."""
    if not os.path.exists(path):
        return False
    try:
        with open(path, 'r+b'):
            return False
    except (IOError, PermissionError, OSError):
        return True


# ---------------------------------------------------------------------------
# F09 — Duplicate codes
# ---------------------------------------------------------------------------

def check_duplicate_codes(df: pd.DataFrame, col: str) -> list:
    """Returns list of codes that appear more than once."""
    if col not in df.columns:
        return []
    dupes = df[df[col].notna() & df[col].duplicated(keep=False)][col].unique().tolist()
    return [str(d) for d in dupes]


# ---------------------------------------------------------------------------
# F10 — Missing fields
# ---------------------------------------------------------------------------

def check_missing_fields(df: pd.DataFrame, cols: list) -> dict:
    """Returns {col_name: count_of_null_rows} for each requested column."""
    result = {}
    for col in cols:
        if col in df.columns:
            count = int(df[col].isna().sum())
            if count > 0:
                result[col] = count
    return result


# ---------------------------------------------------------------------------
# F11 — Merge cells warning (for template files)
# ---------------------------------------------------------------------------

def check_merge_cells(path: str, sheet_index: int = 0) -> ValidationResult:
    """Check if template has merged cells (informational only)."""
    result = ValidationResult(ok=True)
    try:
        wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
        ws = wb.worksheets[sheet_index] if wb.worksheets else None
        if ws and ws.merged_cells:
            count = len(ws.merged_cells.ranges)
            result.add_warning(f"File có {count} vùng ô được gộp (merged cells).")
        wb.close()
    except Exception as e:
        result.add_warning(f"Không thể kiểm tra merged cells: {e}")
    return result


# ---------------------------------------------------------------------------
# F12 — Full validation pipeline
# ---------------------------------------------------------------------------

def run_full_validation(
    path: str,
    required_cols: Optional[dict] = None,
    df: Optional[pd.DataFrame] = None
) -> ValidationResult:
    """
    Orchestrate F06–F11. Returns aggregated ValidationResult.
    df is optional — if provided, enables column/duplicate checks.
    """
    final = ValidationResult(ok=True)

    # F06: File open check
    file_result = validate_input_file(path)
    final.merge(file_result)
    if not file_result.ok:
        return final  # Can't proceed if file unreadable

    # F08: File lock check
    if check_file_lock(path):
        final.add_error("File đang bị mở — đóng Excel rồi thử lại.")

    if df is not None:
        # F07: Column schema
        if required_cols:
            col_result = check_column_schema(df, required_cols)
            final.merge(col_result)

        # F10: Check for missing fields in first 5 cols
        missing = check_missing_fields(df, list(df.columns[:5]))
        for col, count in missing.items():
            final.add_warning(f"Cột '{col}': {count} hàng trống.")

    return final
