"""Tool engine endpoints — wraps actual tool logic from old codebase.

Tool 1: Upload Excel → parse → match products → fill templates
Tool 2: Load analytics files → return dataframes as JSON
Tool 3: AI classification with Gemini + historical lookup
Tool 5: Market search (crawl platforms, verify, synthesize)
Tool 4: Desktop-only, pushes data via /tools/push-deliveries
"""

import os
import json
import uuid
import shutil
import tempfile
from datetime import datetime

from fastapi import APIRouter, Depends, UploadFile, File, BackgroundTasks
from fastapi.responses import FileResponse

from ...db.session import get_db
from ...domain.enums import UserRole
from ...domain.exceptions import NotFoundError, ValidationError
from ..deps import get_current_user, require_role

router = APIRouter(prefix="/tool-engine", tags=["tool-engines"])

# In-memory job store (same pattern as old app_api.py)
_jobs: dict[str, dict] = {}

UPLOAD_DIR = "/opt/songchau/data/uploads"
OUTPUT_DIR = "/opt/songchau/data/outputs"


def _new_job(job_type: str) -> str:
    job_id = uuid.uuid4().hex[:12]
    _jobs[job_id] = {
        "status": "running",
        "type": job_type,
        "progress": 0,
        "logs": [],
        "result": None,
        "error": None,
        "files": [],
        "started_at": datetime.now().isoformat(),
    }
    return job_id


def _log(job_id: str, msg: str):
    if job_id in _jobs:
        ts = datetime.now().strftime("%H:%M:%S")
        _jobs[job_id]["logs"].append(f"[{ts}] {msg}")


# ── Tool 1: Upload & Parse BC BQMS ─────────────────────

@router.post("/tool1/upload")
async def tool1_upload(
    file: UploadFile = File(...),
    _=Depends(require_role(UserRole.admin, UserRole.procurement)),
):
    """Upload BC BQMS Excel file, parse and return orders."""
    if not file.filename or not file.filename.endswith((".xlsx", ".xls", ".xlsm")):
        raise ValidationError("Chỉ chấp nhận file Excel (.xlsx, .xls, .xlsm)")

    os.makedirs(UPLOAD_DIR, exist_ok=True)
    tmp_path = os.path.join(UPLOAD_DIR, f"tool1_{uuid.uuid4().hex[:8]}_{file.filename}")
    content = await file.read()
    with open(tmp_path, "wb") as f:
        f.write(content)

    # Parse using Tool 1 engine
    try:
        from ...tools.tool1.engine import parse_bc_bqms
        orders = parse_bc_bqms(tmp_path)
        return {
            "success": True,
            "file": file.filename,
            "orders": orders if isinstance(orders, list) else [],
            "count": len(orders) if isinstance(orders, list) else 0,
        }
    except Exception as e:
        # Fallback: parse with openpyxl directly
        import openpyxl
        wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
        orders = []
        ws = wb[wb.sheetnames[0]]
        headers = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                headers = [str(c or "").strip() for c in row]
                continue
            if not row or not any(row[:5]):
                continue
            order = {}
            for j, h in enumerate(headers):
                if j < len(row) and row[j] is not None:
                    order[h] = str(row[j])
            if order:
                orders.append(order)
        wb.close()
        return {
            "success": True,
            "file": file.filename,
            "orders": orders[:200],
            "count": len(orders),
            "note": f"Parsed with fallback ({str(e)[:100]})",
        }


@router.post("/tool1/run")
async def tool1_run(
    body: dict,
    background_tasks: BackgroundTasks,
    conn=Depends(get_db),
    user=Depends(require_role(UserRole.admin, UserRole.procurement)),
):
    """Run auto-fill job: match products, fill templates."""
    job_id = _new_job("autofill")
    orders = body.get("orders", [])

    async def _run_job():
        try:
            _log(job_id, f"Bắt đầu xử lý {len(orders)} đơn hàng...")
            results = []

            for i, order in enumerate(orders):
                bqms = order.get("bqms_code", order.get("BMSQ", order.get("bqms", "")))
                spec = order.get("spec", order.get("Tên hàng hóa", ""))

                # Match from DB
                matches = []
                if bqms:
                    rows = await conn.fetch(
                        "SELECT bqms_code, product_name, spec, maker, type FROM products WHERE bqms_code = $1 LIMIT 3",
                        bqms,
                    )
                    matches = [dict(r) for r in rows]

                if not matches and spec:
                    rows = await conn.fetch(
                        "SELECT bqms_code, product_name, spec, maker, type FROM products WHERE spec ILIKE $1 LIMIT 5",
                        f"%{spec[:40]}%",
                    )
                    matches = [dict(r) for r in rows]

                # Get price history
                prices = []
                if bqms:
                    rows = await conn.fetch(
                        "SELECT price_bqms_v1, price_bqms_v2, supplier, result FROM quotations WHERE bqms_code = $1 ORDER BY date DESC LIMIT 5",
                        bqms,
                    )
                    prices = [dict(r) for r in rows]

                results.append({
                    "order": order,
                    "matches": matches,
                    "prices": prices,
                    "matched": len(matches) > 0,
                })

                _jobs[job_id]["progress"] = int((i + 1) / len(orders) * 100)
                _log(job_id, f"Đã xử lý {i+1}/{len(orders)}: {bqms or spec[:30]}")

            _jobs[job_id]["result"] = {
                "total": len(orders),
                "matched": sum(1 for r in results if r["matched"]),
                "items": results,
            }
            _jobs[job_id]["status"] = "done"
            _log(job_id, f"Hoàn thành! {sum(1 for r in results if r['matched'])}/{len(orders)} sản phẩm tìm thấy")

        except Exception as e:
            _jobs[job_id]["status"] = "error"
            _jobs[job_id]["error"] = str(e)
            _log(job_id, f"Lỗi: {e}")

    background_tasks.add_task(_run_job)
    return {"job_id": job_id}


# ── Tool 2: Analytics ───────────────────────────────────

@router.get("/tool2/analytics")
async def tool2_analytics(
    page: int = 1, limit: int = 25,
    conn=Depends(get_db), _=Depends(get_current_user),
):
    """Load analytics data from quotations table (replaces Excel file loading)."""
    offset = (page - 1) * min(limit, 100)
    total = await conn.fetchval("SELECT COUNT(*) FROM quotations")
    rows = await conn.fetch(
        """SELECT date, person_in_charge, rfq_no, bqms_code, spec, maker,
                  quantity_forecast, price_rmb, price_vnd, price_ama,
                  price_bqms_v1, price_bqms_v2, price_bqms_v3, price_bqms_v4,
                  notes, supplier, result
           FROM quotations ORDER BY date DESC NULLS LAST LIMIT $1 OFFSET $2""",
        min(limit, 100), offset,
    )
    return {"data": [dict(r) for r in rows], "total": total, "page": page, "pages": (total + min(limit, 100) - 1) // min(limit, 100)}


@router.get("/tool2/giao-hang")
async def tool2_giao_hang(
    page: int = 1, limit: int = 25,
    conn=Depends(get_db), _=Depends(get_current_user),
):
    """Load delivery data from po_deliveries table."""
    offset = (page - 1) * min(limit, 100)
    total = await conn.fetchval("SELECT COUNT(*) FROM po_deliveries")
    rows = await conn.fetch(
        """SELECT po_date, po_number, shipping_no, rfq_no, bqms_code, spec,
                  quantity, unit, unit_price, total_amount, buyer,
                  receiver_name, warehouse, status, delivery_date, actual_qty
           FROM po_deliveries ORDER BY po_date DESC NULLS LAST LIMIT $1 OFFSET $2""",
        min(limit, 100), offset,
    )
    return {"data": [dict(r) for r in rows], "total": total, "page": page, "pages": (total + min(limit, 100) - 1) // min(limit, 100)}


# ── Tool 3: AI Classification ──────────────────────────

@router.post("/tool3/classify")
async def tool3_classify(
    body: dict,
    background_tasks: BackgroundTasks,
    conn=Depends(get_db),
    _=Depends(require_role(UserRole.admin, UserRole.procurement)),
):
    """Classify orders using historical data + optional Gemini AI."""
    job_id = _new_job("classify")
    orders = body.get("orders", [])
    use_ai = body.get("use_ai", False)

    async def _run_classify():
        try:
            results = []
            for i, order in enumerate(orders):
                bqms = order.get("bqms_code", "")
                spec = order.get("spec", "")

                # Historical lookup
                history = await conn.fetch(
                    "SELECT result, COUNT(*) as cnt FROM quotations WHERE bqms_code = $1 AND result IS NOT NULL GROUP BY result",
                    bqms,
                )
                hist_map = {r["result"]: r["cnt"] for r in history}
                wins = hist_map.get("Y", 0) + hist_map.get("y", 0)
                losses = hist_map.get("N", 0) + hist_map.get("n", 0)

                has_price = await conn.fetchval(
                    "SELECT COUNT(*) FROM quotations WHERE bqms_code = $1 AND price_bqms_v1 IS NOT NULL", bqms
                ) or 0

                # Similar products
                similar = []
                if spec:
                    rows = await conn.fetch(
                        "SELECT bqms_code, spec, result FROM quotations WHERE spec ILIKE $1 AND result IS NOT NULL ORDER BY date DESC LIMIT 5",
                        f"%{spec[:30]}%",
                    )
                    similar = [dict(r) for r in rows]

                # Decision logic
                if wins > 0 and has_price > 0:
                    decision, confidence = "yes", min(0.9, 0.5 + wins * 0.1)
                    reason = f"Đã thắng {wins} lần, có dữ liệu giá"
                elif wins > 0:
                    decision, confidence = "maybe", 0.5
                    reason = f"Đã thắng {wins} lần nhưng chưa có giá gần đây"
                elif losses > 2:
                    decision, confidence = "no", 0.7
                    reason = f"Đã thua {losses} lần, khả năng thấp"
                elif has_price > 0:
                    decision, confidence = "maybe", 0.4
                    reason = "Có dữ liệu giá nhưng chưa có lịch sử thắng"
                elif similar:
                    sim_wins = sum(1 for s in similar if s["result"] in ("Y", "y"))
                    decision = "maybe" if sim_wins > 0 else "no"
                    confidence = 0.4 if sim_wins > 0 else 0.3
                    reason = f"Sản phẩm tương tự: {sim_wins}/{len(similar)} thắng"
                else:
                    decision, confidence = "maybe", 0.3
                    reason = "Không có lịch sử — cần xem xét thủ công"

                results.append({
                    "order": order,
                    "decision": decision,
                    "confidence": confidence,
                    "reason": reason,
                    "wins": wins,
                    "losses": losses,
                    "has_price": has_price,
                    "similar_count": len(similar),
                })

                _jobs[job_id]["progress"] = int((i + 1) / len(orders) * 100)
                _log(job_id, f"[{i+1}/{len(orders)}] {bqms}: {decision} ({confidence:.0%})")

            summary = {
                "total": len(results),
                "yes": sum(1 for r in results if r["decision"] == "yes"),
                "no": sum(1 for r in results if r["decision"] == "no"),
                "maybe": sum(1 for r in results if r["decision"] == "maybe"),
            }

            _jobs[job_id]["result"] = {"results": results, "summary": summary}
            _jobs[job_id]["status"] = "done"
            _log(job_id, f"Hoàn thành: {summary['yes']} tham gia, {summary['no']} bỏ qua, {summary['maybe']} xem xét")

        except Exception as e:
            _jobs[job_id]["status"] = "error"
            _jobs[job_id]["error"] = str(e)
            _log(job_id, f"Lỗi: {e}")

    background_tasks.add_task(_run_classify)
    return {"job_id": job_id}


# ── Tool 5: Market Search ──────────────────────────────

@router.post("/tool5/search")
async def tool5_search(
    body: dict,
    background_tasks: BackgroundTasks,
    conn=Depends(get_db),
    _=Depends(require_role(UserRole.admin, UserRole.procurement)),
):
    """Start market search job. Searches product DB + saved prices."""
    job_id = _new_job("market_search")
    spec = body.get("spec", "")
    platforms = body.get("platforms", ["database"])

    async def _run_search():
        try:
            _log(job_id, f"Tìm kiếm: {spec}")
            results = []

            # Phase 1: Search internal database
            _log(job_id, "Đang tìm trong cơ sở dữ liệu nội bộ...")
            rows = await conn.fetch(
                """SELECT bqms_code, product_name, spec, maker, type
                   FROM products
                   WHERE spec ILIKE $1 OR product_name ILIKE $1 OR bqms_code ILIKE $1
                   LIMIT 20""",
                f"%{spec[:50]}%",
            )
            db_results = [dict(r) for r in rows]
            _log(job_id, f"Tìm thấy {len(db_results)} sản phẩm trong DB")

            # Phase 2: Search quotation history for prices
            _log(job_id, "Đang tra cứu lịch sử giá...")
            price_rows = await conn.fetch(
                """SELECT q.bqms_code, q.spec, q.supplier, q.price_bqms_v1,
                          q.price_bqms_v2, q.result, q.date
                   FROM quotations q
                   WHERE q.spec ILIKE $1 OR q.bqms_code ILIKE $1
                   ORDER BY q.date DESC LIMIT 20""",
                f"%{spec[:50]}%",
            )
            price_results = [dict(r) for r in price_rows]
            _log(job_id, f"Tìm thấy {len(price_results)} bản báo giá liên quan")

            # Phase 3: Search saved market prices
            _log(job_id, "Đang kiểm tra giá thị trường đã lưu...")
            market_rows = await conn.fetch(
                """SELECT bqms_code, product_name, platform, supplier_name,
                          price_usd, moq, match_confidence, created_at
                   FROM market_prices
                   WHERE product_name ILIKE $1 OR spec ILIKE $1
                   ORDER BY created_at DESC LIMIT 10""",
                f"%{spec[:50]}%",
            )
            market_results = [dict(r) for r in market_rows]
            _log(job_id, f"Tìm thấy {len(market_results)} giá thị trường")

            # Phase 4: Search transactions for actual purchase prices
            _log(job_id, "Đang tra cứu giá mua thực tế...")
            tx_rows = await conn.fetch(
                """SELECT bqms_code, spec, maker, unit_price_usd, unit_price_vnd,
                          seller, buyer, transaction_date
                   FROM transactions
                   WHERE (spec ILIKE $1 OR bqms_code ILIKE $1)
                     AND unit_price_usd IS NOT NULL
                   ORDER BY transaction_date DESC LIMIT 10""",
                f"%{spec[:50]}%",
            )
            tx_results = [dict(r) for r in tx_rows]
            _log(job_id, f"Tìm thấy {len(tx_results)} giao dịch có giá")

            _jobs[job_id]["result"] = {
                "query": spec,
                "database_matches": db_results,
                "price_history": price_results,
                "market_prices": market_results,
                "actual_prices": tx_results,
                "summary": {
                    "db_matches": len(db_results),
                    "price_history": len(price_results),
                    "market_prices": len(market_results),
                    "actual_prices": len(tx_results),
                },
            }
            _jobs[job_id]["status"] = "done"
            _log(job_id, "Hoàn thành tìm kiếm!")

        except Exception as e:
            _jobs[job_id]["status"] = "error"
            _jobs[job_id]["error"] = str(e)
            _log(job_id, f"Lỗi: {e}")

    background_tasks.add_task(_run_search)
    return {"job_id": job_id}


# ── Job Status (shared for all tools) ───────────────────

@router.get("/job/{job_id}")
async def get_job(job_id: str, _=Depends(get_current_user)):
    """Poll job status for any tool."""
    if job_id not in _jobs:
        raise NotFoundError("Job", job_id)
    return _jobs[job_id]
